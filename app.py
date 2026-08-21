from flask import Flask, render_template, request, redirect, flash, jsonify, send_from_directory, send_file
from pymongo import MongoClient
from dotenv import load_dotenv
import pandas as pd
import os
import math
from flask_cors import CORS
from bson import ObjectId
from werkzeug.utils import secure_filename
import time
from datetime import datetime
from img_to_text import extract_text_from_image
#from video_to_audio import extract_audio_from_video
from make_contact import create_contact
import tempfile
#import cv2
import os
import time
from flask import session
import random
import requests
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import calendar
import re
import json
import fitz  # PyMuPDF — pip install pymupdf
import secrets
# NEW: Cloudinary (used for Inventory / project image & video uploads)
import cloudinary
import cloudinary.uploader
import cloudinary.api
import cloudinary.utils
from supabase import create_client
import subprocess
import shutil
# NEW: branding overlay for images
from PIL import Image, ImageDraw, ImageFont

import wp
import threading
import socket

# Load env
load_dotenv()

# NEW: Cloudinary config - reads from .env
cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET"),
    secure=True
)

# NEW: Supabase Storage config - reads from .env
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
SUPABASE_PDF_BUCKET = "documents"  # must exist + be set Public in the Supabase dashboard

supabase = None
if SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY:
    supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
else:
    print("[startup] Supabase not configured — SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY missing in .env")



# NEW: Supabase Storage config #2 - used for photo/video media (separate
# project/quota from the PDF bucket above). Reads from .env.
SUPABASE_URL2 = os.getenv("SUPABASE_URL2")
SUPABASE_SERVICE_ROLE_KEY2 = os.getenv("SUPABASE_SERVICE_ROLE_KEY2")
SUPABASE_MEDIA_BUCKET = "media"  # must exist + be set Public in the Supabase dashboard
MAX_MEDIA_UPLOAD_BYTES = 50 * 1024 * 1024  # Supabase Storage free-tier hard cap


supabase2 = None
if SUPABASE_URL2 and SUPABASE_SERVICE_ROLE_KEY2:
    supabase2 = create_client(SUPABASE_URL2, SUPABASE_SERVICE_ROLE_KEY2)
else:
    print("[startup] Supabase (media) not configured — SUPABASE_URL2 / SUPABASE_SERVICE_ROLE_KEY2 missing in .env")


app = Flask(__name__)
app.permanent_session_lifetime = timedelta(days=60)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "supersecretkey")

CORS(app)


# Upload folder
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Mongo Config
MONGO_URI = os.getenv("MONGO_URI")
DB_NAME = os.getenv("DB_NAME")

# PERF: tuned connection pool + fail-fast timeouts instead of Mongo defaults.
# maxPoolSize raised a bit above default so concurrent requests aren't
# queued waiting for a free connection; timeouts prevent a slow/unreachable
# Mongo from hanging every request indefinitely.
client = MongoClient(
    MONGO_URI,
    maxPoolSize=100,
    minPoolSize=5,
    serverSelectionTimeoutMS=5000,
    connectTimeoutMS=5000,
    socketTimeoutMS=20000,
    retryWrites=True
)
db = client[DB_NAME]
dai_collection = db["DAI"]
# CHANGED: was db["project"] - now uses "projects" per the Inventory feature
projects_collection = db["projects"]

# -----------------------------------------------------------------
# DIAGNOSTIC: prints, on startup, which Mongo DB this process is
# actually connected to. Check `docker logs <container>` right after
# the container starts. If MONGO_URI/DB_NAME show as None/empty, or
# the host doesn't match your real DB, your .env is not being loaded
# inside the container (missing env_file / --env-file, or excluded by
# .dockerignore) and the app has silently fallen back to a default
# local Mongo connection instead of your intended database — that
# alone explains "works on localhost, admin missing on VPS": the VPS
# container may be pointed at a different Mongo than you think.
# -----------------------------------------------------------------
print(f"[startup] DB_NAME env var = {DB_NAME!r}")
print(f"[startup] MONGO_URI is set = {bool(MONGO_URI)}")
try:
    print(f"[startup] teamAssign doc count in this DB = {db['teamAssign'].count_documents({})}")
    print(f"[startup] roles present = {sorted(set(db['teamAssign'].distinct('roll')))}")
except Exception as _diag_err:
    print(f"[startup] Could NOT query teamAssign - Mongo connection problem: {_diag_err}")


# =====================================================================
# PERF: INDEXES
# Every field these APIs filter/sort on gets an index here. create_index()
# is idempotent — safe to call on every boot, it's a no-op if the index
# already exists. Without these, Mongo does a full COLLSCAN on every
# query below, which is the single biggest cost in this app as data grows.
# =====================================================================
def ensure_indexes():
    try:
        for coll_name in ["Leads", "RentalLeads", "sellingLeads", "agentLeads"]:
            coll = db[coll_name]
            coll.create_index("Phone Number")
            coll.create_index("Date")
            coll.create_index("Created At")
            coll.create_index("LeadType")
            coll.create_index("AssignToNumber")
            coll.create_index("DateObj")  # NEW indexed range-query field, see backfill_date_obj()

        db["endData"].create_index("Number")  # NOT unique — legacy data may have dupes; keep it that way
        db["endData"].create_index("Call Status")
        db["endData"].create_index("Interest Level")

        db["callLogs"].create_index("DateOnly")
        db["callLogs"].create_index("CreatedAt")
        db["callLogs"].create_index("Number")
        db["callLogs"].create_index("CalledBy")

        db["teamAssign"].create_index("Employee number")
        db["teamAssign"].create_index("roll")

        projects_collection.create_index([("ownerNumber", 1), ("status", 1)])
        projects_collection.create_index("kind")
        projects_collection.create_index("status")
        projects_collection.create_index("uniqueId")
        projects_collection.create_index("createdAt")

        db["requirements"].create_index("submittedByNumber")
        db["requirements"].create_index("broadcastTo")
        db["requirements"].create_index("status")
        db["requirements"].create_index("createdAt")

        db["DAI"].create_index("leadId")

        print("[startup] Indexes ensured.")
    except Exception as idx_err:
        print(f"[startup] Index creation warning (non-fatal): {idx_err}")


ensure_indexes()


# =====================================================================
# PERF: SIMPLE IN-MEMORY TTL CACHE
# For read-heavy, rarely-changing data (settings, etc). Not distributed —
# fine for a single-process deployment; if you scale to multiple workers
# behind a load balancer, swap this for Redis.
# =====================================================================
_cache_store = {}

def cached(key, ttl_seconds, compute_fn):
    now = time.time()
    hit = _cache_store.get(key)
    if hit and (now - hit[0]) < ttl_seconds:
        return hit[1]
    value = compute_fn()
    _cache_store[key] = (now, value)
    return value

def invalidate_cache(key):
    _cache_store.pop(key, None)


# Helper function
def serialize_doc(doc):
    doc["_id"] = str(doc["_id"])

    for key, value in doc.items():
        if isinstance(value, float) and math.isnan(value):
            doc[key] = None
        elif isinstance(value, datetime):
            doc[key] = format_ist(value)

    return doc

def format_ist(dt):
    """Formats a UTC datetime into IST 'hh:mm AM/PM . dd/mm/yyyy'"""
    if not isinstance(dt, datetime):
        return "-"
    ist = dt + timedelta(hours=5, minutes=30)
    return ist.strftime("%I:%M %p . %d/%m/%Y")

def get_collection_data(collection_name, projection=None):
    """PERF: optional projection param — pass only the fields you need to
    cut document transfer size for endpoints that don't need full docs."""
    collection = db[collection_name]
    if projection:
        data = list(collection.find({}, projection))
    else:
        data = list(collection.find())
    return [serialize_doc(doc) for doc in data]


@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET,POST,PUT,DELETE,OPTIONS"
    return response


# -----------------------------------------------------------------
# NEW: PARTNER ACCESS LOCK
# Partners (roll == "partner" in teamAssign) may ONLY reach the
# inventory dashboard, the projects API, uploaded media, and logout.
# Everything else (leads, team, exports, admin/emp dashboards, etc.)
# redirects them straight back to /inventory.
# -----------------------------------------------------------------
PARTNER_ALLOWED_PATHS = {"/inventory", "/logout", "/add-inventory"}
PARTNER_ALLOWED_PREFIXES = ("/api/projects", "/api/ai", "/uploads", "/static")

@app.before_request
def restrict_partner_access():
    if request.method == "OPTIONS":
        return  # let CORS preflight through untouched

    if session.get("role") == "partner":
        path = request.path
        if path in PARTNER_ALLOWED_PATHS:
            return
        if path.startswith(PARTNER_ALLOWED_PREFIXES):
            return
        return redirect("/inventory")


def parse_lead_date(date_str):
    """Leads store Date as DD-MM-YYYY. Defensive about other formats too."""
    if not date_str:
        return None
    date_str = str(date_str).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


# NEW: only show leads dated from July 2026 up to "now". Any lead whose
# Date field is missing or doesn't parse in any known format is dropped
# entirely (never shown as "undated") — same rule applies automatically
# to any new lead added going forward, as long as its Date field parses
# and isn't in the future.
JULY_2026_START = datetime(2026, 7, 1)

def filter_by_july_range(docs):
    """Legacy in-Python filter — kept for any code path still using it.
    Prefer get_filtered_leads() below, which pushes the range filter down
    into Mongo using the indexed DateObj field instead of scanning +
    parsing every document in Python."""
    now = datetime.utcnow()
    kept = []
    for d in docs:
        parsed = parse_lead_date(d.get("Date"))
        if not parsed:
            parsed = parse_created_at_str(d.get("Created At"))
        if not parsed:
            continue
        if JULY_2026_START <= parsed <= now:
            kept.append(d)
    return kept


def get_filtered_leads(collection_name):
    """
    PERF: fast path for the July-2026-onward lead lists.

    Docs that already have a DateObj field (real datetime, written on
    insert by /add-lead, indexed) are fetched with a single indexed Mongo
    range query — no Python date parsing needed.

    Docs from before the DateObj migration (missing the field entirely)
    fall back to the old Python-side parse-and-filter, but only for that
    shrinking legacy subset — not the whole collection. Run
    POST /api/admin/backfill-date-obj once to eliminate the fallback path
    completely; after that this function is a single indexed query.
    """
    now = datetime.utcnow()
    coll = db[collection_name]

    fast_docs = list(coll.find({
        "DateObj": {"$gte": JULY_2026_START, "$lte": now}
    }))

    legacy_docs = list(coll.find({"DateObj": {"$exists": False}}))
    kept_legacy = []
    for d in legacy_docs:
        parsed = parse_lead_date(d.get("Date")) or parse_created_at_str(d.get("Created At"))
        if parsed and JULY_2026_START <= parsed <= now:
            kept_legacy.append(d)

    return [serialize_doc(d) for d in (fast_docs + kept_legacy)]


def backfill_date_obj(collection_name):
    """ONE-TIME (per collection) migration: adds an indexed real-datetime
    DateObj field to every doc that doesn't have one yet, parsed from
    'Date' (falls back to 'Created At'). After running this once per
    collection, get_filtered_leads() runs as a single indexed query with
    no Python fallback needed."""
    coll = db[collection_name]
    docs = list(coll.find({"DateObj": {"$exists": False}}, {"Date": 1, "Created At": 1}))
    updated = 0
    for d in docs:
        parsed = parse_lead_date(d.get("Date")) or parse_created_at_str(d.get("Created At"))
        if parsed:
            coll.update_one({"_id": d["_id"]}, {"$set": {"DateObj": parsed}})
            updated += 1
    return {"scanned": len(docs), "updated": updated}


@app.route("/api/admin/backfill-date-obj", methods=["POST"])
def api_backfill_date_obj():
    """Admin-triggered one-time migration across all 4 lead collections.
    Safe to re-run — it only touches docs still missing DateObj."""
    if session.get("role") != "admin":
        return jsonify({"success": False, "message": "Admin only"}), 403
    try:
        results = {}
        for coll_name in ["Leads", "RentalLeads", "sellingLeads", "agentLeads"]:
            results[coll_name] = backfill_date_obj(coll_name)
        return jsonify({"success": True, "results": results}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


# NEW: ONE-TIME migration, run from the terminal (docker exec) instead of
# a browser route — no session/auth needed since it's only reachable by
# whoever has shell access to the container. Re-fits every EXISTING
# project/inventory banner into the 1080x1080 WhatsApp square (no
# cropping — same fit_to_whatsapp_square() logic new banners use).
# Safe to re-run — docs already fixed are flagged "squareFitted": True
# and get skipped on the next run.
def run_resize_banners_to_square():
    results = {"scanned": 0, "resized": 0, "skipped_no_banner": 0, "failed": []}

    docs = list(projects_collection.find({"squareFitted": {"$ne": True}}))
    results["scanned"] = len(docs)
    print(f"[resize-banners] {len(docs)} doc(s) to check...")

    for doc in docs:
        banner_url = doc.get("bannerUrl") or doc.get("img")
        if not banner_url:
            results["skipped_no_banner"] += 1
            continue

        try:
            resp = requests.get(banner_url, timeout=30)
            resp.raise_for_status()

            content_type = resp.headers.get("Content-Type", "")
            content_len = len(resp.content)
            print(f"[resize-banners] fetched {banner_url} — "
                  f"content-type={content_type!r} bytes={content_len}")

            if "video" in content_type:
                print(f"[resize-banners] SKIP (bannerUrl is a video, not an image): {doc.get('name', doc['_id'])}")
                results["failed"].append({
                    "id": str(doc["_id"]),
                    "error": f"bannerUrl points to a video ({content_type}), not an image"
                })
                continue

            if content_len < 100:
                print(f"[resize-banners] SKIP (empty/near-empty response, first 200 bytes: {resp.content[:200]!r})")
                results["failed"].append({
                    "id": str(doc["_id"]),
                    "error": f"downloaded {content_len} bytes — likely not a real image"
                })
                continue

            squared = fit_to_whatsapp_square(resp.content)

            folder = "projects" if doc.get("kind") == "project" else "inventory"
            new_url, new_object_path = upload_media_to_supabase(
                squared, folder=folder, resource_type="image", ext="jpg"
            )

            update_fields = {"bannerUrl": new_url, "squareFitted": True}

            # Keep img / mediaUrls[0] in sync if they pointed at the same banner
            if doc.get("img") == banner_url:
                update_fields["img"] = new_url
            media_urls = doc.get("mediaUrls") or []
            if media_urls and media_urls[0] == banner_url:
                media_urls[0] = new_url
                update_fields["mediaUrls"] = media_urls

            projects_collection.update_one({"_id": doc["_id"]}, {"$set": update_fields})
            # NOTE: the old (pre-square) banner object is intentionally left
            # in Supabase Storage rather than auto-deleted — we don't have
            # a reliably-tracked object path for bannerUrl on old docs, and
            # a wrong delete would be irreversible. Clean those up manually
            # from the Supabase dashboard if you want to reclaim space.
            results["resized"] += 1
            print(f"[resize-banners] OK: {doc.get('name', doc['_id'])}")

        except Exception as doc_err:
            results["failed"].append({"id": str(doc["_id"]), "error": str(doc_err)})
            print(f"[resize-banners] FAILED: {doc.get('name', doc['_id'])} — {doc_err}")
            continue

    invalidate_cache("inventory_dashboard_stats")
    print(f"[resize-banners] DONE — {results}")
    return results


def get_date_range(period):
    """Returns (start_date, end_date) for the given export period, or (None, None) for 'all'."""
    now = datetime.utcnow()

    if period == "this_month":
        start = datetime(now.year, now.month, 1)
        return start, now

    if period == "last_month":
        # last month + current month combined
        first_of_this_month = datetime(now.year, now.month, 1)
        last_month_end = first_of_this_month - timedelta(days=1)
        start = datetime(last_month_end.year, last_month_end.month, 1)
        return start, now

    if period == "last_3_months":
        # current month + previous 2 months (3 months total, inclusive)
        month = now.month - 2
        year = now.year
        while month <= 0:
            month += 12
            year -= 1
        start = datetime(year, month, 1)
        return start, now

    return None, None  # "all"


# =============================
# DASHBOARD PERIOD FILTER (Today / This Month / Last Month / Last 3 Months / Lifetime)
# Uses the "Created At" field written by /add-lead (format "%Y-%m-%d %H:%M:%S").
# =============================
def get_dashboard_period_range(period):
    """Returns (start_datetime, end_datetime) for the dashboard period filter,
    or (None, None) for 'lifetime' (meaning: no filtering, include everything —
    even leads that have no 'Created At' field at all)."""
    now = datetime.utcnow()

    if period == "today":
        start = datetime(now.year, now.month, now.day)
        return start, now

    if period == "this_month":
        start = datetime(now.year, now.month, 1)
        return start, now

    if period == "last_month":
        first_of_this_month = datetime(now.year, now.month, 1)
        last_month_end = first_of_this_month - timedelta(seconds=1)
        start = datetime(last_month_end.year, last_month_end.month, 1)
        return start, last_month_end

    if period == "last_3_months":
        month = now.month - 2
        year = now.year
        while month <= 0:
            month += 12
            year -= 1
        start = datetime(year, month, 1)
        return start, now

    return None, None  # "lifetime"


def parse_created_at_str(s):
    """Parses the 'Created At' field written by /add-lead: 'YYYY-MM-DD HH:MM:SS'."""
    if not s:
        return None
    try:
        return datetime.strptime(str(s).strip(), "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None

# =============================
# AI PROPERTY AUTOFILL (Mistral)
# =============================
MISTRAL_API_KEY = os.getenv("MISTRAL_API_KEY")
MISTRAL_API_URL = "https://api.mistral.ai/v1/chat/completions"

# Schema for the PARTNER "Add Inventory" form (Image 1)
INVENTORY_FIELDS_SCHEMA = {
    "listingBasis": "Individual property | Project",
    "dealType": "For Sale | For Rent",
    "propertyType": "Apartment | Villa | Plot | Independent House | Commercial | Farmhouse",
    "propertyTitle": "short catchy listing title",
    "locality": "locality / landmark",
    "configuration": "e.g. 2 BHK",
    "furnishing": "Unfurnished | Semi-furnished | Fully furnished",
    "areaUnit": "sqft",
    "carpetArea": "number only, as string",
    "superArea": "number only, as string",
    "floor": "e.g. 3 of 5",
    "bathrooms": "number only, as string",
    "facing": "East | West | North | South | North-East | North-West | South-East | South-West",
    "parking": "number of parking spots, as string",
    "possession": "Ready to move | Under construction",
    "price": "number only, no currency symbol, as string",
    "quickNotes": "short internal note, 1 sentence",
    "description": "polished 3-5 sentence marketing description"
}

# Schema for the ADMIN "Add Project" form (Image 2)
PROJECT_FIELDS_SCHEMA = {
    "name": "project name",
    "location": "location, e.g. Sidon, Himachal Pradesh",
    "propertyType": "Apartment | Villa | Plot | Independent House | Commercial",
    "possession": "New launch | Ready to move | Under construction",
    "configuration": "e.g. Studio / 1 BHK / 2 BHK",
    "startingPrice": "e.g. 70 Lakhs onwards",
    "description": "3-5 sentence project note / marketing description"
}


def extract_text_from_pdf(pdf_path):
    """
    Text-extracts a PDF. Tries the fast direct-text path first (works for
    brochures/typed PDFs). Any page with no extractable text is assumed
    scanned/image-only, so it's rasterized and sent through the existing
    extract_text_from_image() OCR helper instead.
    """
    text_chunks = []
    try:
        doc = fitz.open(pdf_path)
        for page in doc:
            page_text = page.get_text().strip()
            if page_text:
                text_chunks.append(page_text)
                continue

            # Scanned page — rasterize and OCR it
            pix = page.get_pixmap(dpi=200)
            tmp_img = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            tmp_img.close()
            pix.save(tmp_img.name)
            try:
                ocr_text = extract_text_from_image(tmp_img.name)
                if ocr_text:
                    text_chunks.append(ocr_text)
            except Exception as ocr_err:
                print(f"[pdf-extract] OCR fallback failed: {ocr_err}")
            finally:
                os.remove(tmp_img.name)
        doc.close()
    except Exception as e:
        print(f"[pdf-extract] failed: {e}")
    return "\n".join(text_chunks)


def call_mistral_generate(extracted_text, schema):
    """Sends extracted OCR/PDF text to Mistral and asks it to fill the given field schema, returning parsed JSON."""
    if not MISTRAL_API_KEY:
        raise RuntimeError("MISTRAL_API_KEY not configured in .env")

    system_prompt = (
        "You are a real-estate data-entry assistant. Given raw text extracted "
        "from property photos and/or a brochure PDF (may be messy OCR output), "
        "fill in the JSON fields described below as best you can infer. "
        "Return ONLY a valid JSON object — no markdown fences, no commentary. "
        "If a field cannot be determined from the text, return an empty string for it. "
        f"Fields and guidance: {json.dumps(schema)}"
    )

    payload = {
        "model": "mistral-large-latest",
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": extracted_text[:12000] or "No text could be extracted."}
        ],
        "temperature": 0.3,
        "response_format": {"type": "json_object"}
    }

    resp = requests.post(
        MISTRAL_API_URL,
        headers={
            "Authorization": f"Bearer {MISTRAL_API_KEY}",
            "Content-Type": "application/json"
        },
        json=payload,
        timeout=60
    )
    resp.raise_for_status()
    content = resp.json()["choices"][0]["message"]["content"]
    return json.loads(content)

# NEW: separate Mistral key used ONLY for lead-intent classification (Hot/Warm/Cold),
# kept apart from MISTRAL_API_KEY (used for AI autofill) so usage/quota don't mix.
MISTRAL_API_KEY2 = os.getenv("MISTRAL_API_KEY2")


def classify_lead_intent(lead_snapshot: dict, call_log: dict):
    """
    Uses MISTRAL_API_KEY2 to classify a lead's buying intent as Hot / Warm / Cold,
    based on the lead's stored details plus the call log just submitted.
    Returns {"intent": "Hot"|"Warm"|"Cold", "reason": "..."} or None on failure.
    """
    if not MISTRAL_API_KEY2:
        print("[intent] MISTRAL_API_KEY2 not configured — skipping intent classification")
        return None

    context = {
        "lead": {
            "name": lead_snapshot.get("name"),
            "location": lead_snapshot.get("location"),
            "property": lead_snapshot.get("property"),
            "budget": lead_snapshot.get("budget"),
            "timeline": lead_snapshot.get("timeline"),
            "note": lead_snapshot.get("note"),
        },
        "latest_call_log": {
            "callStatus": call_log.get("callStatus"),
            "customerResponse": call_log.get("customerResponse"),
            "interestLevel": call_log.get("interestLevel"),
            "objection": call_log.get("objection"),
            "followupTimeline": call_log.get("followupTimeline"),
            "callPriority": call_log.get("callPriority"),
            "callerRemarks": call_log.get("callerRemarks"),
        }
    }

    system_prompt = (
        "You are a real-estate CRM assistant. Given a lead's profile and the "
        "latest call log recorded by a sales agent, classify the lead's buying "
        "intent as exactly one of: Hot, Warm, Cold. "
        "Hot = ready to move soon, strong interest, budget matches, no major objection. "
        "Warm = interested but hesitant, budget mismatch, or comparing options. "
        "Cold = not interested, wrong number, unreachable repeatedly, or explicitly declined. "
        "Return ONLY a valid JSON object with keys 'intent' and 'reason' "
        "(reason: max one short sentence). No markdown, no commentary."
    )

    payload = {
        "model": "mistral-large-latest",
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": json.dumps(context)}
        ],
        "temperature": 0.2,
        "response_format": {"type": "json_object"}
    }

    try:
        resp = requests.post(
            MISTRAL_API_URL,
            headers={
                "Authorization": f"Bearer {MISTRAL_API_KEY2}",
                "Content-Type": "application/json"
            },
            json=payload,
            timeout=30
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        result = json.loads(content)
        intent = str(result.get("intent", "")).strip().capitalize()
        if intent not in ("Hot", "Warm", "Cold"):
            intent = "Warm"
        return {"intent": intent, "reason": result.get("reason", "")}
    except Exception as e:
        print(f"[intent] classification failed: {e}")
        return None

# =============================
# NEW: BRANDING OVERLAY (images only)
# =============================
# Replace these near the top of app.py
FONT_DIR = os.path.join(os.path.dirname(__file__), "fonts")
FONT_BOLD = os.path.join(FONT_DIR, "DejaVuSans-Bold.ttf")
FONT_REG  = os.path.join(FONT_DIR, "DejaVuSans.ttf")
BRAND_ORANGE = (237, 128, 73)
BRAND_NAVY_SOLID = (16, 19, 28)   # solid, not the old semi-transparent tuple


BRAND_CONTACT_NUMBER = "91 73035 15710"   # <-- change this ONE line if the number is different
BRAND_WEBSITE = "nishahomes.com"           # <-- change this ONE line if the website changes
BRAND_NAME = "NISHA HOMES"

# NEW: WhatsApp chat-bubble preview is a 1:1 square. Every banner (branded
# or not) gets fit into this square — WITHOUT cropping — by scaling it
# down to fit and padding the empty space, never by cutting the photo.
WHATSAPP_SQUARE_SIZE = 1080

def generate_unique_id():
    """Short, URL-safe id used for the public /view/<id> page."""
    return secrets.token_urlsafe(8).replace("_", "").replace("-", "")[:10]

_SYSTEM_FONT_FALLBACKS = {
    "bold": [
        FONT_BOLD,
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    ],
    "regular": [
        FONT_REG,
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ],
}


def upload_pdf_to_supabase(pdf_file, folder):
    """
    Uploads a PDF file (Werkzeug FileStorage from request.files) to Supabase
    Storage and returns its public URL. `folder` groups files, e.g.
    "inventory" or "projects" — purely cosmetic path organization.
    Raises RuntimeError if Supabase isn't configured.
    """
    if not supabase:
        raise RuntimeError("Supabase not configured — check SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY in .env")

    file_bytes = pdf_file.read()
    object_path = f"{folder}/{secrets.token_hex(8)}.pdf"

    supabase.storage.from_(SUPABASE_PDF_BUCKET).upload(
        object_path,
        file_bytes,
        {"content-type": "application/pdf"}
    )

    return supabase.storage.from_(SUPABASE_PDF_BUCKET).get_public_url(object_path)


_IMAGE_CONTENT_TYPES = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "webp": "image/webp", "gif": "image/gif"
}
_VIDEO_CONTENT_TYPES = {
    "mp4": "video/mp4", "mov": "video/quicktime", "avi": "video/x-msvideo",
    "webm": "video/webm", "mkv": "video/x-matroska"
}


def _to_bytes(data):
    """Some of our image/video builders return io.BytesIO, others return
    raw bytes. Normalize to bytes before handing off to Supabase."""
    if hasattr(data, "read"):
        return data.read()
    return data

def upload_media_to_supabase(file_data, folder, resource_type="image", ext="jpg", max_retries=2):
    if not supabase2:
        raise RuntimeError("Supabase media storage not configured — check SUPABASE_URL2 / SUPABASE_SERVICE_ROLE_KEY2 in .env")

    ext = (ext or "jpg").lstrip(".").lower()
    file_bytes = _to_bytes(file_data)

    if len(file_bytes) > MAX_MEDIA_UPLOAD_BYTES:
        raise ValueError(f"File too large for Supabase Storage ({len(file_bytes)/1024/1024:.1f}MB, limit is 50MB)")

    content_type = (
        _VIDEO_CONTENT_TYPES.get(ext, "video/mp4") if resource_type == "video"
        else _IMAGE_CONTENT_TYPES.get(ext, "image/jpeg")
    )

    object_path = f"{folder}/{secrets.token_hex(8)}.{ext}"

    last_err = None
    for attempt in range(1, max_retries + 2):
        try:
            supabase2.storage.from_(SUPABASE_MEDIA_BUCKET).upload(
                object_path, file_bytes, {"content-type": content_type}
            )
            return supabase2.storage.from_(SUPABASE_MEDIA_BUCKET).get_public_url(object_path), object_path
        except Exception as e:
            last_err = e
            print(f"[media] upload attempt {attempt} failed for {object_path}: {e}")
            time.sleep(1.5 * attempt)

    raise last_err



def upload_raw_then_brand(raw_bytes, folder, resource_type, ext, brand_fn=None):
    """
    Uploads the ORIGINAL file to Supabase first. Only if brand_fn is given
    (branding was requested), builds the branded version from the same
    in-memory bytes, uploads it as a separate object, deletes the raw
    object, and returns the branded URL/path instead.

    If branding fails, the raw upload is left in place and used as the
    result — a branding crash never costs the upload itself.

    Returns: (public_url, object_path, was_branded: bool)
    """
    raw_url, raw_path = upload_media_to_supabase(
        raw_bytes, folder=folder, resource_type=resource_type, ext=ext
    )

    if not brand_fn:
        return raw_url, raw_path, False

    try:
        branded_bytes = _to_bytes(brand_fn(raw_bytes))
        branded_url, branded_path = upload_media_to_supabase(
            branded_bytes, folder=folder, resource_type=resource_type, ext=ext
        )
    except Exception as brand_err:
        print(f"[branding] failed for {raw_path}, keeping unbranded upload: {brand_err}")
        return raw_url, raw_path, False

    delete_media_from_supabase(raw_path)
    return branded_url, branded_path, True


def delete_media_from_supabase(object_path):
    """Best-effort delete — never raises, matching how the old Cloudinary
    destroy() calls were wrapped in try/except so a failed cleanup never
    breaks the main request."""
    if not supabase2 or not object_path:
        return
    try:
        supabase2.storage.from_(SUPABASE_MEDIA_BUCKET).remove([object_path])
    except Exception as e:
        print("[media] Supabase delete failed:", e)

def _font(weight, size):
    """Robust font loader. Tries your bundled fonts, then common Linux
    system fonts, and only as an absolute last resort falls back to PIL's
    built-in font — kept at a REQUESTED size instead of the old tiny
    default, which is what made your text invisible."""
    for path in _SYSTEM_FONT_FALLBACKS.get(weight, []):
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            continue
    try:
        return ImageFont.load_default(size=size)   # Pillow >= 10.1
    except TypeError:
        return ImageFont.load_default()


def _format_price_display(price_str):
    """Formats price for the big headline. Pure numbers get comma
    grouping + ₹; free-text prices (e.g. '85 Lakhs') are shown as-is
    with a ₹ prefix if missing."""
    if not price_str:
        return ""
    s = str(price_str).strip()
    cleaned = re.sub(r"[^\d]", "", s)
    if cleaned and cleaned == s.replace(",", ""):
        return f"₹{int(cleaned):,}"
    return s if s.startswith("₹") else f"₹{s}"


def _indian_price_words(price_str):
    """Pure numeric prices get a '2 Crore 12 Lakh' style breakdown line,
    like your reference image. Free-text prices return None (line skipped)."""
    if not price_str:
        return None
    cleaned = re.sub(r"[^\d]", "", str(price_str))
    if not cleaned or not cleaned.isdigit() or int(cleaned) < 100000:
        return None
    n = int(cleaned)
    crore, rem = divmod(n, 10000000)
    lakh, _ = divmod(rem, 100000)
    parts = []
    if crore: parts.append(f"{crore} Crore")
    if lakh: parts.append(f"{lakh} Lakh")
    return " ".join(parts) or None


def build_banner_image(image_bytes, fields):
    """
    Returns branded JPEG bytes styled after the reference design:
    orange top bar -> the ORIGINAL, UNCROPPED photo -> a solid navy info
    panel appended below the photo (title / location / price / config) ->
    divider -> contact bar.

    Title/locality/config lines are wrapped and truncated using REAL
    pixel measurement (draw.textlength), not an estimated "chars per
    line" — the estimate was undersizing for bold/caps text and letting
    lines run off the right edge. panel_h is now measured from the
    actual content instead of a fixed W*0.40, so nothing gets clipped
    or painted over by the contact bar below it.
    """
    CONTACT_NUMBER = "+91 73035 15710"   # fixed branding contact number

    img = Image.open(io.BytesIO(image_bytes)).convert("RGB")

    TARGET_W = 1080
    if img.width != TARGET_W:
        new_h = int(img.height * (TARGET_W / img.width))
        img = img.resize((TARGET_W, new_h), Image.LANCZOS)
    W, H = img.size

    top_bar_h = int(W * 0.095)
    contact_h = int(W * 0.11)
    pad_x = int(W * 0.045)
    panel_pad_top = int(W * 0.09)
    panel_pad_bottom = int(W * 0.035)
    max_text_w = W - (pad_x * 2)

    f_title = _font("bold", int(W * 0.052))
    f_meta  = _font("regular", int(W * 0.030))
    f_price = _font("bold", int(W * 0.062))
    f_words = _font("regular", int(W * 0.026))

    dummy = ImageDraw.Draw(Image.new("RGB", (1, 1)))  # for pre-measuring text before the canvas exists

    def _ellipsize(text, font, max_w):
        if not text:
            return text
        if dummy.textlength(text, font=font) <= max_w:
            return text
        while text and dummy.textlength(text + "…", font=font) > max_w:
            text = text[:-1]
        return text + "…"

    def _wrap_title(text, font, max_w, max_lines):
        """Greedy word-wrap measured in real pixels, capped at max_lines,
        ellipsizing the last line if content still remains."""
        words = (text or "").split()
        if not words:
            return [""]
        lines, current, i = [], "", 0
        while i < len(words):
            trial = (current + " " + words[i]).strip()
            if dummy.textlength(trial, font=font) <= max_w or not current:
                current = trial
                i += 1
            else:
                lines.append(current)
                current = ""
                if len(lines) == max_lines:
                    break
        if current and len(lines) < max_lines:
            lines.append(current)
            i = len(words)
        if i < len(words) and lines:
            last = lines[-1]
            while last and dummy.textlength(last + "…", font=font) > max_w:
                last = last[:-1]
            lines[-1] = last + "…"
        return lines

    # ---- Pre-compute every line of panel text BEFORE we know panel_h ----
    title_lines = _wrap_title(fields.get("propertyTitle") or "", f_title, max_text_w, 2)
    locality_line = _ellipsize("📍 " + (fields.get("locality") or ""), f_meta, max_text_w)
    price_line = _format_price_display(fields.get("price"))
    words_line = _indian_price_words(fields.get("price"))
    sub_line = _ellipsize(" · ".join(filter(None, [
        fields.get("configuration"),
        fields.get("superArea") and f'{fields["superArea"]} sq.ft'
    ])), f_meta, max_text_w)

    # ---- Measure the exact height the content needs, size panel_h to fit ----
    content_h = len(title_lines) * int(f_title.size * 1.22)
    content_h += int(W * 0.02)
    content_h += int(f_meta.size * 1.6)
    content_h += int(f_price.size * 1.05)
    if words_line:
        content_h += int(f_words.size * 1.5)
    else:
        content_h += int(W * 0.01)
    if sub_line:
        content_h += int(f_meta.size * 1.2)

    panel_h = max(int(W * 0.28), content_h + panel_pad_top + panel_pad_bottom)
    total_H = top_bar_h + H + panel_h + contact_h

    canvas = Image.new("RGB", (W, total_H), BRAND_NAVY_SOLID)
    canvas.paste(img, (0, top_bar_h))
    draw = ImageDraw.Draw(canvas, "RGBA")

    # ---------- TOP BAR ----------
    draw.rectangle([0, 0, W, top_bar_h], fill=BRAND_ORANGE)
    f_brand = _font("bold", int(top_bar_h * 0.42))
    f_tag   = _font("regular", int(top_bar_h * 0.24))
    draw.text((int(W * 0.035), top_bar_h * 0.28), "NISHA HOMES", font=f_brand, fill="white")
    tag = "Trusted Real Estate Advisor"
    tag_w = draw.textlength(tag, font=f_tag)
    draw.text((W - tag_w - int(W * 0.035), top_bar_h * 0.40), tag, font=f_tag, fill="white")

    # ---------- DEAL-TYPE PILL ----------
    deal = (fields.get("dealType") or "For Sale").upper()
    f_pill = _font("bold", int(W * 0.032))
    pill_pad_x = int(W * 0.03)
    pill_h = int(W * 0.06)
    pill_w = draw.textlength(deal, font=f_pill) + pill_pad_x * 2
    px, py = int(W * 0.035), top_bar_h + int(W * 0.03)
    draw.rounded_rectangle([px, py, px + pill_w, py + pill_h], radius=pill_h // 2, fill=(255, 255, 255, 245))
    draw.text((px + pill_pad_x, py + pill_h * 0.20), deal, font=f_pill, fill=BRAND_ORANGE)

    # ---------- INFO PANEL (height = panel_h measured above) ----------
    panel_top = top_bar_h + H
    draw.rectangle([0, panel_top, W, panel_top + panel_h], fill=BRAND_NAVY_SOLID)

    y = panel_top + panel_pad_top
    for line in title_lines:
        draw.text((pad_x, y), line, font=f_title, fill="white")
        y += int(f_title.size * 1.22)

    y += int(W * 0.02)
    draw.text((pad_x, y), locality_line, font=f_meta, fill=(210, 214, 224))
    y += int(f_meta.size * 1.6)

    draw.text((pad_x, y), price_line, font=f_price, fill=BRAND_ORANGE)
    y += int(f_price.size * 1.05)

    if words_line:
        draw.text((pad_x, y), f"({words_line})", font=f_words, fill=(180, 186, 198))
        y += int(f_words.size * 1.5)
    else:
        y += int(W * 0.01)

    if sub_line:
        draw.text((pad_x, y), sub_line, font=f_meta, fill=(210, 214, 224))

    # ---------- DIVIDER + CONTACT BAR ----------
    divider_y = panel_top + panel_h
    draw.line([(pad_x, divider_y - 1), (W - pad_x, divider_y - 1)], fill=(255, 255, 255, 40), width=2)
    draw.rectangle([0, divider_y, W, divider_y + contact_h], fill=BRAND_NAVY_SOLID)

    icon_r = int(contact_h * 0.30)
    icon_cx, icon_cy = pad_x + icon_r, divider_y + contact_h // 2
    draw.ellipse([icon_cx - icon_r, icon_cy - icon_r, icon_cx + icon_r, icon_cy + icon_r], fill=BRAND_ORANGE)
    f_icon = _font("bold", int(icon_r * 1.1))
    icon_glyph = "\u260E"
    icon_glyph_w = draw.textlength(icon_glyph, font=f_icon)
    draw.text((icon_cx - icon_glyph_w / 2, icon_cy - f_icon.size * 0.62), icon_glyph, font=f_icon, fill="white")

    f_contact = _font("bold", int(W * 0.034))
    contact_display = f"{CONTACT_NUMBER}  •  {BRAND_WEBSITE}"
    draw.text((icon_cx + icon_r * 1.7, icon_cy - f_contact.size * 0.55), contact_display, font=f_contact, fill="white")

    out = io.BytesIO()
    canvas.save(out, format="JPEG", quality=92)
    out.seek(0)
    return out


def build_simple_branded_image(image_bytes):
    """
    Light branding for every photo EXCEPT the banner: small corner badges
    (logo top-left, contact+website bottom-left) instead of full-width
    strips — nothing gets cropped or covers the photo. Canvas size is
    NEVER changed.
    """
    img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    W, H = img.size
    draw = ImageDraw.Draw(img, "RGBA")

    margin = 10   # distance of badge from the edges (top/left)
    pad = 10      # padding inside each badge around the text

    # ---- TOP-LEFT: brand name badge ----
    f_logo = _font("bold", max(int(W * 0.022), 13))
    logo_w = draw.textlength(BRAND_NAME, font=f_logo)
    logo_h = f_logo.size

    bx, by = margin, margin
    badge_w = logo_w + pad * 2
    badge_h = logo_h + pad * 2
    draw.rounded_rectangle(
        [bx, by, bx + badge_w, by + badge_h],
        radius=8, fill=(*BRAND_ORANGE, 235)
    )
    draw.text((bx + pad, by + pad - 2), BRAND_NAME, font=f_logo, fill="white")

    # ---- BOTTOM-LEFT: contact + website badge ----
    f_contact = _font("regular", max(int(W * 0.017), 11))
    contact_text = f"{BRAND_CONTACT_NUMBER}  |  {BRAND_WEBSITE}"
    contact_w = draw.textlength(contact_text, font=f_contact)
    contact_h = f_contact.size

    cbadge_w = contact_w + pad * 2
    cbadge_h = contact_h + pad * 2
    cx = margin
    cy = H - margin - cbadge_h
    draw.rounded_rectangle(
        [cx, cy, cx + cbadge_w, cy + cbadge_h],
        radius=8, fill=(*BRAND_NAVY_SOLID, 210)
    )
    draw.text((cx + pad, cy + pad - 2), contact_text, font=f_contact, fill="white")

    out = io.BytesIO()
    img.save(out, format="JPEG", quality=92)
    out.seek(0)
    return out


def fit_to_whatsapp_square(image_bytes, size=WHATSAPP_SQUARE_SIZE, bg_color=BRAND_NAVY_SOLID):
    """
    Fits ANY image (a full branded banner OR a plain unbranded photo) into
    a size x size (default 1080x1080) square for WhatsApp's chat-bubble
    preview — WITHOUT cropping anything. The image is scaled down (never
    upscaled) to fit entirely inside the square, keeping its aspect ratio,
    then centered on a solid background so nothing gets cut off, just
    letterboxed top/bottom or left/right as needed.
    """
    img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    w, h = img.size

    scale = min(size / w, size / h)
    scale = min(scale, 1.0)  # never blow up a smaller image, just pad it
    new_w, new_h = max(1, round(w * scale)), max(1, round(h * scale))
    if (new_w, new_h) != (w, h):
        img = img.resize((new_w, new_h), Image.LANCZOS)

    canvas = Image.new("RGB", (size, size), bg_color)
    paste_x = (size - new_w) // 2
    paste_y = (size - new_h) // 2
    canvas.paste(img, (paste_x, paste_y))

    out = io.BytesIO()
    canvas.save(out, format="JPEG", quality=92)
    out.seek(0)
    return out


def _resolve_video_font_path():
    """Same fallback chain used elsewhere, but actually checked for video branding."""
    for path in _SYSTEM_FONT_FALLBACKS["bold"]:
        if os.path.exists(path):
            return path
    return None


def _ffmpeg_escape_text(text):
    """Escapes characters that break ffmpeg's drawtext filter syntax."""
    return (text or "").replace("\\", "\\\\").replace(":", "\\:").replace("'", "\u2019").replace(",", "\\,")


def build_simple_branded_video(video_bytes):
    """
    Light branding burned into a video via ffmpeg. Frame size unchanged.

    Margins and font sizes are computed as a PERCENTAGE of the video's own
    width/height (read via ffprobe) instead of fixed pixel offsets — so a
    portrait 720p clip and a landscape 4K clip both get proportionally
    correct, non-clipped branding instead of one fixed pixel value that
    only looks right on one resolution.

    - Top-left badge (brand name): nudged in from the corner, not flush
      against the edge.
    - Bottom-left badge (name + number): kept close to the left edge, but
      with extra bottom clearance so descenders/text are never cut off.
    - Font sizes are ~20% smaller than the previous version.
    """
    if not shutil.which("ffmpeg") or not shutil.which("ffprobe"):
        raise RuntimeError("ffmpeg/ffprobe not installed on this server — cannot brand video")

    font_path = _resolve_video_font_path()
    if not font_path:
        raise RuntimeError(
            "No usable font file found on server for video branding "
            "(checked bundled fonts + DejaVu/Liberation system paths). "
            "Install fonts, e.g. `apt-get install -y fonts-dejavu-core`."
        )

    tmp_in = tempfile.NamedTemporaryFile(delete=False, suffix=".mp4")
    tmp_in.write(video_bytes); tmp_in.close()
    tmp_out = tempfile.NamedTemporaryFile(delete=False, suffix=".mp4")
    tmp_out.close()

    try:
        # Read the video's REAL dimensions so every margin below scales
        # with its actual resolution/aspect ratio.
        src_w, src_h = _get_video_dimensions(tmp_in.name)

        # ---- Top-left brand-name badge: pushed in a bit from the corner ----
        top_margin_x = max(14, int(src_w * 0.045))
        top_margin_y = max(10, int(src_h * 0.03))

        # ---- Bottom-left contact badge: close to the edge, but with real
        # breathing room at the bottom so text never gets clipped ----
        bottom_margin_x = max(10, int(src_w * 0.025))
        bottom_margin_y = max(24, int(src_h * 0.065))

        # Font sizes: ~20% smaller than the previous 0.026 / 0.020 of height
        name_font_expr = f"main_h*{0.026 * 0.8:.4f}"
        contact_font_expr = f"main_h*{0.020 * 0.8:.4f}"

        name_esc = _ffmpeg_escape_text(BRAND_NAME)
        contact_esc = _ffmpeg_escape_text(f"{BRAND_CONTACT_NUMBER} | {BRAND_WEBSITE}")

        # box=1 + boxborderw draws a background box sized to the text itself
        # (plus the given border/padding), instead of a full-width bar — so
        # it never covers the whole frame and can't get cropped on any
        # screen size. The bottom badge's y is measured from its own text
        # height (th) plus bottom_margin_y, so it stays fully on-screen
        # regardless of resolution.
        vf = (
            f"drawtext=fontfile='{font_path}':text='{name_esc}':"
            f"x={top_margin_x}:y={top_margin_y}:fontsize={name_font_expr}:fontcolor=white:"
            f"box=1:boxcolor=0xED8049@0.9:boxborderw=8,"
            f"drawtext=fontfile='{font_path}':text='{contact_esc}':"
            f"x={bottom_margin_x}:y=h-th-{bottom_margin_y}:fontsize={contact_font_expr}:fontcolor=white:"
            f"box=1:boxcolor=0x10131C@0.85:boxborderw=8"
        )

        cmd = ["ffmpeg", "-y", "-i", tmp_in.name, "-vf", vf,
               "-c:v", "libx264", "-preset", "veryfast", "-crf", "23",
               "-c:a", "aac", "-b:a", "128k", "-movflags", "+faststart", tmp_out.name]
        subprocess.run(cmd, check=True, capture_output=True)
        with open(tmp_out.name, "rb") as f:
            video_bytes_out = f.read()
        print(f"[branding] video branding SUCCEEDED — {len(video_bytes_out)} bytes ({src_w}x{src_h})")
        return io.BytesIO(video_bytes_out)   # <-- was: return f.read() (raw bytes broke cloudinary upload)
    except subprocess.CalledProcessError as e:
        stderr = e.stderr.decode(errors="ignore") if e.stderr else str(e)
        print(f"[branding] ffmpeg video branding FAILED — stderr:\n{stderr}")
        raise
    finally:
        for p in (tmp_in.name, tmp_out.name):
            try: os.remove(p)
            except Exception: pass


def _get_video_dimensions(path):
    """Uses ffprobe (bundled with ffmpeg) to get (width, height) of a video."""
    cmd = [
        "ffprobe", "-v", "error", "-select_streams", "v:0",
        "-show_entries", "stream=width,height",
        "-of", "csv=s=x:p=0", path
    ]
    out = subprocess.check_output(cmd).decode().strip()
    w, h = out.split("x")
    return int(w), int(h)


def _build_brand_bars(fields, video_w):
    """
    Builds the same top-bar / info-panel / contact-bar graphics used for
    branded images, but as two standalone PNGs sized to `video_w` — these
    get composited onto video frames with ffmpeg instead of pasted with PIL.
    Returns: (top_bar_png_bytes, top_bar_h, bottom_bar_png_bytes, bottom_bar_h)
    """
    import textwrap
    CONTACT_NUMBER = "+91 73035 15710"
    W = video_w

    top_bar_h = int(W * 0.095)
    panel_h   = int(W * 0.40)
    contact_h = int(W * 0.11)
    bottom_h  = panel_h + contact_h

    # ---- TOP BAR ----
    top_img = Image.new("RGB", (W, top_bar_h), BRAND_ORANGE)
    draw = ImageDraw.Draw(top_img, "RGBA")
    f_brand = _font("bold", int(top_bar_h * 0.42))
    f_tag   = _font("regular", int(top_bar_h * 0.24))
    draw.text((int(W * 0.035), top_bar_h * 0.28), "NISHA HOMES", font=f_brand, fill="white")
    tag = "Trusted Real Estate Advisor"
    tag_w = draw.textlength(tag, font=f_tag)
    draw.text((W - tag_w - int(W * 0.035), top_bar_h * 0.40), tag, font=f_tag, fill="white")

    deal = (fields.get("dealType") or "For Sale").upper()
    f_pill = _font("bold", int(W * 0.032))
    pill_pad_x = int(W * 0.03)
    pill_h = int(W * 0.06)
    pill_w = draw.textlength(deal, font=f_pill) + pill_pad_x * 2
    px = int(W * 0.035)
    py = max(0, top_bar_h - pill_h - int(W * 0.015))
    draw.rounded_rectangle([px, py, px + pill_w, py + pill_h], radius=pill_h // 2, fill=(255, 255, 255, 245))
    draw.text((px + pill_pad_x, py + pill_h * 0.20), deal, font=f_pill, fill=BRAND_ORANGE)

    # ---- BOTTOM: INFO PANEL + CONTACT BAR ----
    bottom_img = Image.new("RGB", (W, bottom_h), BRAND_NAVY_SOLID)
    draw2 = ImageDraw.Draw(bottom_img, "RGBA")

    pad_x = int(W * 0.045)
    y = int(panel_h * 0.10)

    f_title = _font("bold", int(W * 0.052))
    title = fields.get("propertyTitle") or ""
    max_chars = max(10, int(W / (f_title.size * 0.52)))
    for line in textwrap.wrap(title, width=max_chars)[:2]:
        draw2.text((pad_x, y), line, font=f_title, fill="white")
        y += int(f_title.size * 1.22)

    y += int(panel_h * 0.025)
    f_meta = _font("regular", int(W * 0.030))
    draw2.text((pad_x, y), "📍 " + (fields.get("locality") or ""), font=f_meta, fill=(210, 214, 224))
    y += int(f_meta.size * 1.6)

    f_price = _font("bold", int(W * 0.062))
    draw2.text((pad_x, y), _format_price_display(fields.get("price")), font=f_price, fill=BRAND_ORANGE)
    y += int(f_price.size * 1.05)

    words = _indian_price_words(fields.get("price"))
    if words:
        f_words = _font("regular", int(W * 0.026))
        draw2.text((pad_x, y), f"({words})", font=f_words, fill=(180, 186, 198))
        y += int(f_words.size * 1.5)
    else:
        y += int(W * 0.01)

    sub = " · ".join(filter(None, [
        fields.get("configuration"),
        fields.get("superArea") and f'{fields["superArea"]} sq.ft'
    ]))
    if sub:
        draw2.text((pad_x, y), sub, font=f_meta, fill=(210, 214, 224))

    divider_y = panel_h
    draw2.line([(pad_x, divider_y - 1), (W - pad_x, divider_y - 1)], fill=(255, 255, 255, 40), width=2)

    icon_r = int(contact_h * 0.30)
    icon_cx, icon_cy = pad_x + icon_r, divider_y + contact_h // 2
    draw2.ellipse([icon_cx - icon_r, icon_cy - icon_r, icon_cx + icon_r, icon_cy + icon_r], fill=BRAND_ORANGE)
    f_icon = _font("bold", int(icon_r * 1.1))
    icon_glyph = "\u260E"
    iw = draw2.textlength(icon_glyph, font=f_icon)
    draw2.text((icon_cx - iw / 2, icon_cy - f_icon.size * 0.62), icon_glyph, font=f_icon, fill="white")

    f_contact = _font("bold", int(W * 0.034))
    contact_display = f"{CONTACT_NUMBER}  •  {BRAND_WEBSITE}"
    draw2.text((icon_cx + icon_r * 1.7, icon_cy - f_contact.size * 0.55), contact_display, font=f_contact, fill="white")

    top_buf = io.BytesIO(); top_img.save(top_buf, format="PNG"); top_buf.seek(0)
    bottom_buf = io.BytesIO(); bottom_img.save(bottom_buf, format="PNG"); bottom_buf.seek(0)

    return top_buf.read(), top_bar_h, bottom_buf.read(), bottom_h


def build_branded_video(video_bytes, fields):
    """
    Brands a video the same way build_branded_image brands photos: an
    orange top bar above the clip, a navy info+contact panel below it.
    Requires the `ffmpeg`/`ffprobe` binaries to be installed and on PATH
    on the server (e.g. `apt-get install ffmpeg` in your Dockerfile).
    Returns branded MP4 bytes. Raises RuntimeError if ffmpeg is missing,
    or subprocess.CalledProcessError if the encode fails.
    """
    if not shutil.which("ffmpeg") or not shutil.which("ffprobe"):
        raise RuntimeError("ffmpeg/ffprobe not installed on this server — cannot brand video")

    tmp_in = tempfile.NamedTemporaryFile(delete=False, suffix=".mp4")
    tmp_in.write(video_bytes); tmp_in.close()
    tmp_top = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    tmp_bottom = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    tmp_out = tempfile.NamedTemporaryFile(delete=False, suffix=".mp4")
    tmp_out.close()

    try:
        TARGET_W = 1080
        src_w, src_h = _get_video_dimensions(tmp_in.name)
        scaled_h = int(src_h * (TARGET_W / src_w))
        if scaled_h % 2:
            scaled_h += 1  # ffmpeg needs even dimensions

        top_bytes, top_h, bottom_bytes, bottom_h = _build_brand_bars(fields, TARGET_W)
        tmp_top.write(top_bytes); tmp_top.close()
        tmp_bottom.write(bottom_bytes); tmp_bottom.close()

        total_h = top_h + scaled_h + bottom_h
        if total_h % 2:
            total_h += 1

        filter_complex = (
            f"[0:v]scale={TARGET_W}:{scaled_h},setsar=1[v0];"
            f"[v0]pad={TARGET_W}:{total_h}:0:{top_h}:color=0x10131c[padded];"
            f"[padded][1:v]overlay=0:0[tmp1];"
            f"[tmp1][2:v]overlay=0:{top_h + scaled_h}[vout]"
        )

        cmd = [
            "ffmpeg", "-y",
            "-i", tmp_in.name,
            "-i", tmp_top.name,
            "-i", tmp_bottom.name,
            "-filter_complex", filter_complex,
            "-map", "[vout]", "-map", "0:a?",
            "-c:v", "libx264", "-preset", "veryfast", "-crf", "23",
            "-c:a", "aac", "-b:a", "128k",
            "-movflags", "+faststart",
            tmp_out.name
        ]
        subprocess.run(cmd, check=True, capture_output=True)

        with open(tmp_out.name, "rb") as f:
            return f.read()
    finally:
        for p in (tmp_in.name, tmp_top.name, tmp_bottom.name, tmp_out.name):
            try:
                os.remove(p)
            except Exception:
                pass

@app.route("/")
def loginpage():
    # If already logged in -> redirect based on role
    if "user_id" in session:
        role = session.get("role")

        if role == "admin":
            return redirect("/admin")
        elif role == "emp":
            return redirect("/admin")
        elif role == "partner":
            return redirect("/inventory")
        else:
            # fallback (invalid role)
            session.clear()
            return redirect("/")

    # Not logged in -> show login page
    return render_template("index.html")


#login system
@app.route("/login", methods=["POST"])
def login():
    raw_number = request.form.get("number", "")
    raw_password = request.form.get("password", "")

    number = normalize_number(raw_number)

    if not number:
        flash("Invalid phone number format")
        return redirect("/")

    try:
        number = int(number)
    except ValueError:
        flash("Invalid phone number format")
        return redirect("/")

    password = raw_password.strip()

    remember = request.form.get("remember")

    collection = db["teamAssign"]

    user = collection.find_one({
        "Employee number": number,
        "password": password
    })

    print(f"[login] number={number} found_user={bool(user)}"
          + (f" roll={user.get('roll')!r}" if user else ""))

    if not user:
        flash("Invalid number or password")
        return redirect("/")

    role = (user.get("roll") or "").strip().lower()

    # Store session
    session["user_id"] = str(user["_id"])
    session["role"] = role
    session["employee_name"] = user.get("Employee name")
    session["employee_number"] = user.get("Employee number")
    session.permanent = bool(remember)

    # Redirect based on role
    # Redirect based on role
    if role == "admin":
        return redirect("/admin")
    elif role == "emp":
        return redirect("/admin")
    elif role == "partner":
        return redirect("/inventory")
    else:
        flash("Invalid role")
        return redirect("/")


@app.route("/upload_page")
def upload_page():
    return render_template("upload.html")


@app.route("/admin")
def admin():
    if not session.get("user_id") or session.get("role") not in ("admin", "emp"):
        return redirect("/")

    return render_template(
        "admin.html",
        employee_name=session.get("employee_name"),
        employee_number=session.get("employee_number"),
        role=session.get("role")
    )

@app.route("/emp")
def emp():
    # Only check if logged in
    if not session.get("user_id"):
        return redirect("/")

    return render_template(
        "emp.html",
        employee_name=session.get("employee_name"),
        employee_number=session.get("employee_number")
    )

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

@app.route("/assignehd")
def assign():
    return render_template("assignehd.html")



@app.route("/leadjourney")
def leadjourney():
    return render_template("leadjourney.html")


@app.route("/manageteam")
def manageteam():
    return render_template("manageteam.html")

@app.route("/status")
def status():
    return render_template("status.html")

@app.route("/addtemplete")
def addtemplete():
    return render_template("addtemplete.html")

@app.route("/addlead")
def addlead():
    return render_template("addlead.html")


@app.route("/leadjourney")
def lead_journey():
    return render_template("leadjourney.html")


@app.route("/inventory")
def inventory():
    if not session.get("user_id"):
        return redirect("/")

    role = session.get("role")
    if role not in ("admin", "emp", "partner"):
        session.clear()
        return redirect("/")

    return render_template(
        "inventory_dash.html",
        employee_name=session.get("employee_name"),
        employee_number=session.get("employee_number"),
        role=role
    )


@app.route("/upload", methods=["POST"])
def upload():
    try:
        sheet_name = request.form.get("sheet_name")

        if not sheet_name:
            flash("Collection name is required")
            return redirect("/")

        if "file" not in request.files:
            flash("No file selected")
            return redirect("/")

        file = request.files["file"]

        if file.filename == "":
            flash("Please select a CSV file")
            return redirect("/")

        filepath = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
        file.save(filepath)

        df = pd.read_csv(filepath)
        df = df.where(pd.notnull(df), None)

        data = df.to_dict(orient="records")

        if not data:
            flash("CSV is empty")
            return redirect("/")

        collection = db[sheet_name]
        collection.insert_many(data)

        flash(f"Successfully inserted {len(data)} records into '{sheet_name}' collection!")
        return redirect("/")

    except Exception as e:
        flash(f"Error: {str(e)}")
        return redirect("/")


# =============================
# APIs
# =============================

# CHANGED: now uses get_filtered_leads() — indexed DateObj range query
# instead of fetching the entire collection and filtering in Python.
@app.route("/api/leads")
def leads():
    return jsonify(get_filtered_leads("Leads"))

#single lead
def clean_nan(data):
    for key, value in data.items():
        if isinstance(value, float) and math.isnan(value):
            data[key] = None
    return data


@app.route("/api/get-lead-single", methods=["POST"])
def get_lead():
    try:
        data = request.get_json()

        collection_name = data.get("collection")
        phone_number = data.get("phone")

        if not collection_name or not phone_number:
            return jsonify({"error": "collection and phone are required"}), 400

        collection = db[collection_name]

        lead = collection.find_one({"Phone Number": str(phone_number)})

        if not lead:
            return jsonify({"message": "No Lead"}), 404

        # Remove unwanted fields
        lead.pop("_id", None)
        lead.pop("Phone Number", None)

        lead = clean_nan(lead)

        return jsonify(lead), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


#realtorsdata
@app.route("/api/realtors")
def realtors():
    return jsonify(get_collection_data("Realtors"))


@app.route("/api/update-realtor", methods=["POST"])
def update_realtor():
    try:
        data = request.get_json()

        phone = data.get("phone")
        updates = data.get("updates")  # dict of fields to update

        if not phone or not updates:
            return jsonify({"error": "phone and updates are required"}), 400

        # Clean phone (remove +, spaces, etc.)
        phone = str(phone)
        phone = re.sub(r"\D", "", phone)

        try:
            phone = int(phone)
        except:
            return jsonify({"error": "Invalid phone format"}), 400

        collection = db["Realtors"]

        # Remove invalid fields
        if "_id" in updates:
            del updates["_id"]

        # Clean NaN if coming from frontend
        updates = clean_nan(updates)

        result = collection.update_one(
            {"Phone Number": phone},
            {"$set": updates}
        )

        if result.matched_count == 0:
            return jsonify({"message": "No realtor found"}), 404

        return jsonify({
            "message": "Realtor updated successfully",
            "modified_count": result.modified_count
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/hofcorders")
def hofcorders():
    return jsonify(get_collection_data("orderhouseofcakes"))

# CHANGED: now uses get_filtered_leads()
@app.route("/api/rental-leads")
def rental_leads():
    return jsonify(get_filtered_leads("RentalLeads"))

# CHANGED: now uses get_filtered_leads()
@app.route("/api/agent-leads")
def agent_leads():
    return jsonify(get_filtered_leads("agentLeads"))

# CHANGED: now uses get_filtered_leads()
@app.route("/api/selling-leads")
def selling_leads():
    return jsonify(get_filtered_leads("sellingLeads"))

@app.route("/api/end-data")
def get_end_data():

    collection = db["endData"]   # define collection

    number = request.args.get("number")

    # If number provided -> return single lead
    if number:
        try:
            lead = collection.find_one({"Number": int(number)})
        except:
            lead = collection.find_one({"Number": number})

        if lead:
            return jsonify(serialize_doc(lead))
        return jsonify({"error": "Not found"}), 404

    # If no number -> return all leads
    leads = list(collection.find())
    return jsonify([serialize_doc(l) for l in leads])


@app.route("/api/get-team-assign", methods=["GET"])
def get_team_assign():
    collection = db["teamAssign"]
    data = list(collection.find())

    return jsonify([serialize_doc(doc) for doc in data])


@app.route("/api/add-team-assign", methods=["POST"])
def add_team_member():
    try:
        data = request.json

        name = data.get("name")
        number = data.get("number")
        role = data.get("role", "emp")  # default emp

        if role not in ("admin", "emp", "partner"):
            return jsonify({"success": False, "message": "Invalid role"}), 400

        if not name or not number:
            return jsonify({"success": False, "message": "Missing fields"}), 400

        # CLEAN NUMBER
        number = str(number).strip()
        number = number.replace("+", "")
        number = "".join(filter(str.isdigit, number))

        if not number:
            return jsonify({"success": False, "message": "Invalid phone number"}), 400

        # CONVERT TO INT64
        number = int(number)

        collection = db["teamAssign"]

        # prevent duplicate
        existing = collection.find_one({"Employee number": number})
        if existing:
            return jsonify({"success": False, "message": "Employee already exists"}), 400

        # GENERATE PASSWORD
        clean_name = name.lower().replace(" ", "")
        rand_digits = random.randint(100, 9999)  # 3-4 digits
        password = f"{clean_name}@{rand_digits}"

        new_member = {
            "Employee name": name,
            "Employee number": number,
            "password": password,
            "roll": role,  # as requested (roll, not role)
            "Leads": [],
            "Active": True
        }

        collection.insert_one(new_member)

        # SEND TO N8N WEBHOOK
        try:
            requests.post(
                "https://n8n.phishnix.site/webhook/recevingdataofteammember",
                json={
                    "name": name,
                    "number": number,
                    "password": password,
                    "login_url": "https://crm.nishahomes.com/",
                    "message": f"Welcome {name}, your account has been created"
                },
                timeout=5
            )
        except Exception as webhook_error:
            print("Webhook failed:", webhook_error)  # don't break main flow

        return jsonify({
            "success": True,
            "message": "Team member added successfully"
        })

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/remove-team-assign/<number>", methods=["DELETE"])
def remove_team_member(number):
    try:
        collection = db["teamAssign"]

        number = str(number).strip()
        number = number.replace("+", "")
        number = "".join(filter(str.isdigit, number))

        if not number:
            return jsonify({"success": False, "message": "Invalid number"}), 400

        number = int(number)

        result = collection.delete_one({"Employee number": number})

        if result.deleted_count == 0:
            return jsonify({"success": False, "message": "Member not found"}), 404

        return jsonify({"success": True, "message": "Team member removed"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500



@app.route("/api/reset-team-password/<number>", methods=["POST"])
def reset_team_password(number):
    """
    Regenerates a plaintext password for an existing team member
    (covers older records created before/without a password field,
    or whenever an admin wants to issue a fresh one).
    """
    try:
        if session.get("role") != "admin":
            return jsonify({"success": False, "message": "Admin only"}), 403

        number = str(number).strip()
        number = number.replace("+", "")
        number = "".join(filter(str.isdigit, number))
        if not number:
            return jsonify({"success": False, "message": "Invalid number"}), 400
        number = int(number)

        collection = db["teamAssign"]
        member = collection.find_one({"Employee number": number})
        if not member:
            return jsonify({"success": False, "message": "Team member not found"}), 404

        name = member.get("Employee name", "user")
        clean_name = name.lower().replace(" ", "")
        rand_digits = random.randint(100, 9999)
        new_password = f"{clean_name}@{rand_digits}"

        collection.update_one(
            {"_id": member["_id"]},
            {"$set": {"password": new_password}}
        )

        # Notify via the same webhook used on creation (best-effort, won't break the response)
        try:
            requests.post(
                "https://n8n.phishnix.site/webhook/recevingdataofteammember",
                json={
                    "name": name,
                    "number": number,
                    "password": new_password,
                    "login_url": "https://api.phishnix.site",
                    "message": f"Hi {name}, your password has been reset"
                },
                timeout=5
            )
        except Exception as webhook_error:
            print("Webhook failed:", webhook_error)

        return jsonify({
            "success": True,
            "message": "Password reset successfully",
            "password": new_password
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

#post apis



def normalize_number(number):
    if not number:
        return ""

    number = str(number).strip()
    number = number.replace("+", "")
    number = number.replace("@s.whatsapp.net", "")
    number = number.replace("@c.us", "")
    number = "".join(filter(str.isdigit, number))

    return number


def normalize_phone_91(raw):
    """
    Normalizes ANY phone representation into a clean bare digit-only
    '91XXXXXXXXXX' string — no '+', no spaces, no leading zeros, and
    always exactly ONE '91' country-code prefix.

    This is stricter than normalize_number() alone, which just strips
    non-digit characters and leaves the '91' prefixing to a naive
    startswith() check at each call site — that's what was letting
    malformed numbers (leading zeros, missing prefix, accidental double
    prefix) slip into DAI/endData with the wrong shape.
    """
    digits = re.sub(r"\D", "", str(raw or ""))
    digits = digits.lstrip("0")  # drop stray leading zeros e.g. "0919876543210"

    if not digits:
        return ""

    # Bare 10-digit Indian mobile number -> add the prefix
    if len(digits) == 10:
        return "91" + digits

    # Already exactly "91" + 10 digits -> good as-is
    if len(digits) == 12 and digits.startswith("91"):
        return digits

    # Accidental double prefix e.g. "9191XXXXXXXXXX" -> collapse to one
    if len(digits) == 14 and digits.startswith("9191"):
        return digits[2:]

    # Anything longer than 12 that still starts with 91 -> keep the last
    # 12 digits (guards against extra junk prefixed by upstream sources)
    if len(digits) > 12 and digits.startswith("91"):
        return digits[-12:]

    # Fallback: doesn't cleanly fit the 10/12-digit shape — prefix as-is
    # rather than silently dropping the number.
    if not digits.startswith("91"):
        return "91" + digits

    return digits


@app.route("/api/assign-lead", methods=["POST"])
def assign_lead():
    try:
        data = request.json or {}

        collection_name = str(data.get("collection") or "").strip()
        lead_id = str(data.get("leadId") or "").strip()            # Mongo _id of the lead doc
        assign_to_number = data.get("assignToNumber")  # employee number, NOT name

        if not collection_name or not lead_id or not assign_to_number:
            return jsonify({"success": False, "message": "Missing fields"}), 400

        if collection_name not in ("Leads", "RentalLeads", "sellingLeads", "agentLeads"):
            return jsonify({"success": False, "message": f"Invalid collection: {collection_name}"}), 400

        try:
            assign_to_number = int(str(assign_to_number).strip())
        except ValueError:
            return jsonify({"success": False, "message": "Invalid employee number"}), 400

        try:
            obj_id = ObjectId(lead_id)
        except Exception:
            return jsonify({"success": False, "message": "Invalid lead id"}), 400

        employee = db["teamAssign"].find_one({"Employee number": assign_to_number})
        if not employee:
            return jsonify({"success": False, "message": "Employee not found"}), 404

        assigner_name = session.get("employee_name") or "Unknown"
        assigner_number = session.get("employee_number")

        history_entry = {
            "by": assigner_name,
            "byNumber": assigner_number,
            "to": employee.get("Employee name"),
            "toNumber": assign_to_number,
            "at": datetime.utcnow()
        }

        result = db[collection_name].update_one(
            {"_id": obj_id},
            {
                "$set": {
                    "AssignTo": employee.get("Employee name"),
                    "AssignToNumber": assign_to_number,
                    "AssignedBy": assigner_name,
                    "AssignedByNumber": assigner_number,
                    "AssignedAt": datetime.utcnow()
                },
                "$push": {"AssignmentHistory": history_entry}
            }
        )

        if result.matched_count == 0:
            return jsonify({"success": False, "message": "Lead not found"}), 404

        return jsonify({"success": True})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

# CHANGED: every user (admin or emp) now only sees leads assigned to the
# employee number they are CURRENTLY logged in as — even admin-to-admin
# assignments stay private to the assignee. Previously any admin saw
# every admin's assigned leads.
@app.route('/api/assigned-leads')
def assigned_leads():
    if not session.get("user_id"):
        return jsonify({"success": False, "message": "Not authenticated"}), 401

    employee_number = session.get("employee_number")  # stored as int at login time

    collection_type_map = {
        "Leads": "buying",
        "RentalLeads": "rental",
        "sellingLeads": "selling",
        "agentLeads": "agent",
    }

    if not employee_number:
        return jsonify({"success": False, "message": "No employee number in session"}), 400

    query = {
        "AssignTo": {"$exists": True, "$nin": [None, ""]},
        "AssignToNumber": employee_number
    }

    # Cap per-collection to keep this fast/light with ~8k+ assigned docs.
    PER_COLLECTION_LIMIT = 500

    all_docs = []
    try:
        for collection_name, lead_type in collection_type_map.items():
            docs = (
                db[collection_name]
                .find(query)
                .sort("AssignedAt", -1)
                .limit(PER_COLLECTION_LIMIT)
            )
            for d in docs:
                d = serialize_doc(d)
                d["_leadType"] = lead_type
                d["_collection"] = collection_name
                all_docs.append(d)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

    def sort_key(doc):
        return doc.get("AssignedAt") or ""

    all_docs.sort(key=sort_key, reverse=True)

    return jsonify({"success": True, "data": all_docs})



@app.route("/api/get-lead-by-id", methods=["POST"])
def get_lead_by_id():
    try:
        data = request.json or {}
        collection_name = data.get("collection")
        lead_id = data.get("id")

        if not collection_name or not lead_id:
            return jsonify({"error": "collection and id are required"}), 400

        lead = db[collection_name].find_one({"_id": ObjectId(lead_id)})
        if not lead:
            return jsonify({"error": "Lead not found"}), 404

        return jsonify({"success": True, "data": serialize_doc(lead)})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/update-lead-by-id", methods=["POST"])
def update_lead_by_id():
    try:
        data = request.json or {}
        collection_name = data.get("collection")
        lead_id = data.get("id")
        set_fields = data.get("set", {})

        if not collection_name or not lead_id:
            return jsonify({"error": "collection and id are required"}), 400
        if not set_fields:
            return jsonify({"error": "No fields to update"}), 400

        set_fields.pop("_id", None)

        result = db[collection_name].update_one(
            {"_id": ObjectId(lead_id)},
            {"$set": set_fields}
        )

        if result.matched_count == 0:
            return jsonify({"error": "Lead not found"}), 404

        updated = db[collection_name].find_one({"_id": ObjectId(lead_id)})
        return jsonify({"success": True, "data": serialize_doc(updated)})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/api/bulk-assign-leads", methods=["POST"])
def bulk_assign_leads():
    try:
        data = request.json or {}

        collection_name = str(data.get("collection") or "").strip()
        lead_ids = [str(i).strip() for i in (data.get("leadIds") or []) if str(i).strip()]
        assign_to_number = data.get("assignToNumber")

        if not collection_name or not lead_ids or not assign_to_number:
            return jsonify({"success": False, "message": "Missing fields"}), 400

        if collection_name not in ("Leads", "RentalLeads", "sellingLeads", "agentLeads"):
            return jsonify({"success": False, "message": f"Invalid collection: {collection_name}"}), 400

        try:
            assign_to_number = int(str(assign_to_number).strip())
        except ValueError:
            return jsonify({"success": False, "message": "Invalid employee number"}), 400

        employee = db["teamAssign"].find_one({"Employee number": assign_to_number})
        if not employee:
            return jsonify({"success": False, "message": "Employee not found"}), 404

        try:
            obj_ids = [ObjectId(i) for i in lead_ids]
        except Exception:
            return jsonify({"success": False, "message": "Invalid lead id in list"}), 400

        assigner_name = session.get("employee_name") or "Unknown"
        assigner_number = session.get("employee_number")

        history_entry = {
            "by": assigner_name,
            "byNumber": assigner_number,
            "to": employee.get("Employee name"),
            "toNumber": assign_to_number,
            "at": datetime.utcnow()
        }

        result = db[collection_name].update_many(
            {"_id": {"$in": obj_ids}},
            {
                "$set": {
                    "AssignTo": employee.get("Employee name"),
                    "AssignToNumber": assign_to_number,
                    "AssignedBy": assigner_name,
                    "AssignedByNumber": assigner_number,
                    "AssignedAt": datetime.utcnow()
                },
                "$push": {"AssignmentHistory": history_entry}
            }
        )

        return jsonify({"success": True, "assignedCount": result.modified_count})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


#reassign function


@app.route("/api/reassign-lead", methods=["POST"])
def reassign_lead():
    try:
        data = request.json

        phone = data.get("phone")
        new_employee_number = data.get("newEmployeeNumber")
        collection_name = data.get("collection")

        # Validate input
        if not phone or not new_employee_number or not collection_name:
            return jsonify({"error": "Missing required fields"}), 400

        phone = phone.replace("+", "").strip()
        formatted_phone = f"+{phone}"

        team_collection = db["teamAssign"]
        lead_collection = db[collection_name]

        # ESCAPE REGEX PROPERLY
        safe_phone_regex = re.escape(formatted_phone)

        # 1. Find current employee safely
        current_employee = team_collection.find_one({
            "Leads": {"$regex": safe_phone_regex}
        })

        if not current_employee:
            return jsonify({"error": "Lead not found in any employee"}), 404

        # 2. Remove lead from old employee
        old_leads_string = current_employee.get("Leads", "{}")

        old_list = old_leads_string.strip("{}").split(",")
        old_list = [l.strip() for l in old_list if l.strip()]

        updated_old_list = [l for l in old_list if l != formatted_phone]

        new_old_string = "{" + ", ".join(updated_old_list) + "}"

        team_collection.update_one(
            {"_id": current_employee["_id"]},
            {"$set": {"Leads": new_old_string}}
        )

        # 3. Add lead to new employee
        new_employee = team_collection.find_one({
            "Employee number": int(new_employee_number)
        })

        if not new_employee:
            return jsonify({"error": "New employee not found"}), 404

        new_leads_string = new_employee.get("Leads", "{}")

        new_list = new_leads_string.strip("{}").split(",")
        new_list = [l.strip() for l in new_list if l.strip()]

        if formatted_phone not in new_list:
            new_list.append(formatted_phone)

        updated_new_string = "{" + ", ".join(new_list) + "}"

        team_collection.update_one(
            {"_id": new_employee["_id"]},
            {"$set": {"Leads": updated_new_string}}
        )

        # 4. Update AssignTo in Lead document safely
        lead_doc = lead_collection.find_one({"Phone Number": phone})

        if not lead_doc:
            return jsonify({"error": "Lead not found in lead collection"}), 404

        current_assign = lead_doc.get("AssignTo")

        new_employee_name = new_employee["Employee name"]

        if not current_assign:
            updated_assign = new_employee_name
        else:
            current_assign = str(current_assign)
            names = [n.strip() for n in current_assign.split(",") if n.strip()]

            if new_employee_name not in names:
                names.append(new_employee_name)

            updated_assign = ", ".join(names)

        lead_collection.update_one(
            {"_id": lead_doc["_id"]},
            {"$set": {"AssignTo": updated_assign}}
        )

        return jsonify({"success": True})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/call-attempt", methods=["POST"])
def call_attempt():
    """
    Fires whenever ANY 'Call' button is clicked, from ANY page (All Leads,
    Assigned Leads, Followups, Hot/Warm/Cold, etc). Increments Total Calls,
    pushes a {At, By} entry into CallHistory (array, so every call attempt
    is kept, not just the latest), and also writes a lightweight row into
    callLogs so the dashboard's "Team Activity - Calls Today" picks it up
    automatically (that widget already aggregates callLogs by DateOnly).
    """
    try:
        data = request.json or {}
        number = data.get("number")

        if not number:
            return jsonify({"error": "Number is required"}), 400

        # Clean number
        number = str(number).replace("+", "").strip()

        employee_name = session.get("employee_name") or data.get("employee") or "Unknown"
        now = datetime.utcnow()
        formatted_dt = format_ist(now)
        today_str = now.strftime("%Y-%m-%d")

        end_collection = db["endData"]

        end_collection.update_one(
            {"Number": number},
            {
                "$inc": {"Call_attempt": 1},
                "$setOnInsert": {"Number": number},
                "$push": {
                    "CallHistory": {
                        "$each": [{
                            "At": now,
                            "AtFormatted": formatted_dt,
                            "By": employee_name
                        }],
                        "$slice": -100
                    }
                }
            },
            upsert=True
        )

        # Also record it in callLogs so Team Activity / Total Calls count it
        call_logs_collection.insert_one({
            "Number": number,
            "Name": data.get("name", ""),
            "LeadType": data.get("leadType", ""),
            "CallAttemptOnly": True,
            "CallDateTimeFormatted": formatted_dt,
            "CalledBy": employee_name,
            "CreatedAt": now,
            "DateOnly": today_str
        })

        updated_doc = end_collection.find_one({"Number": number})

        return jsonify({
            "success": True,
            "Number": number,
            "Call_attempt": updated_doc.get("Call_attempt", 1),
            "CalledBy": employee_name,
            "CallDateTime": formatted_dt
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


#adding call logs

call_logs_collection = db["callLogs"]

# ============================================================
# CALL LOG - replaces the n8n webhook, writes straight to Mongo
# ============================================================

COLLECTION_MAP = {
    "buying": "Leads",
    "rental": "RentalLeads",
    "selling": "sellingLeads",
    "agent": "agentLeads",
    "other": "Leads"
}


@app.route("/api/call-log", methods=["POST"])
def add_call_log():
    try:
        data = request.json
        if not data:
            return jsonify({"error": "No JSON body provided"}), 400

        phone = data.get("phone")
        if not phone:
            return jsonify({"error": "Phone is required"}), 400

        number = normalize_number(phone)
        if not number:
            return jsonify({"error": "Invalid phone number"}), 400

        if not number.startswith("91"):
            number = "91" + number

        employee_name = session.get("employee_name") or data.get("employee") or "Unknown"

        now = datetime.utcnow()
        today_str = now.strftime("%Y-%m-%d")
        formatted_dt = format_ist(now)

        end_collection = db["endData"]

        # Figure out which call attempt number this is, BEFORE incrementing
        existing_end_doc = end_collection.find_one({"Number": number})
        current_attempt = (existing_end_doc or {}).get("Call_attempt", 0)
        attempt_number = current_attempt + 1

        log_entry = {
            "Number": number,
            "Name": data.get("name", ""),
            "LeadType": data.get("leadType", ""),
            "CallAttemptNumber": attempt_number,
            "CallDateTimeFormatted": formatted_dt,
            "CallStatus": data.get("callStatus", ""),
            "CustomerResponse": data.get("customerResponse", ""),
            "InterestLevel": data.get("interestLevel", ""),
            "Configuration": data.get("configuration", ""),
            "Objection": data.get("objection", ""),
            "FollowupTimeline": data.get("followupTimeline", ""),
            "NextCallDate": data.get("nextCallDate", ""),
            "CallPriority": data.get("callPriority", ""),
            "Status": data.get("status", "Pending"),
            "CallerRemarks": data.get("callerRemarks", ""),
            "LeadSnapshot": {
                "Location": data.get("location", ""),
                "Property": data.get("property", ""),
                "Budget": data.get("budget", ""),
                "Timeline": data.get("timeline", ""),
                "Note": data.get("note", "")
            },
            "CalledBy": employee_name,
            "CreatedAt": now,
            "DateOnly": today_str
        }

        call_logs_collection.insert_one(log_entry)

        lead_type = data.get("leadType")
        collection_name = COLLECTION_MAP.get(lead_type)
        if collection_name:
            lead_doc = db[collection_name].find_one({"Phone Number": {"$regex": number}})
            if lead_doc:
                db["endData"].update_one(
                    {"Number": number},
                    {"$set": {"LeadId": str(lead_doc["_id"])}},
                    upsert=True
                )
                db[collection_name].update_one(
                    {"_id": lead_doc["_id"]},
                    {"$set": {"callBy": session.get("user_id")}}
                )

        update_fields = {
            "Call Status": data.get("callStatus", ""),
            "Customer Response": data.get("customerResponse", ""),
            "Interest Level": data.get("interestLevel", ""),
            "Configuration": data.get("configuration", ""),
            "Objection / Reason": data.get("objection", ""),
            "Next Follow-up Timeline": data.get("followupTimeline", ""),
            "Next Call Date": data.get("nextCallDate", ""),
            "Call Priority": data.get("callPriority", ""),
            "Status": data.get("status", "Pending"),
            "Caller Remarks": data.get("callerRemarks", ""),
            "Location Interested In": data.get("location", ""),
            "Property Type": data.get("property", ""),
            "Budget Range": data.get("budget", ""),
            "Customer Name": data.get("name", ""),
            "lastCallBy": employee_name,
            "lastCallAt": now,
            "lastCallAtFormatted": formatted_dt,
            "lastUpdatedAt": now
        }
        update_fields = {k: v for k, v in update_fields.items() if v not in [None, ""]}

        update_fields["LastLeadType"] = data.get("leadType", "")

        end_collection.update_one(
            {"Number": number},
            {
                "$set": update_fields,
                "$inc": {"Call_attempt": 1},
                "$push": {
                    "RecentLogs": {
                        "$each": [{
                            "CallAttemptNumber": attempt_number,
                            "CallDateTimeFormatted": formatted_dt,
                            "CallStatus": data.get("callStatus", ""),
                            "CustomerResponse": data.get("customerResponse", ""),
                            "CalledBy": employee_name,
                            "At": now,
                            "Remarks": data.get("callerRemarks", "")
                        }],
                        "$slice": -10
                    }
                }
            },
            upsert=True
        )

        # NEW: AI intent classification (Hot / Warm / Cold) using MISTRAL_API_KEY2
        intent_result = classify_lead_intent(
            {
                "name": data.get("name", ""),
                "location": data.get("location", ""),
                "property": data.get("property", ""),
                "budget": data.get("budget", ""),
                "timeline": data.get("timeline", ""),
                "note": data.get("note", "")
            },
            data
        )
        if intent_result:
            end_collection.update_one(
                {"Number": number},
                {"$set": {
                    "AI Intent": intent_result["intent"],
                    "AI Intent Reason": intent_result.get("reason", ""),
                    "AI Intent At": now
                }}
            )

        updated_doc = end_collection.find_one({"Number": number})

        return jsonify({
            "success": True,
            "message": "Call log saved",
            "attempt_number": attempt_number,
            "call_datetime": formatted_dt,
            "data": serialize_doc(updated_doc) if updated_doc else None
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/call-logs/<phone>", methods=["GET"])
def get_call_logs(phone):
    try:
        number = normalize_number(phone)
        if not number.startswith("91"):
            number = "91" + number

        logs = list(call_logs_collection.find({"Number": number}).sort("CreatedAt", -1))
        for l in logs:
            l["_id"] = str(l["_id"])
            if isinstance(l.get("CreatedAt"), datetime):
                l["CreatedAt"] = l["CreatedAt"].isoformat()

        return jsonify({"success": True, "count": len(logs), "logs": logs}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/dashboard-stats", methods=["GET"])
def dashboard_stats():
    try:
        end_collection = db["endData"]
        today_str = datetime.utcnow().strftime("%Y-%m-%d")

        # PERF: projection — only pull the fields actually used below.
        all_docs = list(end_collection.find({}, {
            "Number": 1, "Status": 1, "Next Follow-up Timeline": 1
        }))

        followup_count = 0
        pending_count = 0
        done_count = 0

        for d in all_docs:
            status = (d.get("Status") or "").strip().lower()
            followup = (d.get("Next Follow-up Timeline") or "").strip()
            if followup:
                followup_count += 1
            if status == "done":
                done_count += 1
            else:
                pending_count += 1

        today_logs = list(call_logs_collection.find({"DateOnly": today_str}, {"CalledBy": 1}))
        calls_today_total = len(today_logs)

        by_employee = {}
        for l in today_logs:
            name = l.get("CalledBy", "Unknown")
            by_employee[name] = by_employee.get(name, 0) + 1

        return jsonify({
            "success": True,
            "followup_count": followup_count,
            "pending_count": pending_count,
            "done_count": done_count,
            "calls_today_total": calls_today_total,
            "calls_today_by_employee": by_employee
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


# =============================
# NEW: FOLLOW-UPS DUE TODAY / THIS WEEK (for admin dashboard widget)
# Scans every lead across all 4 collections, looks up its endData
# doc for "Next Call Date", and buckets it into today / this week.
# Visible to admin & emp; shows follow-ups set by ANY employee/admin.
# =============================
def get_followup_period_range(period):
    """Returns (start_date, end_date) as date objects for filtering
    follow-ups by their Next Call Date (both bounds inclusive).
    (None, None) means no restriction — "lifetime" = every follow-up
    that has a Next Call Date, whenever it falls."""
    today = datetime.utcnow().date()

    if period == "today":
        return today, today

    if period == "this_week":
        start = today - timedelta(days=today.weekday())  # Monday
        return start, start + timedelta(days=6)           # Sunday

    if period == "this_month":
        start = today.replace(day=1)
        end = today.replace(day=calendar.monthrange(today.year, today.month)[1])
        return start, end

    if period == "last_month":
        first_of_this_month = today.replace(day=1)
        last_month_end = first_of_this_month - timedelta(days=1)
        return last_month_end.replace(day=1), last_month_end

    if period == "last_3_months":
        month = today.month - 2
        year = today.year
        while month <= 0:
            month += 12
            year -= 1
        start = datetime(year, month, 1).date()
        end = today.replace(day=calendar.monthrange(today.year, today.month)[1])
        return start, end

    return None, None  # "lifetime"



@app.route("/api/dashboard-followups", methods=["GET"])
def dashboard_followups():
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False, "message": "Staff only"}), 403

    period = request.args.get("period", "today")
    range_start, range_end = get_followup_period_range(period)

    now = datetime.utcnow()

    def parse_followup_date(s):
        if not s:
            return None
        s = str(s).strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    end_collection = db["endData"]
    collections = {
        "buying": "Leads", "rental": "RentalLeads",
        "selling": "sellingLeads", "agent": "agentLeads"
    }

    followups_list = []

    try:
        # PERF: collect every candidate lead + its normalized phone FIRST,
        # with zero endData queries yet — then do exactly ONE batched
        # $in lookup for all their endData docs instead of one find_one()
        # per lead (this was the N+1 hotspot in this endpoint).
        candidate_leads = []   # (lead, lead_type, phone)
        phones_needed = set()

        for lead_type, coll_name in collections.items():
            for lead in db[coll_name].find():
                lead_date = parse_lead_date(lead.get("Date"))
                if not lead_date or lead_date < JULY_2026_START or lead_date > now:
                    continue

                phone = normalize_number(lead.get("Phone Number", ""))
                if not phone:
                    continue
                if not phone.startswith("91"):
                    phone = "91" + phone

                candidate_leads.append((lead, lead_type, phone))
                phones_needed.add(phone)

        end_data_map = {}
        if phones_needed:
            for ed in end_collection.find({"Number": {"$in": list(phones_needed)}}):
                end_data_map[ed.get("Number")] = ed

        def _clean(v, default="-"):
            if v is None or v == "":
                return default
            if isinstance(v, float) and math.isnan(v):
                return default
            return v

        for lead, lead_type, phone in candidate_leads:
            ed = end_data_map.get(phone)
            if not ed:
                continue

            ed_created = ed.get("lastUpdatedAt")
            if isinstance(ed_created, datetime):
                if ed_created < JULY_2026_START or ed_created > now:
                    continue

            fdate = parse_followup_date(ed.get("Next Call Date"))
            if not fdate:
                continue

            if range_start is not None and not (range_start <= fdate <= range_end):
                continue

            actual_type = _normalize_lead_type_field(lead.get("LeadType"))

            entry = {
                "id": str(lead["_id"]),
                "collection": collections[lead_type],
                "leadType": actual_type,
                "name": _clean(lead.get("Lead Name") or lead.get("Name"), "Unknown"),
                "phone": phone,
                "assignedTo": _clean(lead.get("AssignTo"), "Unassigned"),
                "nextCallDate": _clean(ed.get("Next Call Date")),
                "nextFollowupTimeline": _clean(ed.get("Next Follow-up Timeline")),
                "callStatus": _clean(ed.get("Call Status")),
                "location": _clean(lead.get("Location Interested In") or lead.get("Property Location")),
                "propertyType": _clean(lead.get("Property Type") or ed.get("Property Type")),
                "budget": _clean(lead.get("Budget Range") or lead.get("Expected Price")),
                "customerResponse": _clean(ed.get("Customer Response")),
                "interestLevel": _clean(ed.get("Interest Level")),
                "callerRemarks": _clean(ed.get("Caller Remarks")),
                "callAttempts": ed.get("Call_attempt") if isinstance(ed.get("Call_attempt"), int) else 0,
            }

            followups_list.append(entry)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

    return jsonify({
        "success": True,
        "period": period,
        "count": len(followups_list),
        "followups": followups_list
    }), 200


@app.route("/api/dashboard-intent-leads", methods=["GET"])
def dashboard_intent_leads():
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False, "message": "Staff only"}), 403

    def map_interest_to_bucket(level):
        lvl = (level or "").strip().lower()
        if lvl in ("very high", "high"):
            return "Hot"
        if lvl == "medium":
            return "Warm"
        if lvl == "cold":
            return "Cold"
        return None

    end_collection = db["endData"]
    docs = list(end_collection.find({
        "Interest Level": {"$in": ["Very High", "High", "Medium", "Cold"]}
    }))

    def _clean(v, default="-"):
        if v is None or v == "":
            return default
        if isinstance(v, float) and math.isnan(v):
            return default
        return v

    buckets = {"Hot": [], "Warm": [], "Cold": []}
    for d in docs:
        bucket = map_interest_to_bucket(d.get("Interest Level"))
        if bucket not in buckets:
            continue
        buckets[bucket].append({
            "number": d.get("Number", ""),
            "name": _clean(d.get("Customer Name"), "Unknown"),
            "location": _clean(d.get("Location Interested In")),
            "propertyType": _clean(d.get("Property Type")),
            "budget": _clean(d.get("Budget Range")),
            "callStatus": _clean(d.get("Call Status")),
            "customerResponse": _clean(d.get("Customer Response")),
            "interestLevel": _clean(d.get("Interest Level")),
            "callerRemarks": _clean(d.get("Caller Remarks")),
            "nextFollowupTimeline": _clean(d.get("Next Follow-up Timeline")),
            "nextCallDate": _clean(d.get("Next Call Date")),
            "callAttempts": d.get("Call_attempt") if isinstance(d.get("Call_attempt"), int) else 0,
            "leadId": d.get("LeadId", ""),
            "type": d.get("LastLeadType", "buying") or "buying",
        })

    return jsonify({
        "success": True,
        "hotCount": len(buckets["Hot"]),
        "warmCount": len(buckets["Warm"]),
        "coldCount": len(buckets["Cold"]),
        "hotLeads": buckets["Hot"],
        "warmLeads": buckets["Warm"],
        "coldLeads": buckets["Cold"]
    }), 200


# =============================
# DASHBOARD OVERVIEW — single period-aware endpoint that powers every
# widget on the Dashboard page: KPI totals, Pipeline Health, Follow-ups,
# Team Activity, Lead Intent (AI), Trends and Breakdown.
#
# period = "today" | "this_month" | "last_month" | "last_3_months" | "lifetime"
#
# Filtering is based on each lead's "Created At" field. For "lifetime",
# every lead is included — even ones with no "Created At" at all (as
# requested). For any other period, leads without a parseable "Created At"
# are excluded (there's no reliable way to place them in a specific window).
# =============================
@app.route("/api/dashboard-overview", methods=["GET"])
def dashboard_overview():
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False, "message": "Staff only"}), 403

    period = request.args.get("period", "lifetime")
    start, end = get_dashboard_period_range(period)
    now = datetime.utcnow()

    collections = {
        "buying": "Leads", "rental": "RentalLeads",
        "selling": "sellingLeads", "agent": "agentLeads"
    }

    totals = {"buying": 0, "rental": 0, "selling": 0, "agent": 0, "other": 0}
    trend_counts = {}       # "YYYY-MM-DD" -> count
    location_counts = {}    # location -> count
    qualifying_phones = {}  # normalized "91xxxxxxxxxx" -> leadType

    try:
        for lead_type, coll_name in collections.items():
            for lead in db[coll_name].find():
                created_dt = parse_created_at_str(lead.get("Created At"))

                if start is not None:
                    if not created_dt or created_dt < start or created_dt > end:
                        continue
                # lifetime (start is None): include regardless of Created At —
                # leads with no Created At only ever show up under "lifetime"

                # CHANGED: bucket by the lead's ACTUAL "LeadType" field
                # (same mapping /api/dashboard-leads-by-type already uses),
                # not by which collection the doc happens to live in.
                actual_type = _normalize_lead_type_field(lead.get("LeadType"))
                totals[actual_type] = totals.get(actual_type, 0) + 1

                if created_dt:
                    d_str = created_dt.strftime("%Y-%m-%d")
                    trend_counts[d_str] = trend_counts.get(d_str, 0) + 1

                loc = (lead.get("Location Interested In") or lead.get("Property Location")
                       or lead.get("Operating City") or "").strip()
                if loc:
                    location_counts[loc] = location_counts.get(loc, 0) + 1

                phone = normalize_number(lead.get("Phone Number", ""))
                if phone:
                    if not phone.startswith("91"):
                        phone = "91" + phone
                    qualifying_phones[phone] = actual_type

        total_leads = sum(totals.values())

        # ---- Pipeline health / status breakdown / intent / follow-ups ----
        # PERF: this was already using a batched $in lookup — kept as-is.
        end_collection = db["endData"]
        followup_count = pending_count = done_count = lost_count = 0
        status_counts = {}
        intent_buckets = {"Hot": [], "Warm": [], "Cold": []}
        today_list, week_list = [], []

        def map_interest_to_bucket(level):
            lvl = (level or "").strip().lower()
            if lvl in ("very high", "high"):
                return "Hot"
            if lvl == "medium":
                return "Warm"
            if lvl == "cold":
                return "Cold"
            return None

        def _clean(v, default="-"):
            if v is None or v == "":
                return default
            if isinstance(v, float) and math.isnan(v):
                return default
            return v

        def parse_followup_date(s):
            if not s:
                return None
            s = str(s).strip()
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
            return None

        today = now.date()
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)

        if qualifying_phones:
            for ed in end_collection.find({"Number": {"$in": list(qualifying_phones.keys())}}):
                number = ed.get("Number")
                lead_type = qualifying_phones.get(number, "buying")

                status_val = (ed.get("Status") or "").strip().lower()
                followup_val = (ed.get("Next Follow-up Timeline") or "").strip()
                if followup_val:
                    followup_count += 1
                if status_val == "done":
                    done_count += 1
                else:
                    pending_count += 1

                call_status = ed.get("Call Status") or "No Log"
                status_counts[call_status] = status_counts.get(call_status, 0) + 1

                if _is_lost_lead(ed):
                    lost_count += 1

                bucket = map_interest_to_bucket(ed.get("Interest Level"))
                if bucket:
                    intent_buckets[bucket].append({
                        "number": number, "name": _clean(ed.get("Customer Name"), "Unknown"),
                        "location": _clean(ed.get("Location Interested In")),
                        "propertyType": _clean(ed.get("Property Type")),
                        "budget": _clean(ed.get("Budget Range")),
                        "callStatus": _clean(ed.get("Call Status")),
                        "customerResponse": _clean(ed.get("Customer Response")),
                        "interestLevel": _clean(ed.get("Interest Level")),
                        "callerRemarks": _clean(ed.get("Caller Remarks")),
                        "nextFollowupTimeline": _clean(ed.get("Next Follow-up Timeline")),
                        "nextCallDate": _clean(ed.get("Next Call Date")),
                        "callAttempts": ed.get("Call_attempt") if isinstance(ed.get("Call_attempt"), int) else 0,
                        "leadId": ed.get("LeadId", ""),
                        "type": lead_type,
                    })

                fdate = parse_followup_date(ed.get("Next Call Date"))
                if fdate and fdate >= today:
                    entry = {
                        "leadType": lead_type,
                        "name": _clean(ed.get("Customer Name"), "Unknown"),
                        "phone": number,
                        "assignedTo": "-",
                        "nextCallDate": _clean(ed.get("Next Call Date")),
                        "nextFollowupTimeline": _clean(ed.get("Next Follow-up Timeline")),
                        "callStatus": _clean(ed.get("Call Status")),
                        "location": _clean(ed.get("Location Interested In")),
                        "propertyType": _clean(ed.get("Property Type")),
                        "budget": _clean(ed.get("Budget Range")),
                        "customerResponse": _clean(ed.get("Customer Response")),
                        "interestLevel": _clean(ed.get("Interest Level")),
                        "callerRemarks": _clean(ed.get("Caller Remarks")),
                        "callAttempts": ed.get("Call_attempt") if isinstance(ed.get("Call_attempt"), int) else 0,
                    }
                    if fdate == today:
                        today_list.append(entry)
                    elif week_start <= fdate <= week_end:
                        week_list.append(entry)

        # ---- Team activity: calls within the selected period, by the call log's own timestamp ----
        call_query = {}
        if start is not None:
            call_query["CreatedAt"] = {"$gte": start, "$lte": end}
        by_employee = {}
        calls_period_total = 0
        for log in call_logs_collection.find(call_query, {"CalledBy": 1}):
            calls_period_total += 1
            name = log.get("CalledBy", "Unknown")
            by_employee[name] = by_employee.get(name, 0) + 1

        top_locations = dict(sorted(location_counts.items(), key=lambda x: -x[1])[:5])

        return jsonify({
            "success": True,
            "period": period,
            "totals": {"total": total_leads, **totals},
            "pipeline": {
                "followup_count": followup_count,
                "pending_count": pending_count,
                "done_count": done_count,
                "calls_period_total": calls_period_total,
                "lost_count": lost_count
            },
            "team_activity": {"by_employee": by_employee},
            "intent": {
                "hotCount": len(intent_buckets["Hot"]),
                "warmCount": len(intent_buckets["Warm"]),
                "coldCount": len(intent_buckets["Cold"]),
                "hotLeads": intent_buckets["Hot"],
                "warmLeads": intent_buckets["Warm"],
                "coldLeads": intent_buckets["Cold"]
            },
            "followups": {
                "todayCount": len(today_list), "weekCount": len(week_list),
                "todayLeads": today_list, "weekLeads": week_list
            },
            "trend": trend_counts,
            "status_breakdown": status_counts,
            "top_locations": top_locations
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


def _normalize_lead_type_field(lt):
    """Maps the raw LeadType FIELD on a doc to a filter bucket — same mapping
    the frontend's normalizeLeadType() uses. NOT which collection it lives in."""
    v = str(lt or "").strip().lower()
    return {
        "buyer_purchase": "buying",
        "buyer_rental": "rental",
        "seller": "selling",
        "agent": "agent",
    }.get(v, "other")


# =============================
# "LOST" LEADS — unreachable call statuses OR explicit "Not Interested"
# response, used by the Dashboard's Lost KPI card.
# =============================
LOST_CALL_STATUSES = {
    "Call Disconnected", "Call Not Picked", "Invalid Number", "Line Busy",
    "No Response", "Number Switched Off", "Wrong Number"
}

def _is_lost_lead(ed):
    """True if this endData doc counts as a 'lost' lead — either the call
    status is one of the unreachable statuses, or the customer explicitly
    said Not Interested."""
    call_status = (ed.get("Call Status") or "").strip()
    customer_response = (ed.get("Customer Response") or "").strip()
    return call_status in LOST_CALL_STATUSES or customer_response == "Not Interested"

@app.route("/api/dashboard-leads-by-type", methods=["GET"])
def dashboard_leads_by_type():
    """Powers the Dashboard KPI-card popups. Same period logic as
    /api/dashboard-overview, but returns the actual lead list (merged with
    endData call info) filtered by the LeadType FIELD instead of just counts.
    type = all | buying | rental | selling | agent
    """
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False, "message": "Staff only"}), 403

    period = request.args.get("period", "lifetime")
    want_type = request.args.get("type", "all")
    start, end = get_dashboard_period_range(period)

    coll_names = ["Leads", "RentalLeads", "sellingLeads", "agentLeads"]
    end_collection = db["endData"]

    def _clean(v, default="-"):
        if v is None or v == "":
            return default
        if isinstance(v, float) and math.isnan(v):
            return default
        return v

    try:
        # PERF: two-pass approach like dashboard_lost_leads — collect all
        # candidate leads + phones first (zero endData queries), then do
        # ONE batched $in lookup instead of a find_one() per lead.
        candidate_leads = []  # (lead, coll_name, phone, phone_raw)
        phones_needed = set()

        for coll_name in coll_names:
            for lead in db[coll_name].find():
                created_dt = parse_created_at_str(lead.get("Created At"))
                if start is not None:
                    if not created_dt or created_dt < start or created_dt > end:
                        continue

                lead_type = _normalize_lead_type_field(lead.get("LeadType"))
                if want_type != "all" and lead_type != want_type:
                    continue

                phone_raw = lead.get("Phone Number", "")
                phone = normalize_number(phone_raw)
                if phone and not phone.startswith("91"):
                    phone = "91" + phone

                candidate_leads.append((lead, coll_name, phone, phone_raw, lead_type))
                if phone:
                    phones_needed.add(phone)

        end_data_map = {}
        if phones_needed:
            for ed in end_collection.find({"Number": {"$in": list(phones_needed)}}):
                end_data_map[ed.get("Number")] = ed

        out = []
        for lead, coll_name, phone, phone_raw, lead_type in candidate_leads:
            ed = end_data_map.get(phone, {}) if phone else {}

            out.append({
                "id": str(lead["_id"]),
                "collection": coll_name,
                "leadType": lead_type,
                "name": _clean(lead.get("Lead Name") or lead.get("Name"), "Unknown"),
                "phone": phone or _clean(phone_raw, ""),
                "date": _clean(lead.get("Date")),
                "location": _clean(lead.get("Location Interested In") or lead.get("Property Location")),
                "propertyType": _clean(lead.get("Property Type")),
                "budget": _clean(lead.get("Budget Range") or lead.get("Expected Price")),
                "assignedTo": _clean(lead.get("AssignTo"), "Unassigned"),
                "callStatus": _clean(ed.get("Call Status")),
                "customerResponse": _clean(ed.get("Customer Response")),
                "interestLevel": _clean(ed.get("Interest Level")),
                "callerRemarks": _clean(ed.get("Caller Remarks")),
                "nextFollowupTimeline": _clean(ed.get("Next Follow-up Timeline")),
                "nextCallDate": _clean(ed.get("Next Call Date")),
                "callAttempts": ed.get("Call_attempt") if isinstance(ed.get("Call_attempt"), int) else 0,
                "lastCallBy": _clean(ed.get("lastCallBy")),
                "lastCallAtFormatted": _clean(ed.get("lastCallAtFormatted")),
            })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

    return jsonify({"success": True, "period": period, "type": want_type,
                     "count": len(out), "data": out}), 200


@app.route("/api/dashboard-lost-leads", methods=["GET"])
def dashboard_lost_leads():
    """Powers the 'Lost' KPI card popup. Same period logic as
    /api/dashboard-overview, but returns leads whose endData Call Status is
    one of the unreachable statuses, or whose Customer Response is
    'Not Interested'. status=<exact status | 'Not Interested' | 'all'>
    narrows to just that reason.

    PERFORMANCE: batches the endData lookup into ONE query instead of one
    find_one() per lead — with hundreds of leads the old per-lead lookup
    was doing hundreds of round-trips to Mongo, which is why this was slow.
    """
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False, "message": "Staff only"}), 403

    period = request.args.get("period", "lifetime")
    want_status = request.args.get("status", "all")
    start, end = get_dashboard_period_range(period)

    coll_names = ["Leads", "RentalLeads", "sellingLeads", "agentLeads"]
    end_collection = db["endData"]

    def _clean(v, default="-"):
        if v is None or v == "":
            return default
        if isinstance(v, float) and math.isnan(v):
            return default
        return v

    try:
        # ---- PASS 1: collect period-matching leads + their normalized phones,
        # with ZERO Mongo calls to endData yet ----
        candidate_leads = []   # list of (lead_doc, coll_name, phone)
        phones_needed = set()

        for coll_name in coll_names:
            for lead in db[coll_name].find():
                created_dt = parse_created_at_str(lead.get("Created At"))
                if start is not None:
                    if not created_dt or created_dt < start or created_dt > end:
                        continue

                phone_raw = lead.get("Phone Number", "")
                phone = normalize_number(phone_raw)
                if phone and not phone.startswith("91"):
                    phone = "91" + phone

                candidate_leads.append((lead, coll_name, phone, phone_raw))
                if phone:
                    phones_needed.add(phone)

        # ---- PASS 2: ONE batched query for every endData doc we might need ----
        end_data_map = {}
        if phones_needed:
            for ed in end_collection.find({"Number": {"$in": list(phones_needed)}}):
                end_data_map[ed.get("Number")] = ed

        # ---- PASS 3: filter + build output using the in-memory map ----
        out = []
        for lead, coll_name, phone, phone_raw in candidate_leads:
            ed = end_data_map.get(phone)
            if not ed or not _is_lost_lead(ed):
                continue

            call_status = (ed.get("Call Status") or "").strip()
            customer_response = (ed.get("Customer Response") or "").strip()

            if want_status == "Not Interested":
                if customer_response != "Not Interested":
                    continue
            elif want_status != "all":
                if call_status != want_status:
                    continue

            lead_type = _normalize_lead_type_field(lead.get("LeadType"))

            out.append({
                "id": str(lead["_id"]),
                "collection": coll_name,
                "leadType": lead_type,
                "name": _clean(lead.get("Lead Name") or lead.get("Name"), "Unknown"),
                "phone": phone or _clean(phone_raw, ""),
                "date": _clean(lead.get("Date")),
                "location": _clean(lead.get("Location Interested In") or lead.get("Property Location")),
                "propertyType": _clean(lead.get("Property Type")),
                "budget": _clean(lead.get("Budget Range") or lead.get("Expected Price")),
                "assignedTo": _clean(lead.get("AssignTo"), "Unassigned"),
                "callStatus": _clean(ed.get("Call Status")),
                "customerResponse": _clean(ed.get("Customer Response")),
                "callerRemarks": _clean(ed.get("Caller Remarks")),
                "nextFollowupTimeline": _clean(ed.get("Next Follow-up Timeline")),
                "nextCallDate": _clean(ed.get("Next Call Date")),
                "callAttempts": ed.get("Call_attempt") if isinstance(ed.get("Call_attempt"), int) else 0,
            })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

    return jsonify({"success": True, "period": period, "status": want_status,
                     "count": len(out), "data": out}), 200
    

@app.route("/api/delete-lead", methods=["DELETE"])
def delete_lead():
    try:
        data = request.json

        lead_id = data.get("id")
        collection_name = data.get("collection")

        if not lead_id or not collection_name:
            return jsonify({"error": "Missing id or collection"}), 400

        if collection_name not in [
            "Leads",
            "RentalLeads",
            "sellingLeads",
            "agentLeads",
            "endData"
        ]:
            return jsonify({"error": "Invalid collection"}), 400

        collection = db[collection_name]

        result = collection.delete_one({
            "_id": ObjectId(lead_id)
        })

        if result.deleted_count == 0:
            return jsonify({"error": "Lead not found"}), 404

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

#excel

@app.route("/api/export-leads", methods=["POST"])
def export_leads():
    """
    Builds an Excel export of leads.
    """
    try:
        data = request.json or {}
        leads_input = data.get("leads", [])
        period = data.get("period")                 # "this_month" | "last_month" | "last_3_months" | "all" | None
        lead_type_filter = data.get("type", "all")   # "all" | "buying" | "rental" | "selling" | "agent"

        # Date-range based export (used when no explicit page-selection is sent)
        if not leads_input and period:
            start_date, end_date = get_date_range(period)
            types_to_scan = (
                [lead_type_filter] if lead_type_filter != "all"
                else ["buying", "rental", "selling", "agent"]
            )

            leads_input = []
            for t in types_to_scan:
                collection_name = COLLECTION_MAP.get(t, "Leads")

                try:
                    cursor = db[collection_name].find()
                except Exception as scan_err:
                    print(f"[export] Failed scanning collection '{collection_name}': {scan_err}")
                    continue

                for d in cursor:
                    try:
                        if start_date:
                            lead_date = parse_lead_date(d.get("Date"))
                            if not lead_date or lead_date < start_date or lead_date > end_date:
                                continue
                        leads_input.append({
                            "id": str(d["_id"]),
                            "phone": d.get("Phone Number", ""),
                            "type": t,
                            "name": d.get("Lead Name") or d.get("Name") or ""
                        })
                    except Exception as row_err:
                        print(f"[export] Skipping malformed doc in '{collection_name}': {row_err}")
                        continue

        if not leads_input:
            return jsonify({"error": "No leads found for the selected filters"}), 400

        end_collection = db["endData"]
        rows = []
        max_calls = 0
        skipped = 0

        for item in leads_input:
            try:
                lead_id = item.get("id")
                lead_type = item.get("type", "buying")
                collection_name = COLLECTION_MAP.get(lead_type, "Leads")

                raw_phone = item.get("phone", "")
                phone = normalize_number(raw_phone)

                valid_phone = bool(phone) and len(phone) >= 8
                if valid_phone and not phone.startswith("91"):
                    phone = "91" + phone

                lead_doc = None

                if lead_id:
                    try:
                        lead_doc = db[collection_name].find_one({"_id": ObjectId(lead_id)})
                    except Exception:
                        lead_doc = None

                if not lead_doc and valid_phone:
                    try:
                        safe_phone = re.escape(phone)
                        lead_doc = db[collection_name].find_one(
                            {"Phone Number": {"$regex": safe_phone}}
                        )
                    except Exception:
                        lead_doc = None

                lead_doc = lead_doc or {}

                end_doc = {}
                call_logs = []
                if valid_phone:
                    try:
                        end_doc = end_collection.find_one({"Number": phone}) or {}
                    except Exception as end_err:
                        print(f"[export] end-data lookup failed for {phone}: {end_err}")
                    try:
                        call_logs = list(
                            call_logs_collection.find({"Number": phone}).sort("CreatedAt", 1)
                        )
                    except Exception as log_err:
                        print(f"[export] call-log lookup failed for {phone}: {log_err}")

                max_calls = max(max_calls, len(call_logs))

                rows.append({
                    "name": lead_doc.get("Lead Name") or lead_doc.get("Name") or item.get("name") or "Unknown",
                    "phone": ("+" + phone) if valid_phone else (str(raw_phone) or "-"),
                    "type": str(lead_type).capitalize(),
                    "location": lead_doc.get("Location Interested In") or lead_doc.get("Property Location") or "-",
                    "property": lead_doc.get("Property Type", "-"),
                    "budget": lead_doc.get("Budget Range") or lead_doc.get("Expected Price") or "-",
                    "assigned_to": lead_doc.get("AssignTo", "-"),
                    "call_status": end_doc.get("Call Status", "-"),
                    "interest_level": end_doc.get("Interest Level", "-"),
                    "next_followup": end_doc.get("Next Follow-up Timeline", "-"),
                    "next_call_date": end_doc.get("Next Call Date", "-"),
                    "total_calls": len(call_logs),
                    "calls": call_logs
                })

            except Exception as item_err:
                skipped += 1
                print(f"[export] Skipping row due to error: {item_err}")
                continue

        if not rows:
            return jsonify({"error": "No leads could be exported (all rows failed)"}), 400

        # Build workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads Export"

        base_headers = [
            "Lead Name", "Phone", "Type", "Location", "Property", "Budget",
            "Assigned To", "Current Status", "Interest Level",
            "Next Follow-up", "Next Call Date", "Total Calls"
        ]
        call_headers = []
        for i in range(1, max_calls + 1):
            call_headers += [
                f"Call {i} - Date & Time", f"Call {i} - Status",
                f"Call {i} - Response", f"Call {i} - Remarks"
            ]
        headers = base_headers + call_headers
        ws.append(headers)

        header_fill = PatternFill(start_color="2D3142", end_color="2D3142", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(style="thin", color="D1D5DB")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 28

        accent_fill = PatternFill(start_color="FDF1EB", end_color="FDF1EB", fill_type="solid")
        for r_idx, row in enumerate(rows, start=2):
            base_values = [
                row["name"], row["phone"], row["type"], row["location"], row["property"],
                row["budget"], row["assigned_to"], row["call_status"], row["interest_level"],
                row["next_followup"], row["next_call_date"], row["total_calls"]
            ]
            call_values = []
            for c in row["calls"]:
                call_values += [
                    format_ist(c.get("CreatedAt")),
                    c.get("CallStatus", "-"),
                    c.get("CustomerResponse", "-"),
                    c.get("CallerRemarks", "-")
                ]
            call_values += ["-"] * (len(call_headers) - len(call_values))

            full_row = base_values + call_values
            for col_idx, val in enumerate(full_row, start=1):
                cell = ws.cell(row=r_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if r_idx % 2 == 0:
                    cell.fill = accent_fill

        widths = [22, 16, 10, 18, 16, 14, 16, 20, 14, 18, 14, 10] + [20, 16, 18, 22] * max_calls
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        if skipped:
            print(f"[export] Completed with {skipped} row(s) skipped out of {len(leads_input)} requested")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"Leads_Export_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


#wp templetes

@app.route("/api/wp-template", methods=["POST"])
def create_wp_template():
    try:
        name = request.form.get("name")
        message = request.form.get("message")
        file = request.files.get("media")

        if not name or not message:
            return jsonify({"error": "Name and message required"}), 400

        filename = None
        file_type = None

        if file:
            ext = file.filename.split('.')[-1]
            unique_name = str(int(time.time())) + "_" + secure_filename(file.filename)
            file_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_name)
            file.save(file_path)

            filename = unique_name

            if ext.lower() in ["jpg", "jpeg", "png", "webp"]:
                file_type = "image"
            elif ext.lower() in ["mp4", "mov", "avi"]:
                file_type = "video"
            else:
                file_type = "file"

        data = {
            "name": name,
            "message": message,
            "media": filename,
            "type": file_type,
            "createdAt": datetime.utcnow()
        }

        db.wp.insert_one(data)

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/wp-template", methods=["GET"])
def get_wp_templates():
    try:
        templates = list(db.wp.find().sort("createdAt", -1))

        for t in templates:
            t["_id"] = str(t["_id"])

        return jsonify(templates)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)

@app.route("/api/wp-template/<id>", methods=["DELETE"])
def delete_wp_template(id):
    try:
        template = db.wp.find_one({"_id": ObjectId(id)})

        if not template:
            return jsonify({"error": "Template not found"}), 404

        # Delete file from uploads folder
        if template.get("media"):
            file_path = os.path.join(app.config["UPLOAD_FOLDER"], template["media"])
            if os.path.exists(file_path):
                os.remove(file_path)

        # Delete from MongoDB
        db.wp.delete_one({"_id": ObjectId(id)})

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


update_lead_db = client["NishaHomesData"]
update_lead_collection = update_lead_db["endData"]


@app.route("/update-lead", methods=["POST"])
def update_lead():
    try:
        data = request.get_json()

        if not data or not isinstance(data, list):
            return jsonify({"error": "Invalid input format"}), 400

        output = data[0].get("output", {})
        number = output.get("number")

        if not number:
            return jsonify({"error": "Number is required"}), 400

        # Fields we want to update (excluding _id, Number, Call_attempt)
        update_fields = {
            "Customer Name": output.get("customerName"),
            "Lead Source": output.get("leadSource"),
            "Property Type": output.get("propertyType"),
            "Preferred Location": output.get("preferredLocation"),
            "Budget Range": output.get("budgetRange"),
            "Call Status": output.get("callStatus"),
            "Transaction Type": output.get("transactionType"),
            "Configuration": output.get("configuration"),
            "Customer Response": output.get("customerResponse"),
            "Interest Level": output.get("interestLevel"),
            "Objection / Reason": output.get("objectionReason"),
            "Next Follow-up Timeline": output.get("nextFollowupTimeline"),
            "Caller Remarks": output.get("callerRemarks"),
            "Call Priority": output.get("callPriority"),
            "Next Call Date": output.get("nextCallDate"),
            "done": output.get("done"),
            "Lead type": output.get("leadType"),
            "Lead score": output.get("leadScore"),
            "lastUpdatedAt": datetime.utcnow()
        }

        # Remove None values (clean update)
        update_fields = {k: v for k, v in update_fields.items() if v is not None}

        result = update_lead_collection.update_one(
            {"Number": number},   # Find by Number
            {
                "$set": update_fields,
                "$inc": {"Call_attempt": 1}  # increment safely
            },
            upsert=False  # Do NOT create new document automatically
        )

        if result.matched_count == 0:
            return jsonify({"message": "No document found with this number"}), 404

        return jsonify({"message": "Lead updated successfully"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/check-assign")
def check_assign():
    return "assign route section reached"

#https://www.karmandrones.com/

# Allowed values for the LeadType field on every Leads-family document
VALID_LEAD_TYPES = {"buyer_purchase", "buyer_rental", "seller", "agent"}


@app.route("/add-lead", methods=["POST"])
def add_lead():
    try:
        data = request.json

        # 1. Get collection name dynamically
        collection_name = data.get("collection")
        if not collection_name:
            return jsonify({"error": "Collection name is required"}), 400

        collection = db[collection_name]

        # 2. Extract phone number (required for upsert)
        phone_number = data.get("Phone Number")
        if not phone_number:
            return jsonify({"error": "Phone Number is required"}), 400

        # NEW: normalize/validate LeadType — defaults to "buyer_purchase"
        # if missing or not one of the accepted values.
        lead_type = str(data.get("LeadType", "")).strip()
        if lead_type not in VALID_LEAD_TYPES:
            lead_type = "buyer_purchase"
        data["LeadType"] = lead_type

        # 3. Remove collection key from document
        data.pop("collection", None)

        # NEW: "Created At" timestamp, formatted as "YYYY-MM-DD HH:MM:SS"
        # (e.g. 2026-07-14 13:58:12). Uses $setOnInsert so it's only written
        # once — the first time this lead is created — and never gets
        # overwritten on later upserts/updates to the same phone number.
        data.pop("Created At", None)  # never let incoming payload override it
        now = datetime.utcnow()
        created_at_str = now.strftime("%Y-%m-%d %H:%M:%S")

        # NEW: also write "Date" (DD-MM-YYYY) on first insert — this is the
        # field parse_lead_date()/filter_by_july_range() actually filter on
        # for /api/leads, /api/rental-leads, /api/selling-leads, /api/agent-leads.
        # Without it, leads created via this endpoint were silently excluded
        # from every list even though they existed in Mongo.
        if not str(data.get("Date", "")).strip():
            data["Date"] = now.strftime("%d-%m-%Y")

        # PERF: also write DateObj — a real, indexed datetime — so the fast
        # path in get_filtered_leads() can serve this lead with a single
        # indexed range query instead of falling back to Python parsing.
        data.pop("DateObj", None)  # never let incoming payload override it

        # 4. Upsert (update if exists, insert if not)
        result = collection.update_one(
            {"Phone Number": phone_number},
            {
                "$set": data,
                "$setOnInsert": {"Created At": created_at_str},
                "$currentDate": {},  # placeholder no-op kept for clarity of intent
            },
            upsert=True
        )
        # Set DateObj explicitly right after (covers both insert + update
        # cases uniformly without fighting $setOnInsert semantics above).
        collection.update_one(
            {"Phone Number": phone_number},
            {"$set": {"DateObj": now}}
        )

        name = data.get("Lead Name")

        create_contact(name, phone_number)

        return jsonify({
            "success": True,
            "matched_count": result.matched_count,
            "modified_count": result.modified_count,
            "upserted_id": str(result.upserted_id) if result.upserted_id else None
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/modify-document", methods=["POST"])
def modify_document():
    try:
        data = request.json

        if not data:
            return jsonify({"error": "No JSON body provided"}), 400

        # 1. Validate collection
        collection_name = data.get("collection")
        if not collection_name:
            return jsonify({"error": "Collection name is required"}), 400

        # Optional: restrict collections (recommended)
        allowed_collections = [
            "Leads",
            "RentalLeads",
            "sellingLeads",
            "agentLeads",
            "endData",
            "teamAssign",
            "orderhouseofcakes",
            "tasks",
            "Realtors"
        ]

        if collection_name not in allowed_collections:
            return jsonify({"error": "Invalid collection"}), 400

        collection = db[collection_name]

        # 2. Validate phone / number
        raw_number = (
            data.get("Phone Number") or
            data.get("Number") or
            data.get("Employee number")
        )

        if not raw_number:
            return jsonify({"error": "Phone Number / Number / Employee number is required"}), 400

        normalized_number = normalize_number(raw_number)

        if not normalized_number:
            return jsonify({"error": "Invalid phone number"}), 400

        # 3. Build filter dynamically
        if collection_name == "endData":
            filter_query = {"Number": normalized_number}
        else:
            filter_query = {
                "$or": [
                    {"Phone Number": normalized_number},
                    {"Employee number": normalized_number},
                    {"Number": normalized_number},

                    {"Phone Number": {"$regex": str(normalized_number)}},
                    {"Employee number": {"$regex": str(normalized_number)}},
                    {"Number": {"$regex": str(normalized_number)}},
                ]
            }

        # 4. Build update operations
        update_query = {}

        # SET (add/update fields)
        set_fields = data.get("set", {})
        if set_fields:
            update_query["$set"] = set_fields

        # UNSET (delete fields)
        unset_fields = data.get("unset", [])
        if unset_fields:
            update_query["$unset"] = {field: "" for field in unset_fields}

        # PUSH (append to array)
        push_fields = data.get("push", {})
        if push_fields:
            update_query["$push"] = push_fields

        # INC (increment numbers)
        inc_fields = data.get("inc", {})
        if inc_fields:
            update_query["$inc"] = inc_fields

        if not update_query:
            return jsonify({"error": "No update operations provided"}), 400

        # 5. Perform update (Upsert allowed)
        result = collection.update_one(
            filter_query,
            update_query,
            upsert=True
        )

        # 6. Return updated document
        updated_doc = collection.find_one(filter_query)

        return jsonify({
            "success": True,
            "matched_count": result.matched_count,
            "modified_count": result.modified_count,
            "upserted_id": str(result.upserted_id) if result.upserted_id else None,
            "updated_document": serialize_doc(updated_doc) if updated_doc else None
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/add-task', methods=['POST'])
def add_task():
    data = request.json

    emp = data.get("phone")
    task_text = data.get("task")
    status = data.get("status", "pending")

    if not emp or not task_text:
        return jsonify({"success": False, "message": "Missing data"})

    normalized_number = normalize_number(emp)

    task_id = f"{normalized_number}_{int(time.time()*1000)}"

    task_obj = {
        "id": task_id,
        "text": task_text,
        "status": status,
        "created_at": int(time.time())
    }

    result = db.teamAssign.update_one(
        {
            "$or": [
                {"Employee number": normalized_number},
                {"Employee number": {"$regex": str(normalized_number)}},
            ]
        },
        {
            "$push": {"tasks": task_obj}
        },
        upsert=True
    )

    return jsonify({
        "success": True,
        "message": "Task added",
        "task_id": task_id
    })


@app.route('/update-task-status', methods=['POST'])
def update_task_status():
    data = request.json

    emp = data.get("phone")
    task_id = data.get("task_id")
    status = data.get("status")

    if not emp or not task_id or not status:
        return jsonify({"success": False, "message": "Missing fields"})

    result = db.teamassign.update_one(
        {
            "Phone Number": emp,
            "tasks.id": task_id
        },
        {
            "$set": {
                "tasks.$.status": status
            }
        }
    )

    if result.modified_count:
        return jsonify({"success": True, "message": "Status updated"})
    else:
        return jsonify({"success": False, "message": "Task not found"})


@app.route('/get-tasks/<phone>', methods=['GET'])
def get_tasks(phone):
    user = db.teamassign.find_one({"Phone Number": phone})

    if not user:
        return jsonify([])

    return jsonify(user.get("tasks", []))


# Endpoint 1: Image -> Text
@app.route('/image-to-text', methods=['POST'])
def image_to_text():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']

    temp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
    temp.close()

    file.save(temp.name)

    text = extract_text_from_image(temp.name)

    os.remove(temp.name)  # cleanup

    return jsonify({'text': text})

# Endpoint 2: Video -> Audio (returns HTML player)
@app.route('/video-to-audio', methods=['POST'])
def video_to_audio():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']

    # TEMP VIDEO (not permanent)
    temp_video = tempfile.NamedTemporaryFile(delete=False, suffix=".mp4")
    temp_video.close()
    file.save(temp_video.name)

    # FINAL AUDIO (saved in uploads)
    audio_filename = f"{int(time.time())}.mp3"
    audio_path = os.path.join(app.config["UPLOAD_FOLDER"], audio_filename)

    # Extract audio
    extract_audio_from_video(temp_video.name, audio_path)

    # Delete temp video
    os.remove(temp_video.name)

    # Return usable URL
    return jsonify({
        "audio_url": f"/uploads/{audio_filename}"
    })


@app.route('/get-audio')
def get_audio():
    path = request.args.get('path')
    return send_file(path, mimetype='audio/mpeg')

#project


@app.route("/api/projects", methods=["GET"])
def get_projects():
    try:
        role = session.get("role")
        query = {}

        if role == "partner":
            query = {"ownerNumber": session.get("employee_number")}

        status_filter = request.args.get("status")
        if status_filter:
            query["status"] = status_filter

        # NEW
        kind_filter = request.args.get("kind")
        if kind_filter:
            query["kind"] = kind_filter

        projects = list(projects_collection.find(query).sort("createdAt", -1))

        return jsonify({
            "success": True,
            "data": [serialize_doc(p) for p in projects]
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "data": [], "error": str(e)}), 500

# =============================
# CRM STAGE (extends projects)
# =============================
LIST_STAGES = [
    "Approved", "Customer Shared", "Site Visit Scheduled", "Negotiation",
    "Token Received", "Agreement Done", "Registration Done",
    "Sold", "Rented", "Closed", "Cancelled"
]

@app.route("/api/projects/stage/<project_id>", methods=["POST"])
def update_project_stage(project_id):
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"status": "error", "message": "Staff only"}), 403

    data = request.json or {}
    stage = data.get("stage")
    if stage not in LIST_STAGES:
        return jsonify({"status": "error", "message": "Invalid stage"}), 400

    project = projects_collection.find_one({"_id": ObjectId(project_id)})
    if not project:
        return jsonify({"status": "error", "message": "Project not found"}), 404

    history_entry = {
        "action": f"Stage: {stage}",
        "remark": data.get("remark", ""),
        "at": datetime.utcnow(),
        "by": session.get("employee_name")
    }

    projects_collection.update_one(
        {"_id": ObjectId(project_id)},
        {
            "$set": {"stage": stage, "lastUpdatedAt": datetime.utcnow()},
            "$push": {"history": history_entry}
        }
    )

    updated = projects_collection.find_one({"_id": ObjectId(project_id)})
    return jsonify({"status": "success", "data": serialize_doc(updated)}), 200


# =============================
# REQUIREMENTS DESK
# =============================
requirements_collection = db["requirements"]

REQ_STATUS_LABELS = {
    "new": "New", "broadcasted": "Broadcasted", "responses": "Responses",
    "matched": "Matched", "visit": "Visit scheduled", "closed": "Closed",
    "cancelled": "Cancelled", "expired": "Expired", "rejected": "Rejected"
}


@app.route("/api/requirements", methods=["GET"])
def get_requirements():
    if not session.get("user_id"):
        return jsonify({"success": False, "data": []}), 401

    role = session.get("role")
    num = session.get("employee_number")
    query = {}
    if role == "partner":
        query = {"$or": [
            {"submittedByNumber": num},
            {"broadcastTo": "all"},
            {"broadcastTo": num}
        ]}

    docs = list(requirements_collection.find(query).sort("createdAt", -1))
    result = []
    for d in docs:
        d = serialize_doc(d)
        if role == "partner" and d.get("submittedByNumber") != num and not d.get("showClientContact"):
            d.pop("clientName", None)
            d.pop("clientMobile", None)
        result.append(d)
    return jsonify({"success": True, "data": result}), 200


@app.route("/api/requirements/add", methods=["POST"])
def add_requirement():
    if not session.get("user_id"):
        return jsonify({"success": False, "message": "Login required"}), 401

    data = request.json or {}
    role = session.get("role")

    if role == "partner":
        submitted_by_number = session.get("employee_number")
        submitted_by_name = session.get("employee_name")
    else:
        submitted_by_number = data.get("onBehalfNumber") or session.get("employee_number")
        submitted_by_name = data.get("onBehalfName") or session.get("employee_name")

    location = (data.get("location") or "").strip()
    if not location:
        return jsonify({"success": False, "message": "Location is required"}), 400

    doc = {
        "reqType": data.get("reqType", "Buy"),
        "propertyType": data.get("propertyType", ""),
        "config": data.get("config", ""),
        "location": location,
        "budgetMin": data.get("budgetMin", ""),
        "budgetMax": data.get("budgetMax", ""),
        "areaMin": data.get("areaMin", ""),
        "areaMax": data.get("areaMax", ""),
        "furnishing": data.get("furnishing", ""),
        "possession": data.get("possession", ""),
        "notes": data.get("notes", ""),
        "priority": data.get("priority", "Medium"),
        "priority": data.get("priority", "Medium"),
        "clientName": data.get("clientName", ""),
        "clientMobile": data.get("clientMobile", ""),
        "showClientContact": bool(data.get("showClientContact", False)),   # <-- ADD THIS LINE
        "submittedByNumber": submitted_by_number,
        "submittedByNumber": submitted_by_number,
        "submittedByName": submitted_by_name,
        "status": "new",
        "broadcastTo": [],
        "responses": [],
        "history": [{
            "action": "Submitted", "remark": "",
            "at": datetime.utcnow(), "by": submitted_by_name
        }],
        "createdAt": datetime.utcnow()
    }

    result = requirements_collection.insert_one(doc)
    return jsonify({"success": True, "id": str(result.inserted_id)}), 201


@app.route("/api/requirements/status/<req_id>", methods=["POST"])
def update_requirement_status(req_id):
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False, "message": "Staff only"}), 403

    data = request.json or {}
    status = data.get("status")
    if status not in REQ_STATUS_LABELS:
        return jsonify({"success": False, "message": "Invalid status"}), 400

    req_doc = requirements_collection.find_one({"_id": ObjectId(req_id)})
    if not req_doc:
        return jsonify({"success": False, "message": "Not found"}), 404

    requirements_collection.update_one(
        {"_id": ObjectId(req_id)},
        {
            "$set": {"status": status},
            "$push": {"history": {
                "action": REQ_STATUS_LABELS.get(status, status),
                "remark": data.get("remark", ""),
                "at": datetime.utcnow(),
                "by": session.get("employee_name")
            }}
        }
    )
    return jsonify({"success": True}), 200


@app.route("/api/requirements/broadcast/<req_id>", methods=["POST"])
def broadcast_requirement(req_id):
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False, "message": "Staff only"}), 403

    data = request.json or {}
    to = data.get("to")  # "all" or a list of employee numbers

    req_doc = requirements_collection.find_one({"_id": ObjectId(req_id)})
    if not req_doc:
        return jsonify({"success": False, "message": "Not found"}), 404

    if to == "all":
        broadcast_to = "all"
    else:
        existing = req_doc.get("broadcastTo", [])
        if existing == "all":
            existing = []
        existing = list(set(existing + (to or [])))
        broadcast_to = existing

    new_status = "broadcasted" if req_doc.get("status") == "new" else req_doc.get("status")

    requirements_collection.update_one(
        {"_id": ObjectId(req_id)},
        {
            "$set": {"broadcastTo": broadcast_to, "status": new_status},
            "$push": {"history": {
                "action": "Broadcasted", "remark": "",
                "at": datetime.utcnow(), "by": session.get("employee_name")
            }}
        }
    )
    return jsonify({"success": True}), 200


@app.route("/api/requirements/respond/<req_id>", methods=["POST"])
def respond_requirement(req_id):
    if session.get("role") != "partner":
        return jsonify({"success": False, "message": "Partner only"}), 403

    data = request.json or {}
    partner_number = session.get("employee_number")
    partner_name = session.get("employee_name")

    resp = {
        "partnerNumber": partner_number,
        "partnerName": partner_name,
        "type": data.get("type", "Need More Details"),
        "at": datetime.utcnow(),
        "property": data.get("property", {})
    }

    req_doc = requirements_collection.find_one({"_id": ObjectId(req_id)})
    if not req_doc:
        return jsonify({"success": False, "message": "Not found"}), 404

    responses = [r for r in req_doc.get("responses", []) if r.get("partnerNumber") != partner_number]
    responses.append(resp)

    new_status = "responses" if req_doc.get("status") == "broadcasted" else req_doc.get("status")

    requirements_collection.update_one(
        {"_id": ObjectId(req_id)},
        {
            "$set": {"responses": responses, "status": new_status},
            "$push": {"history": {
                "action": f"{partner_name}: {resp['type']}",
                "remark": "", "at": datetime.utcnow(), "by": partner_name
            }}
        }
    )
    return jsonify({"success": True}), 200


@app.route("/api/requirements/delete/<req_id>", methods=["DELETE"])
def delete_requirement(req_id):
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False, "message": "Staff only"}), 403
    requirements_collection.delete_one({"_id": ObjectId(req_id)})
    return jsonify({"success": True}), 200


# =============================
# PARTNER MANAGEMENT (extends teamAssign)
# =============================
@app.route("/api/toggle-team-active/<number>", methods=["POST"])
def toggle_team_active(number):
    if session.get("role") != "admin":
        return jsonify({"success": False, "message": "Admin only"}), 403

    try:
        number = int(normalize_number(number))
    except ValueError:
        return jsonify({"success": False, "message": "Invalid number"}), 400

    collection = db["teamAssign"]
    member = collection.find_one({"Employee number": number})
    if not member:
        return jsonify({"success": False, "message": "Not found"}), 404

    new_active = not member.get("Active", True)
    collection.update_one({"_id": member["_id"]}, {"$set": {"Active": new_active}})
    return jsonify({"success": True, "active": new_active}), 200


@app.route("/api/update-team-areas", methods=["POST"])
def update_team_areas():
    if session.get("role") != "admin":
        return jsonify({"success": False, "message": "Admin only"}), 403

    data = request.json or {}
    try:
        number = int(normalize_number(str(data.get("number", ""))))
    except ValueError:
        return jsonify({"success": False, "message": "Invalid number"}), 400

    db["teamAssign"].update_one(
        {"Employee number": number},
        {"$set": {"Areas": data.get("areas", "")}}
    )
    return jsonify({"success": True}), 200


# =============================
# SETTINGS (corporate/agent share settings)
# =============================
settings_collection = db["settings"]

@app.route("/api/settings", methods=["GET"])
def get_settings_api():
    if not session.get("user_id"):
        return jsonify({"success": False}), 401
    # PERF: settings almost never change — cache for 30s to skip the
    # round-trip on every dashboard/page load that reads it.
    doc = cached("global_settings", 30, lambda: settings_collection.find_one({"_id": "global"}) or {})
    doc = dict(doc)
    doc.pop("_id", None)
    return jsonify({"success": True, "data": doc}), 200


@app.route("/api/settings", methods=["POST"])
def save_settings_api():
    if session.get("role") != "admin":
        return jsonify({"success": False, "message": "Admin only"}), 403
    data = request.json or {}
    fields = {k: data.get(k, "") for k in ["corporate", "agent", "advisorName", "website", "landing", "cta"]}
    settings_collection.update_one({"_id": "global"}, {"$set": fields}, upsert=True)
    invalidate_cache("global_settings")  # PERF: bust the cache immediately on save
    return jsonify({"success": True}), 200


# =============================
# NEW: WHATSAPP SHARE TEXT BUILDER
# =============================
@app.route("/api/projects/share-text/<project_id>", methods=["GET"])
def project_share_text(project_id):
    try:
        p = projects_collection.find_one({"_id": ObjectId(project_id)})
        if not p:
            return jsonify({"success": False, "message": "Not found"}), 404
        s = settings_collection.find_one({"_id": "global"}) or {}
        return jsonify({"success": True, "text": build_whatsapp_share_text(p, s)})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


def build_whatsapp_share_text(p, settings):
    base_url = os.getenv("PUBLIC_BASE_URL", "https://crm.nishahomes.com")
    view_url = f"{base_url}/view/{p.get('uniqueId', '')}"
    L = [
        p.get("name") or p.get("propertyTitle") or "",
        f"📍 {p.get('location') or p.get('locality') or ''}",
        f"💰 {p.get('budget') or p.get('startingPrice') or ''}",
        "",
        "🔗 See complete details of this property here:",
        view_url,
        "",
        "🏡 Nisha Homes — Your Trusted Real Estate Advisor",
        f"💬 {BRAND_CONTACT_NUMBER}",
        f"🌐 https://{BRAND_WEBSITE}",
    ]
    return "\n".join(L)

# =============================
# COORDINATOR DASHBOARD STATS
# =============================
@app.route("/api/inventory-dashboard-stats", methods=["GET"])
def inventory_dashboard_stats():
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False}), 403

    def _compute():
        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        closed_stages = {"Sold", "Rented", "Closed", "Cancelled"}

        all_projects = list(projects_collection.find())
        all_reqs = list(requirements_collection.find())
        all_partners = list(db["teamAssign"].find({"roll": "partner"}))

        today_inventory = sum(1 for p in all_projects if p.get("createdAt") and p["createdAt"] >= today_start)
        today_requirements = sum(1 for r in all_reqs if r.get("createdAt") and r["createdAt"] >= today_start)
        pending_inventory = sum(1 for p in all_projects if p.get("status") == "pending")
        pending_requirements = sum(1 for r in all_reqs if r.get("status") == "new")
        live_stock = sum(1 for p in all_projects if p.get("status") == "approved")
        sold = sum(1 for p in all_projects if p.get("stage") == "Sold")
        rented = sum(1 for p in all_projects if p.get("stage") == "Rented")
        visits = (sum(1 for p in all_projects if p.get("stage") == "Site Visit Scheduled") +
                  sum(1 for r in all_reqs if r.get("status") == "visit"))
        inventory_closed = sum(1 for p in all_projects if p.get("stage") in closed_stages)
        requirements_closed = sum(1 for r in all_reqs if r.get("status") in ("closed", "matched"))

        perf = []
        for p in all_partners:
            num = p.get("Employee number")
            p_inv = [x for x in all_projects if x.get("ownerNumber") == num]
            p_req = [x for x in all_reqs if x.get("submittedByNumber") == num]
            approved = sum(1 for x in p_inv if x.get("status") == "approved" or x.get("stage") in closed_stages)
            deals = sum(1 for x in p_inv if x.get("stage") in ("Sold", "Rented"))
            conv = round(deals * 100 / len(p_inv)) if p_inv else 0
            perf.append({
                "name": p.get("Employee name"),
                "inventory": len(p_inv),
                "requirements": len(p_req),
                "approved": approved,
                "deals": deals,
                "conversion": conv
            })
        perf.sort(key=lambda x: (-x["deals"], -x["inventory"]))

        return {
            "success": True,
            "todayInventory": today_inventory,
            "todayRequirements": today_requirements,
            "pendingInventory": pending_inventory,
            "pendingRequirements": pending_requirements,
            "liveStock": live_stock,
            "sold": sold,
            "rented": rented,
            "visits": visits,
            "inventoryClosed": inventory_closed,
            "requirementsClosed": requirements_closed,
            "partnerPerformance": perf
        }

    # PERF: this endpoint scans 3 whole collections + does O(n*m) partner
    # matching in Python — cache the computed result briefly so rapid
    # dashboard refreshes/tab-switches don't recompute it every time.
    result = cached("inventory_dashboard_stats", 15, _compute)
    return jsonify(result), 200


# =============================
# AI: SHARED GENERATE-PROPERTY ENDPOINT
# =============================
@app.route("/api/ai/generate-property", methods=["POST"])
def ai_generate_property():
    if not session.get("user_id"):
        return jsonify({"success": False, "message": "Login required"}), 401

    try:
        form_type = request.form.get("form_type", "inventory")
        combined_text = []

    

        # NEW: reference screenshot(s) — OCR'd for context only, NEVER uploaded/saved
        # to Cloudinary or Mongo. Purely used to extract extra text for Mistral.
        for f in request.files.getlist("screenshot"):
            if not f or not f.filename:
                continue
            ext = os.path.splitext(f.filename)[-1] or ".png"
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
            tmp.close()
            f.save(tmp.name)
            try:
                txt = extract_text_from_image(tmp.name)
                if txt:
                    combined_text.append(txt)
            except Exception as scr_err:
                print(f"[ai-generate] screenshot OCR failed: {scr_err}")
            finally:
                os.remove(tmp.name)

        pdf_file = request.files.get("pdf")
        if pdf_file and pdf_file.filename:
            tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            tmp_pdf.close()
            pdf_file.save(tmp_pdf.name)
            try:
                txt = extract_text_from_pdf(tmp_pdf.name)
                if txt:
                    combined_text.append(txt)
            finally:
                os.remove(tmp_pdf.name)

        full_text = "\n\n".join(combined_text).strip()
        if not full_text:
            return jsonify({"success": False, "message": "Could not extract any text from the uploaded photos/PDF"}), 400

        schema = INVENTORY_FIELDS_SCHEMA if form_type == "inventory" else PROJECT_FIELDS_SCHEMA
        result = call_mistral_generate(full_text, schema)

        return jsonify({"success": True, "data": result})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


# -------------------------------
# POST /api/projects/upload
# -------------------------------
@app.route("/api/projects/upload", methods=["POST"])
def upload_project():
    try:
        if not session.get("user_id"):
            return jsonify({"status": "error", "message": "Login required"}), 401

        name = request.form.get("name")
        location = request.form.get("location")
        description = request.form.get("description")
        budget = request.form.get("budget")
        category = request.form.get("category")
        file = request.files.get("media")

        # Validation
        if not all([name, location, description, budget, category, file]):
            return jsonify({
                "status": "error",
                "message": "All fields required"
            }), 400

        # Detect image vs video from extension
        # Detect image vs video from extension
        ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
        video_exts = {"mp4", "mov", "avi", "webm", "mkv"}
        resource_type = "video" if ext in video_exts else "image"

        # CHANGED: was Cloudinary — now Supabase Storage
        file_url, public_id = upload_media_to_supabase(
            file.read(), folder="projects", resource_type=resource_type, ext=ext or "jpg"
        )

        # PDF / brochure upload for the project — now goes to Supabase Storage
        pdf_url = None
        pdf_file = request.files.get("pdf")
        if pdf_file and pdf_file.filename:
            pdf_url = upload_pdf_to_supabase(pdf_file, folder="projects")

        # Save in DB
        project_data = {
            "name": name,
            "location": location,
            "description": description,
            "budget": budget,
            "category": category,
            "img": file_url,
            "mediaUrl": file_url,
            "mediaPublicId": public_id,
            "type": resource_type,
            "pdfUrl": pdf_url,
            "status": "pending" if session.get("role") == "partner" else "approved",
            "ownerNumber": session.get("employee_number"),
            "ownerName": session.get("employee_name"),
            "ownerRole": session.get("role"),
            "createdAt": datetime.utcnow()
        }

        result = projects_collection.insert_one(project_data)
        invalidate_cache("inventory_dashboard_stats")

        return jsonify({
            "status": "success",
            "message": "Project uploaded successfully",
            "id": str(result.inserted_id),
            "url": file_url
        }), 201

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


# -------------------------------
# PARTNER "Add Inventory" upload
# (also used by admin/emp via the same modal — branding + screenshot-OCR aware)
# -------------------------------
@app.route("/api/projects/upload-inventory", methods=["POST"])
def upload_inventory():
    try:
        if not session.get("user_id"):
            return jsonify({"status": "error", "message": "Login required"}), 401

        f = request.form.get
        required = ["propertyType", "propertyTitle", "locality", "price"]
        if not all(f(k) for k in required):
            return jsonify({"status": "error", "message": "Property type, title, locality and price are required"}), 400

        branding = request.form.get("branding") == "true"
        brand_fields = {
            "dealType": f("dealType", ""), "propertyTitle": f("propertyTitle", ""),
            "locality": f("locality", ""), "price": f("price", ""),
            "configuration": f("configuration", ""), "superArea": f("superArea", "")
        }

        media_urls, media_public_ids = [], []
        banner_url, banner_public_id = None, None
        failed_files = []
        photo_files = [f for f in request.files.getlist("photos") if f and f.filename]

        for idx, file in enumerate(photo_files):
            try:
                raw = file.read()
                if len(raw) > MAX_MEDIA_UPLOAD_BYTES:
                    print(f"[upload-inventory] photo '{file.filename}' rejected — over 50MB")
                    failed_files.append(file.filename)
                    continue

                if idx == 0:
                    # BANNER photo — ALWAYS fit into the 1080x1080 WhatsApp
                    # square, no cropping, regardless of the branding toggle.
                    if branding:
                        brand_fn = lambda b, _fields=brand_fields: fit_to_whatsapp_square(
                            build_banner_image(b, _fields).getvalue()
                        )
                    else:
                        brand_fn = lambda b: fit_to_whatsapp_square(b)
                else:
                    brand_fn = (lambda b: build_simple_branded_image(b)) if branding else None

                url, object_path, _ = upload_raw_then_brand(
                    raw, folder="inventory", resource_type="image", ext="jpg", brand_fn=brand_fn
                )
                media_urls.append(url)
                media_public_ids.append(object_path)
                if idx == 0:
                    banner_url, banner_public_id = url, object_path
            except Exception as photo_err:
                print(f"[upload-inventory] photo '{file.filename}' failed: {photo_err}")
                failed_files.append(file.filename)

        for file in request.files.getlist("videos"):
            if not file or not file.filename:
                continue
            try:
                ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "mp4"
                raw = file.read()
                if len(raw) > MAX_MEDIA_UPLOAD_BYTES:
                    print(f"[upload-inventory] video '{file.filename}' rejected — over 50MB")
                    failed_files.append(file.filename)
                    continue

                brand_fn = build_simple_branded_video if branding else None

                url, object_path, _ = upload_raw_then_brand(
                    raw, folder="inventory", resource_type="video", ext=ext, brand_fn=brand_fn
                )
                media_urls.append(url)
                media_public_ids.append(object_path)
            except Exception as video_err:
                print(f"[upload-inventory] video '{file.filename}' failed: {video_err}")
                failed_files.append(file.filename)

        pdf_url = None
        pdf_file = request.files.get("pdf")
        if pdf_file and pdf_file.filename:
            pdf_url = upload_pdf_to_supabase(pdf_file, folder="inventory")

        inventory_data = {
            "listingBasis": f("listingBasis", ""),
            "dealType": f("dealType", ""),
            "category": f("propertyType", ""),
            "name": f("propertyTitle", ""),
            "location": f("locality", ""),
            "configuration": f("configuration", ""),
            "furnishing": f("furnishing", ""),
            "areaUnit": f("areaUnit", "sqft"),
            "carpetArea": f("carpetArea", ""),
            "superArea": f("superArea", ""),
            "floor": f("floor", ""),
            "bathrooms": f("bathrooms", ""),
            "facing": f("facing", ""),
            "parking": f("parking", ""),
            "possession": f("possession", ""),
            "budget": f("price", ""),
            "landingPageLink": f("landingPageLink", ""),
            "mapLink": f("mapLink", ""),
            "quickNotes": f("quickNotes", ""),
            "description": f("description", ""),
            "img": media_urls[0] if media_urls else None,
            "bannerUrl": banner_url or (media_urls[0] if media_urls else None),
            "mediaUrls": media_urls,
            "mediaPublicIds": media_public_ids,
            "pdfUrl": pdf_url,
            "type": "inventory",
            "uniqueId": generate_unique_id(),
            "status": "pending" if session.get("role") == "partner" else "approved",
            "ownerNumber": session.get("employee_number"),
            "ownerName": session.get("employee_name"),
            "ownerRole": session.get("role"),
            "createdAt": datetime.utcnow()
        }

        embed_and_attach(inventory_data)
        result = projects_collection.insert_one(inventory_data)
        invalidate_cache("inventory_dashboard_stats")
        return jsonify({
            "status": "success",
            "id": str(result.inserted_id),
            "failedFiles": failed_files
        }), 201

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

# PUT/POST-style update - allows editing fields and/or replacing media
@app.route("/api/projects/update/<project_id>", methods=["POST"])
def update_project(project_id):
    try:
        if not session.get("user_id"):
            return jsonify({"status": "error", "message": "Login required"}), 401

        query = {"_id": ObjectId(project_id)}
        # Partners can only edit their own listings
        if session.get("role") == "partner":
            query["ownerNumber"] = session.get("employee_number")

        project = projects_collection.find_one(query)
        if not project:
            return jsonify({"status": "error", "message": "Project not found or not yours"}), 404

        update_fields = {}
        for field in ["name", "location", "description", "budget", "category"]:
            val = request.form.get(field)
            if val:
                update_fields[field] = val

        file = request.files.get("media")
        if file and file.filename:
            ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
            video_exts = {"mp4", "mov", "avi", "webm", "mkv"}
            resource_type = "video" if ext in video_exts else "image"

            # Remove old asset from Supabase before uploading the new one
            old_object_path = project.get("mediaPublicId")
            if old_object_path:
                delete_media_from_supabase(old_object_path)

            new_url, new_object_path = upload_media_to_supabase(
                file.read(), folder="projects", resource_type=resource_type, ext=ext or "jpg"
            )
            update_fields["img"] = new_url
            update_fields["mediaUrl"] = new_url
            update_fields["mediaPublicId"] = new_object_path
            update_fields["type"] = resource_type

        if not update_fields:
            return jsonify({"status": "error", "message": "No fields to update"}), 400

        update_fields["lastUpdatedAt"] = datetime.utcnow()
        merged_preview = {**project, **update_fields}
        update_fields["embedding"] = get_embedding(build_inventory_embedding_text(merged_preview))
        projects_collection.update_one({"_id": ObjectId(project_id)}, {"$set": update_fields})
        invalidate_cache("inventory_dashboard_stats")

        updated = projects_collection.find_one({"_id": ObjectId(project_id)})
        return jsonify({"status": "success", "data": serialize_doc(updated)}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


# delete a project (and its Cloudinary asset)
@app.route("/api/projects/delete/<project_id>", methods=["DELETE"])
def delete_project(project_id):
    try:
        if not session.get("user_id"):
            return jsonify({"status": "error", "message": "Login required"}), 401

        query = {"_id": ObjectId(project_id)}
        # Partners can only delete their own listings
        if session.get("role") == "partner":
            query["ownerNumber"] = session.get("employee_number")

        project = projects_collection.find_one(query)
        if not project:
            return jsonify({"status": "error", "message": "Project not found or not yours"}), 404

        # CHANGED: was Cloudinary destroy — now Supabase delete.
        # Also cleans up mediaPublicIds (plural) for multi-photo inventory
        # / project-v2 listings, which the old Cloudinary code never cleaned up.
        object_path = project.get("mediaPublicId")
        if object_path:
            delete_media_from_supabase(object_path)

        for op in (project.get("mediaPublicIds") or []):
            delete_media_from_supabase(op)

        projects_collection.delete_one({"_id": ObjectId(project_id)})
        invalidate_cache("inventory_dashboard_stats")
        return jsonify({"status": "success", "message": "Project deleted"}), 200

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500



# Approve / send-back a pending listing. Admin only.
@app.route("/api/projects/approve/<project_id>", methods=["POST"])
def approve_project(project_id):
    try:
        if not session.get("user_id") or session.get("role") != "admin":
            return jsonify({"status": "error", "message": "Admin login required"}), 403

        project = projects_collection.find_one({"_id": ObjectId(project_id)})
        if not project:
            return jsonify({"status": "error", "message": "Project not found"}), 404

        action = (request.json or {}).get("action", "approve")  # "approve" | "reject"
        new_status = "approved" if action == "approve" else "pending"

        projects_collection.update_one(
            {"_id": ObjectId(project_id)},
            {"$set": {
                "status": new_status,
                "reviewedBy": session.get("employee_name"),
                "reviewedAt": datetime.utcnow()
            }}
        )
        invalidate_cache("inventory_dashboard_stats")

        updated = projects_collection.find_one({"_id": ObjectId(project_id)})
        return jsonify({"status": "success", "data": serialize_doc(updated)}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500



#new
@app.route("/view/<unique_id>")
def view_property(unique_id):
    p = projects_collection.find_one({"uniqueId": unique_id, "status": "approved"})
    if not p:
        return render_template("property_not_found.html"), 404
    return render_template("view_property.html", unique_id=unique_id)


@app.route("/api/projects/by-unique/<unique_id>")
def get_project_by_unique(unique_id):
    p = projects_collection.find_one({"uniqueId": unique_id, "status": "approved"})
    if not p:
        return jsonify({"success": False, "message": "Not found"}), 404

    data = {
        "name": p.get("name") or p.get("propertyTitle"),
        "location": p.get("location") or p.get("locality"),
        "category": p.get("category") or p.get("propertyType"),
        "configuration": p.get("configuration"),
        "furnishing": p.get("furnishing"),
        "budget": p.get("budget") or p.get("startingPrice"),
        "description": p.get("description"),
        "possession": p.get("possession"),
        "carpetArea": p.get("carpetArea"),
        "superArea": p.get("superArea"),
        "bathrooms": p.get("bathrooms"),
        "facing": p.get("facing"),
        "parking": p.get("parking"),
        "floor": p.get("floor"),
        "dealType": p.get("dealType"),
        "bannerUrl": p.get("bannerUrl") or p.get("img") or p.get("mediaUrl"),
        "mediaUrls": p.get("mediaUrls") or ([p.get("mediaUrl")] if p.get("mediaUrl") else []),
        "pdfUrl": p.get("pdfUrl"),
        "type": p.get("type"),
        "contact": BRAND_CONTACT_NUMBER,
    }
    return jsonify({"success": True, "data": data}), 200



@app.route("/add-inventory")
def add_inventory_page():
    if not session.get("user_id"):
        return redirect("/")
    return render_template("add_inventory.html")


@app.route("/api/add-end-data", methods=["POST"])
def add_end_data():
    try:
        data = request.json

        if not data:
            return jsonify({"error": "No JSON body provided"}), 400

        collection_name = data.get("collection")

        if collection_name != "endData":
            return jsonify({"error": "Invalid collection"}), 400

        collection = db["endData"]

        raw_number = data.get("Number")

        if not raw_number:
            return jsonify({"error": "Number is required"}), 400

        # Normalize number
        number = normalize_number(raw_number)

        if not number:
            return jsonify({"error": "Invalid number"}), 400

        # Ensure Indian format (add 91 if missing)
        if not number.startswith("91"):
            number = "91" + number

        # Clean payload
        data.pop("collection", None)
        data["Number"] = number

        # Remove empty / None values (IMPORTANT)
        clean_data = {k: v for k, v in data.items() if v not in [None, "", []]}

        # Add timestamp
        clean_data["lastUpdatedAt"] = datetime.utcnow()

        # Check if record exists
        existing_doc = collection.find_one({"Number": number})

        if existing_doc:
            # Update ONLY provided fields
            update_query = {
                "$set": clean_data,
                "$inc": {"Call_attempt": int(data.get("Call_attempt", 1))}
            }

            collection.update_one({"Number": number}, update_query)

            message = "Data updated successfully"

        else:
            # Insert new document
            clean_data["Call_attempt"] = int(data.get("Call_attempt", 1))

            collection.insert_one(clean_data)

            message = "Data inserted successfully"

        # Fetch updated doc
        updated_doc = collection.find_one({"Number": number})

        return jsonify({
            "success": True,
            "message": message,
            "data": serialize_doc(updated_doc)
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/get-lead-by-number", methods=["POST"])
def get_lead_by_number():
    collection = db["endData"]

    data = request.get_json()

    if not data or ("number" not in data and "leadId" not in data):
        return jsonify({"error": "number or leadId is required"}), 400

    lead_id = data.get("leadId")
    number = data.get("number")

    lead = None

    # Try matching by LeadId (the Mongo _id of the Lead doc) first
    if lead_id:
        lead = collection.find_one({"LeadId": str(lead_id)})

    # Fallback: match through number (91number), same as before
    if not lead and number:
        normalized = normalize_number(number)
        if normalized and not normalized.startswith("91"):
            normalized = "91" + normalized

        try:
            lead = collection.find_one({"Number": int(normalized)})
        except:
            lead = collection.find_one({"Number": normalized})

        if not lead:
            try:
                lead = collection.find_one({"Number": int(number)})
            except:
                pass
        if not lead:
            lead = collection.find_one({"Number": str(number)})

    if lead:
        return jsonify({
            "success": True,
            "data": serialize_doc(lead)
        })

    return jsonify({
        "success": False,
        "error": "Lead not found"
    }), 404


# =============================
# DAI (Disable-AI) collection
# =============================

@app.route("/api/toggle-ai", methods=["POST"])
def toggle_ai():
    """
    Toggles whether AI is disabled for a lead.
    """
    try:
        data = request.json or {}
        lead_id = data.get("leadId")
        phone = data.get("phone")
        name = data.get("name")

        if not lead_id:
            return jsonify({"success": False, "message": "leadId is required"}), 400

        lead_id = str(lead_id)
        existing = dai_collection.find_one({"leadId": lead_id})

        if existing:
            dai_collection.delete_one({"leadId": lead_id})
            return jsonify({
                "success": True,
                "disabled": False,
                "message": "AI re-enabled for this lead"
            })
        else:
            normalized_phone = normalize_phone_91(phone)

            doc = {
                "leadId": lead_id,
                "Phone Number": normalized_phone,
                "Lead Name": name or "",
                "createdAt": datetime.utcnow()
            }
            dai_collection.insert_one(doc)
            return jsonify({
                "success": True,
                "disabled": True,
                "message": "AI disabled for this lead"
            })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/dai-list", methods=["GET"])
def get_dai_list():
    """Returns every lead currently DAI-flagged (AI disabled)."""
    try:
        docs = list(dai_collection.find())
        return jsonify([serialize_doc(d) for d in docs])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/fix-dai-numbers", methods=["POST"])
def fix_dai_numbers():
    """ONE-TIME cleanup: re-normalizes every existing DAI doc's 'Phone
    Number' field to the clean '91XXXXXXXXXX' shape using
    normalize_phone_91(). Safe to re-run — a doc already in the correct
    shape is simply left unchanged (matched_count 0)."""
    if session.get("role") != "admin":
        return jsonify({"success": False, "message": "Admin only"}), 403
    try:
        fixed = 0
        for doc in dai_collection.find():
            old_phone = doc.get("Phone Number", "")
            new_phone = normalize_phone_91(old_phone)
            if new_phone and new_phone != old_phone:
                dai_collection.update_one({"_id": doc["_id"]}, {"$set": {"Phone Number": new_phone}})
                fixed += 1
        return jsonify({"success": True, "fixed": fixed}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


def remove_assign_to_from_leads():
    """
    ONE-TIME CLEANUP: strips AssignTo (and related assignment fields)
    from every document in the Leads collection.
    Run once, then delete this function and its call below.
    """
    result = db["Leads"].update_many(
        {},
        {
            "$unset": {
                "AssignTo": "",
                "AssignToNumber": "",
                "AssignedBy": "",
                "AssignedByNumber": "",
                "AssignedAt": ""
            }
        }
    )
    print(f"[cleanup] matched: {result.matched_count}, modified: {result.modified_count}")



# =============================
# ADMIN: "Manage Project" — Add Project (multi-photo/video, AI-fillable)
# Saves into the SAME projects_collection as inventory, tagged kind="project"
# Branding + screenshot-OCR aware, same as upload_inventory
# =============================
@app.route("/api/projects/upload-project", methods=["POST"])
def upload_project_v2():
    try:
        if not session.get("user_id"):
            return jsonify({"status": "error", "message": "Login required"}), 401

        f = request.form.get
        required = ["name", "location", "propertyType"]
        if not all(f(k) for k in required):
            return jsonify({"status": "error", "message": "Project name, location and property type are required"}), 400

        video_links_raw = f("videoLinks", "") or ""
        video_links = [v.strip() for v in video_links_raw.splitlines() if v.strip()]

        branding = request.form.get("branding") == "true"
        brand_fields = {
            "dealType": "New Launch", "propertyTitle": f("name", ""),
            "locality": f("location", ""), "price": f("startingPrice", ""),
            "configuration": f("configuration", ""), "superArea": ""
        }

        media_urls, media_public_ids = [], []
        has_image, has_video = False, False
        failed_files = []

        banner_url = None
        photo_files = [f for f in request.files.getlist("photos") if f and f.filename]
        for idx, file in enumerate(photo_files):
            try:
                raw = file.read()
                if len(raw) > MAX_MEDIA_UPLOAD_BYTES:
                    print(f"[upload-project] photo '{file.filename}' rejected — over 50MB")
                    failed_files.append(file.filename)
                    continue

                if idx == 0:
                    # BANNER photo — ALWAYS fit into the 1080x1080 WhatsApp
                    # square, no cropping, regardless of the branding toggle.
                    if branding:
                        brand_fn = lambda b, _fields=brand_fields: fit_to_whatsapp_square(
                            build_banner_image(b, _fields).getvalue()
                        )
                    else:
                        brand_fn = lambda b: fit_to_whatsapp_square(b)
                else:
                    brand_fn = (lambda b: build_simple_branded_image(b)) if branding else None

                url, object_path, _ = upload_raw_then_brand(
                    raw, folder="projects", resource_type="image", ext="jpg", brand_fn=brand_fn
                )
                media_urls.append(url)
                media_public_ids.append(object_path)
                if idx == 0:
                    banner_url = url
                has_image = True
            except Exception as photo_err:
                print(f"[upload-project] photo '{file.filename}' failed: {photo_err}")
                failed_files.append(file.filename)

        for file in request.files.getlist("videos"):
            if not file or not file.filename:
                continue
            try:
                ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "mp4"
                raw = file.read()
                if len(raw) > MAX_MEDIA_UPLOAD_BYTES:
                    print(f"[upload-project] video '{file.filename}' rejected — over 50MB")
                    failed_files.append(file.filename)
                    continue

                brand_fn = build_simple_branded_video if branding else None

                url, object_path, _ = upload_raw_then_brand(
                    raw, folder="projects", resource_type="video", ext=ext, brand_fn=brand_fn
                )
                media_urls.append(url)
                media_public_ids.append(object_path)
                has_video = True
            except Exception as video_err:
                print(f"[upload-project] video '{file.filename}' failed: {video_err}")
                failed_files.append(file.filename)

        pdf_url = None
        pdf_file = request.files.get("pdf")
        if pdf_file and pdf_file.filename:
            pdf_url = upload_pdf_to_supabase(pdf_file, folder="projects")

        starting_price = f("startingPrice", "")

        project_data = {
            "kind": "project",
            "name": f("name", ""),
            "location": f("location", ""),
            "propertyType": f("propertyType", ""),
            "category": f("propertyType", ""),
            "possession": f("possession", ""),
            "configuration": f("configuration", ""),
            "startingPrice": starting_price,
            "budget": starting_price,
            "description": f("description", ""),
            "videoLinks": video_links,
            "img": media_urls[0] if media_urls else None,
            "mediaUrl": media_urls[0] if media_urls else None,
            "bannerUrl": banner_url or (media_urls[0] if media_urls else None),
            "mediaUrls": media_urls,
            "mediaPublicIds": media_public_ids,
            "type": "video" if (has_video and not has_image) else "image",
            "pdfUrl": pdf_url,
            "uniqueId": generate_unique_id(),
            "status": "approved",
            "ownerNumber": session.get("employee_number"),
            "ownerName": session.get("employee_name"),
            "ownerRole": session.get("role"),
            "createdAt": datetime.utcnow()
        }

        embed_and_attach(project_data)
        result = projects_collection.insert_one(project_data)
        invalidate_cache("inventory_dashboard_stats")
        return jsonify({
            "status": "success",
            "id": str(result.inserted_id),
            "failedFiles": failed_files
        }), 201

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


#embddings

# =============================
# VECTOR SEARCH / RAG - INVENTORY EMBEDDINGS
# =============================
MISTRAL_EMBED_URL = "https://api.mistral.ai/v1/embeddings"
VECTOR_INDEX_NAME = "searchdata"  # must match your Atlas Search index name exactly


def get_embedding(text: str):
    """Returns a 1024-dim embedding vector for the given text using Mistral, or None on failure."""
    if not text or not text.strip():
        return None
    if not MISTRAL_API_KEY:
        print("[embed] MISTRAL_API_KEY not configured")
        return None
    try:
        resp = requests.post(
            MISTRAL_EMBED_URL,
            headers={
                "Authorization": f"Bearer {MISTRAL_API_KEY}",
                "Content-Type": "application/json"
            },
            json={"model": "mistral-embed", "input": [text[:8000]]},
            timeout=30
        )
        resp.raise_for_status()
        return resp.json()["data"][0]["embedding"]
    except Exception as e:
        print(f"[embed] failed: {e}")
        return None


def build_inventory_embedding_text(doc: dict) -> str:
    """Flattens the fields that matter for semantic search into one text blob."""
    parts = [
        doc.get("name", ""),
        doc.get("category", ""),
        doc.get("propertyType", ""),
        doc.get("location", ""),
        doc.get("configuration", ""),
        doc.get("furnishing", ""),
        doc.get("possession", ""),
        doc.get("dealType", ""),
        doc.get("listingBasis", ""),
        doc.get("budget", "") or doc.get("startingPrice", ""),
        doc.get("description", ""),
        doc.get("quickNotes", "")
    ]
    return " | ".join(str(p) for p in parts if p)


def embed_and_attach(doc: dict) -> dict:
    """Mutates doc in-place, adding an 'embedding' field. Returns doc for convenience."""
    text = build_inventory_embedding_text(doc)
    doc["embedding"] = get_embedding(text)
    return doc


def backfill_inventory_embeddings():
    """ONE-TIME: embeds every existing project/inventory doc missing an 'embedding' field."""
    docs = list(projects_collection.find({"embedding": {"$exists": False}}))
    print(f"[backfill] {len(docs)} docs need embedding")
    for d in docs:
        text = build_inventory_embedding_text(d)
        emb = get_embedding(text)
        if emb:
            projects_collection.update_one({"_id": d["_id"]}, {"$set": {"embedding": emb}})
            print(f"[backfill] embedded: {d.get('name')}")
        else:
            print(f"[backfill] SKIPPED (no text/embed failed): {d.get('name')}")


@app.route("/api/ai/backfill-embeddings", methods=["POST"])
def api_backfill_embeddings():
    """Manually trigger backfill via HTTP instead of editing __main__. Admin only."""
    if session.get("role") != "admin":
        return jsonify({"success": False, "message": "Admin only"}), 403
    try:
        backfill_inventory_embeddings()
        return jsonify({"success": True, "message": "Backfill complete, check server logs"}), 200
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/ai/search-inventory", methods=["POST"])
def search_inventory():
    """
    RAG search endpoint for the WhatsApp AI bot (n8n etc.) to call.
    Body: { "query": "3bhk under 80 lakhs in Sidon", "limit": 5,
            "propertyType": "Apartment", "dealType": "For Sale" }  <- filters optional
    """
    try:
        data = request.json or {}
        query_text = (data.get("query") or "").strip()
        limit = int(data.get("limit", 5))

        if not query_text:
            return jsonify({"success": False, "message": "query is required"}), 400

        query_vector = get_embedding(query_text)
        if not query_vector:
            return jsonify({"success": False, "message": "Could not embed query"}), 500

        match_filter = {"status": "approved"}
        if data.get("propertyType"):
            match_filter["category"] = data["propertyType"]
        if data.get("dealType"):
            match_filter["dealType"] = data["dealType"]

        pipeline = [
            {
                "$vectorSearch": {
                    "index": VECTOR_INDEX_NAME,
                    "path": "embedding",
                    "queryVector": query_vector,
                    "numCandidates": 100,
                    "limit": limit,
                    "filter": match_filter
                }
            },
            {
                "$project": {
                    "embedding": 0,
                    "score": {"$meta": "vectorSearchScore"}
                }
            }
        ]

        results = list(projects_collection.aggregate(pipeline))
        return jsonify({"success": True, "data": [serialize_doc(r) for r in results]}), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

# =============================
# CAMPAIGN BUILDER (OpenAI-backed, no Mongo)
# =============================
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
OPENROUTER_API_URL = "https://openrouter.ai/api/v1/chat/completions"


@app.route("/campaign-builder")
def campaign_builder_page():
    if not session.get("user_id"):
        return redirect("/")
    # file must live at: templates/campaign-builder.html
    return render_template("campaign-builder.html")


@app.route("/api/ai/campaign", methods=["POST"])
def ai_campaign():
    if not session.get("user_id"):
        return jsonify({"success": False, "message": "Login required"}), 401

    if not OPENROUTER_API_KEY:
        return jsonify({"success": False, "message": "OPENROUTER_API_KEY not set in .env"}), 500

    resp = None
    try:
        data = request.json or {}
        system_prompt = data.get("system", "")
        user_prompt = data.get("prompt", "")
        max_tokens = int(data.get("max_tokens", 4000))
        temperature = float(data.get("temperature", 0.7))
        image = data.get("image")  # optional: {"media_type": "...", "data": "<base64>"}

        if not user_prompt:
            return jsonify({"success": False, "message": "prompt is required"}), 400

        user_content = []
        if image and image.get("data"):
            user_content.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:{image.get('media_type', 'image/png')};base64,{image['data']}"
                }
            })
        user_content.append({"type": "text", "text": user_prompt})

        payload = {
            "model": data.get("model") or "openai/gpt-4o-mini",
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content}
            ],
            "temperature": temperature,
            "max_tokens": max_tokens
        }

        resp = requests.post(
            OPENROUTER_API_URL,
            headers={
                "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                "Content-Type": "application/json",
                "HTTP-Referer": "https://nishahomes.com",   # OpenRouter recommends/requires this
                "X-Title": "Nisha Homes Campaign Builder"
            },
            json=payload,
            timeout=90
        )
        resp.raise_for_status()
        result = resp.json()
        text = result["choices"][0]["message"]["content"]

        return jsonify({"success": True, "text": text})

    except requests.exceptions.HTTPError as e:
        try:
            err_body = resp.json() if resp is not None else {}
        except Exception:
            err_body = {}
        msg = (err_body.get("error") or {}).get("message") or str(e)
        status = resp.status_code if resp is not None else 500
        return jsonify({"success": False, "message": msg}), status
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

import uuid

# =============================
# IMPORT LEADS FROM CSV / EXCEL (with field-mapping)
# =============================

IMPORT_TEMP_DIR = os.path.join(tempfile.gettempdir(), "lead_imports")
os.makedirs(IMPORT_TEMP_DIR, exist_ok=True)

# import_id -> {"path": ..., "ext": ..., "createdAt": time.time()}
_import_cache = {}

def _cleanup_old_imports(max_age_seconds=3600):
    """Drops any parsed-but-never-committed uploads older than 1hr."""
    now = time.time()
    stale = [k for k, v in _import_cache.items() if now - v["createdAt"] > max_age_seconds]
    for k in stale:
        try:
            os.remove(_import_cache[k]["path"])
        except Exception:
            pass
        _import_cache.pop(k, None)


# The fields a lead can be mapped into. "Phone Number" is the only hard
# requirement (everything is upserted on it, same as /add-lead).
LEAD_TARGET_FIELDS = [
    {"key": "Lead Name",               "label": "Lead Name",        "required": True},
    {"key": "Phone Number",            "label": "Phone Number",     "required": True},
    {"key": "Date",                    "label": "Date (DD-MM-YYYY)","required": False},
    {"key": "Location Interested In",  "label": "Location",         "required": False},
    {"key": "Property Type",           "label": "Property Type",    "required": False},
    {"key": "Budget Range",            "label": "Budget Range",     "required": False},
    {"key": "Configuration",           "label": "Configuration",    "required": False},
    {"key": "AssignTo",                "label": "Assigned To",      "required": False},
    {"key": "LeadType",                "label": "Lead Type",        "required": False},
    {"key": "Operating City",          "label": "Operating City",   "required": False},
    {"key": "Note",                    "label": "Note / Remarks",   "required": False},
]

_MAPPING_KEYWORDS = {
    "Lead Name": ["name", "lead name", "full name", "customer name", "client name"],
    "Phone Number": ["phone", "phone number", "mobile", "contact", "contact number", "number", "whatsapp"],
    "Date": ["date", "created date", "lead date", "enquiry date"],
    "Location Interested In": ["location", "city", "area", "locality", "location interested in"],
    "Property Type": ["property type", "property", "type", "unit type"],
    "Budget Range": ["budget", "budget range", "price", "expected price"],
    "Configuration": ["configuration", "config", "bhk"],
    "AssignTo": ["assign to", "assigned to", "agent", "executive"],
    "LeadType": ["lead type", "leadtype", "category"],
    "Operating City": ["operating city", "branch city"],
    "Note": ["note", "notes", "remarks", "comment", "comments"],
}

def _auto_guess_mapping(headers):
    """Best-effort auto-suggestion so the mapping modal isn't empty by default."""
    guesses = {}
    lower_map = {str(h).strip().lower(): h for h in headers}
    for target, keywords in _MAPPING_KEYWORDS.items():
        for kw in keywords:
            if kw in lower_map:
                guesses[target] = lower_map[kw]
                break
    return guesses


def _read_import_file(path, ext):
    if ext == "csv":
        return pd.read_csv(path, dtype=str, keep_default_na=False)
    return pd.read_excel(path, dtype=str, keep_default_na=False)


@app.route("/api/import-leads/parse", methods=["POST"])
def import_leads_parse():
    """Step 1: user picks a file. We save it to a temp dir, parse headers +
    a small preview, and hand back an importId the frontend re-uses in the
    commit call — the file is NOT re-uploaded on step 2."""
    if not session.get("user_id"):
        return jsonify({"success": False, "message": "Login required"}), 401

    try:
        if "file" not in request.files:
            return jsonify({"success": False, "message": "No file uploaded"}), 400
        file = request.files["file"]
        if not file.filename:
            return jsonify({"success": False, "message": "No file selected"}), 400

        ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
        if ext not in ("csv", "xlsx", "xls"):
            return jsonify({"success": False, "message": "Only .csv, .xlsx or .xls files are supported"}), 400

        _cleanup_old_imports()

        import_id = uuid.uuid4().hex
        saved_path = os.path.join(IMPORT_TEMP_DIR, f"{import_id}.{ext}")
        file.save(saved_path)

        try:
            df = _read_import_file(saved_path, ext)
        except Exception as parse_err:
            try:
                os.remove(saved_path)
            except Exception:
                pass
            return jsonify({"success": False, "message": f"Could not read file: {parse_err}"}), 400

        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all")

        if df.empty or len(df.columns) == 0:
            os.remove(saved_path)
            return jsonify({"success": False, "message": "File has no readable data"}), 400

        headers = list(df.columns)
        preview_rows = df.head(5).fillna("").astype(str).to_dict(orient="records")

        _import_cache[import_id] = {"path": saved_path, "ext": ext, "createdAt": time.time()}

        return jsonify({
            "success": True,
            "importId": import_id,
            "fileName": file.filename,
            "headers": headers,
            "previewRows": preview_rows,
            "totalRows": int(len(df)),
            "targetFields": LEAD_TARGET_FIELDS,
            "autoMapping": _auto_guess_mapping(headers)
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/import-leads/commit", methods=["POST"])
def import_leads_commit():
    """Step 2: user has matched sheet columns -> lead fields in the popup.
    Every imported/updated lead gets the two required fields:
      - "Created At": "YYYY-MM-DD HH:MM:SS" (only set on first insert)
      - "adon_leads": 0
    Also writes "Date"/"DateObj" the same way /add-lead does, so imported
    leads show up correctly in the existing July-2026-onward lead lists.
    """
    if not session.get("user_id"):
        return jsonify({"success": False, "message": "Login required"}), 401

    try:
        data = request.json or {}
        import_id = data.get("importId")
        collection_name = "Leads"
        mapping = data.get("mapping", {})          # { targetField: sheetColumnName }
        default_lead_type = data.get("defaultLeadType", "buyer_purchase")

        if collection_name not in ("Leads", "RentalLeads", "sellingLeads", "agentLeads"):
            return jsonify({"success": False, "message": "Invalid collection"}), 400

        if not mapping.get("Phone Number"):
            return jsonify({"success": False, "message": "Phone Number must be mapped to a column"}), 400

        entry = _import_cache.get(import_id)
        if not entry or not os.path.exists(entry["path"]):
            return jsonify({"success": False, "message": "Import session expired — please re-upload the file"}), 400

        df = _read_import_file(entry["path"], entry["ext"])
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all")

        collection = db[collection_name]
        now = datetime.utcnow()
        created_at_str = now.strftime("%Y-%m-%d %H:%M:%S")   # e.g. 2026-08-03 12:12:13

        inserted, updated, skipped = 0, 0, 0
        errors = []

        for idx, row in df.iterrows():
            try:
                phone_col = mapping.get("Phone Number")
                raw_phone = str(row.get(phone_col, "")).strip()
                phone_clean = normalize_number(raw_phone)
                if not phone_clean:
                    skipped += 1
                    continue

                doc = {}
                for target_field, sheet_col in mapping.items():
                    if not sheet_col:
                        continue
                    val = row.get(sheet_col, "")
                    val = "" if val is None else str(val).strip()
                    if val:
                        doc[target_field] = val

                doc["Phone Number"] = phone_clean

                lead_type = str(doc.get("LeadType", "")).strip()
                if lead_type not in VALID_LEAD_TYPES:
                    lead_type = default_lead_type if default_lead_type in VALID_LEAD_TYPES else "buyer_purchase"
                doc["LeadType"] = lead_type

                if not str(doc.get("Date", "")).strip():
                    doc["Date"] = now.strftime("%d-%m-%Y")

                doc["DateObj"] = now
                doc["adon_leads"] = 0   # NEW required field on every imported lead

                existing = collection.find_one({"Phone Number": phone_clean})
                if existing:
                    # Keep the original Created At — never overwrite it on re-import
                    collection.update_one({"_id": existing["_id"]}, {"$set": doc})
                    updated += 1
                else:
                    doc["Created At"] = created_at_str   # NEW required field, set once
                    collection.insert_one(doc)
                    inserted += 1

            except Exception as row_err:
                skipped += 1
                errors.append(f"Row {idx + 2}: {row_err}")
                continue

        try:
            os.remove(entry["path"])
        except Exception:
            pass
        _import_cache.pop(import_id, None)

        return jsonify({
            "success": True,
            "totalRows": int(len(df)),
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:20]
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/import-leads/cancel", methods=["POST"])
def import_leads_cancel():
    """Cleans up the temp file if the user closes the mapping popup without importing."""
    data = request.json or {}
    entry = _import_cache.pop(data.get("importId"), None)
    if entry:
        try:
            os.remove(entry["path"])
        except Exception:
            pass
    return jsonify({"success": True})


@app.route("/import-leads")
def import_leads_page():
    if not session.get("user_id"):
        return redirect("/")
    return render_template("import_leads.html")

    import sys
    #remove_assign_to_from_leads()

    if len(sys.argv) > 1 and sys.argv[1] == "resize-banners":
        # Terminal-only migration — see run_resize_banners_to_square() above.
        run_resize_banners_to_square()
        sys.exit(0)
        
        
        

# ---------------------------------------------------------------------
# 1) PERIOD RANGE HELPER
# Same shape as get_dashboard_period_range(), plus "this_week".
# Returns (start_datetime, end_datetime) or (None, None) for "lifetime".
# ---------------------------------------------------------------------
def get_team_status_period_range(period):
    now = datetime.utcnow()
 
    if period == "today":
        start = datetime(now.year, now.month, now.day)
        return start, now
 
    if period == "this_week":
        start_date = now - timedelta(days=now.weekday())  # Monday
        start = datetime(start_date.year, start_date.month, start_date.day)
        return start, now
 
    if period == "this_month":
        start = datetime(now.year, now.month, 1)
        return start, now
 
    if period == "last_month":
        first_of_this_month = datetime(now.year, now.month, 1)
        last_month_end = first_of_this_month - timedelta(seconds=1)
        start = datetime(last_month_end.year, last_month_end.month, 1)
        return start, last_month_end
 
    if period == "last_3_months":
        month = now.month - 2
        year = now.year
        while month <= 0:
            month += 12
            year -= 1
        start = datetime(year, month, 1)
        return start, now
 
    return None, None  # "lifetime"
 
 
# ---------------------------------------------------------------------
# 2) PAGE ROUTE
# ---------------------------------------------------------------------
@app.route("/team-status")
def team_status_page():
    if not session.get("user_id") or session.get("role") not in ("admin", "emp"):
        return redirect("/")
 
    return render_template(
        "team_status.html",
        employee_name=session.get("employee_name"),
        employee_number=session.get("employee_number"),
        role=session.get("role")
    )
 
 
# ---------------------------------------------------------------------
# 3) MEMBERS LIST — powers the <select> dropdown
# ---------------------------------------------------------------------
@app.route("/api/team-status/members", methods=["GET"])
def team_status_members():
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False, "message": "Staff only"}), 403
 
    try:
        members = list(db["teamAssign"].find(
            {"roll": {"$in": ["admin", "emp"]}},
            {"Employee name": 1, "Employee number": 1, "roll": 1, "Active": 1}
        ).sort("Employee name", 1))
 
        data = [{
            "name": m.get("Employee name", "Unknown"),
            "number": m.get("Employee number"),
            "role": (m.get("roll") or "").strip().lower(),
            "active": m.get("Active", True)
        } for m in members if m.get("Employee number") is not None]
 
        return jsonify({"success": True, "data": data}), 200
 
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500
 
 
# ---------------------------------------------------------------------
# 4) OVERVIEW — the main stats endpoint
# GET /api/team-status/overview?number=<Employee number>&period=<today|this_week|this_month|last_month|last_3_months|lifetime>
# ---------------------------------------------------------------------
@app.route("/api/team-status/overview", methods=["GET"])
def team_status_overview():
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False, "message": "Staff only"}), 403
 
    try:
        raw_number = request.args.get("number")
        period = request.args.get("period", "today")
 
        if not raw_number:
            return jsonify({"success": False, "message": "number is required"}), 400
 
        try:
            employee_number = int(str(raw_number).strip())
        except ValueError:
            return jsonify({"success": False, "message": "Invalid employee number"}), 400
 
        member = db["teamAssign"].find_one({"Employee number": employee_number})
        if not member:
            return jsonify({"success": False, "message": "Team member not found"}), 404
 
        employee_name = member.get("Employee name", "Unknown")
        start, end = get_team_status_period_range(period)
 
        def _clean(v, default="-"):
            if v is None or v == "":
                return default
            if isinstance(v, float) and math.isnan(v):
                return default
            return v
 
        # ---------------- CALLS + FOLLOW-UPS + INTENT ----------------
        # Pulled from callLogs (every call attempt this employee logged),
        # then deduped to the LATEST full log per lead number so a lead
        # called 5 times only counts once toward Hot/Warm/Cold/Followups.
        # "Call attempt only" pings (the plain Call button, no form) still
        # count toward Total Calls but never overwrite a real log's status.
        call_query = {"CalledBy": employee_name}
        if start is not None:
            call_query["CreatedAt"] = {"$gte": start, "$lte": end}
 
        total_calls = 0
        latest_by_number = {}
 
        for log in call_logs_collection.find(call_query).sort("CreatedAt", 1):
            total_calls += 1
            if log.get("CallAttemptOnly"):
                continue
            num = log.get("Number")
            if num:
                latest_by_number[num] = log
 
        def bucket_for(level):
            lvl = (level or "").strip().lower()
            if lvl in ("very high", "high"):
                return "Hot"
            if lvl == "medium":
                return "Warm"
            if lvl == "cold":
                return "Cold"
            return None
 
        hot_list, warm_list, cold_list, followup_list = [], [], [], []
 
        for num, log in latest_by_number.items():
            entry = {
                "number": num,
                "name": _clean(log.get("Name"), "Unknown"),
                "leadType": _clean(log.get("LeadType")),
                "callStatus": _clean(log.get("CallStatus")),
                "customerResponse": _clean(log.get("CustomerResponse")),
                "interestLevel": _clean(log.get("InterestLevel")),
                "callerRemarks": _clean(log.get("CallerRemarks")),
                "followupTimeline": _clean(log.get("FollowupTimeline")),
                "nextCallDate": _clean(log.get("NextCallDate")),
                "lastCallAt": _clean(log.get("CallDateTimeFormatted")),
            }
 
            bucket = bucket_for(log.get("InterestLevel"))
            if bucket == "Hot":
                hot_list.append(entry)
            elif bucket == "Warm":
                warm_list.append(entry)
            elif bucket == "Cold":
                cold_list.append(entry)
 
            if (log.get("FollowupTimeline") or "").strip() or (log.get("NextCallDate") or "").strip():
                followup_list.append(entry)
 
        # sort newest-first for the popup lists
        for lst in (hot_list, warm_list, cold_list, followup_list):
            lst.sort(key=lambda e: e.get("lastCallAt") or "", reverse=True)
 
        # ---------------- INVENTORY / PROJECTS ADDED ----------------
        proj_query = {"ownerNumber": employee_number}
        if start is not None:
            proj_query["createdAt"] = {"$gte": start, "$lte": end}
 
        owned_docs = list(projects_collection.find(proj_query, {
            "kind": 1, "status": 1, "name": 1, "location": 1, "createdAt": 1,
            "budget": 1, "startingPrice": 1, "category": 1, "propertyType": 1,
            "uniqueId": 1
        }))
 
        inventory_items, project_items = [], []
        inventory_pending = inventory_approved = 0
        project_pending = project_approved = 0
 
        for d in owned_docs:
            item = {
                "id": str(d["_id"]),
                "name": _clean(d.get("name")),
                "location": _clean(d.get("location")),
                "status": _clean(d.get("status"), "pending"),
                "budget": _clean(d.get("budget") or d.get("startingPrice")),
                "propertyType": _clean(d.get("category") or d.get("propertyType")),
                "createdAt": format_ist(d.get("createdAt")) if isinstance(d.get("createdAt"), datetime) else "-",
                "uniqueId": d.get("uniqueId", "")
            }
            if d.get("kind") == "project":
                project_items.append(item)
                if d.get("status") == "approved":
                    project_approved += 1
                else:
                    project_pending += 1
            else:
                inventory_items.append(item)
                if d.get("status") == "approved":
                    inventory_approved += 1
                else:
                    inventory_pending += 1
 
        for lst in (inventory_items, project_items):
            lst.sort(key=lambda e: e.get("createdAt") or "", reverse=True)
 
        return jsonify({
            "success": True,
            "period": period,
            "employee": {
                "name": employee_name,
                "number": employee_number,
                "role": (member.get("roll") or "").strip().lower(),
                "active": member.get("Active", True)
            },
            "calls": {
                "totalCalls": total_calls,
                "uniqueLeadsCalled": len(latest_by_number)
            },
            "followups": {
                "count": len(followup_list),
                "list": followup_list
            },
            "intent": {
                "hotCount": len(hot_list),
                "warmCount": len(warm_list),
                "coldCount": len(cold_list),
                "hotList": hot_list,
                "warmList": warm_list,
                "coldList": cold_list
            },
            "inventory": {
                "totalAdded": len(inventory_items) + len(project_items),
                "inventoryCount": len(inventory_items),
                "projectCount": len(project_items),
                "inventoryApproved": inventory_approved,
                "inventoryPending": inventory_pending,
                "projectApproved": project_approved,
                "projectPending": project_pending,
                "inventoryList": inventory_items,
                "projectList": project_items
            }
        }), 200
 
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500
         


# =====================================================================
# AUTOMATIC WHATSAPP FOLLOW-UP SYSTEM
# =====================================================================
 
FOLLOWUP_COLLECTIONS = {"Leads": "buying", "RentalLeads": "rental"}   # agent/selling intentionally excluded
FOLLOWUP_LEAD_TYPES_EXCLUDED = {"agent", "selling"}
 
FOLLOWUP_WINDOW_DAYS = 7            # only leads created in the last 7 days are ever followed up
FOLLOWUP_MIN_GAP_MINUTES = 5        # don't message if the customer texted in the last 5 minutes
FOLLOWUP_MAX_GAP_DAYS = 7           # don't message if it's been over a week since they last texted
FOLLOWUP_MAX_ATTEMPTS = 3           # no automatic followups after the 3rd — becomes "manual required"
 
FOLLOWUP_HOURLY_LIMIT = 35          # leads processed per batch
 
FOLLOWUP_SEND_WAIT_MIN_MINUTES = 1  # random wait between individual sends: 1-3 min,
FOLLOWUP_SEND_WAIT_MAX_MINUTES = 3  # can land on a decimal (1.2, 2.6, ...), not just whole minutes
 
FOLLOWUP_BATCH_COOLDOWN_MIN_MINUTES = 10   # after a 35-message batch, cool down 7-10 min
FOLLOWUP_BATCH_COOLDOWN_MAX_MINUTES = 20  # before starting on any leads left over
 
FOLLOWUP_BUSINESS_START_HOUR = 10   # 10:00 AM IST — no sends before this
FOLLOWUP_BUSINESS_END_HOUR = 19     # 7:00 PM IST — no sends at/after this
 
FOLLOWUP_RESCAN_INTERVAL_SEC = 3600     # re-scan for newly-eligible leads every hour (during business hours)
 
FOLLOWUP_MISTRAL_KEY = MISTRAL_API_KEY  # reuses the existing autofill key; swap to MISTRAL_API_KEY2 if you'd rather keep quotas separate
 
# Brand sign-off + CTA-word library, from your Nisha Homes message-format
# doc. Scoped to just the follow-up system so it doesn't touch the
# existing BRAND_CONTACT_NUMBER/BRAND_WEBSITE constants used by the
# image/video branding code elsewhere in this file.
FOLLOWUP_CONTACT_LINE = "📞 7303515710 / 7303755710\n🌐 www.nishahomes.com"
FOLLOWUP_CTA_LIBRARY = ["YES", "OPTIONS", "HELP", "SEARCH", "CALL", "PRICE", "LOCATION", "BUDGET"]
 
# Maps the 3 automatic attempts we currently support onto the "purpose
# per message" idea from your doc (Remind -> Understand requirement ->
# Offer help / human check-in). Attempt 4+ never happens automatically
# (that's the "manual calling" bucket), so only 1-3 are defined.
FOLLOWUP_STAGE_GUIDANCE = {
    1: (
        "This is the FIRST follow-up. Purpose: a warm welcome/acknowledgement — remind them "
        "who Nisha Homes is and that you're here to help with their property/accommodation "
        "requirement in Delhi NCR. Keep it low-pressure, not salesy."
    ),
    2: (
        "This is the SECOND follow-up. Purpose: get them talking with an easy, low-effort "
        "question about their requirement (buy / rent / sell / invest / PG, their location, "
        "budget, or timeline) — something they can answer in one word or a short reply."
    ),
    3: (
        "This is the THIRD and FINAL automatic follow-up before this lead is handed to a "
        "human caller. Purpose: a polite final check-in — ask directly if they're still "
        "looking, and mention that a simple one-word reply is completely fine. Don't sound "
        "desperate or pushy."
    ),
}
 
followup_logs_collection = db["followupLogs"]
system_locks_collection = db["systemLocks"]
 
 
def parse_last_user_msg_at(s):
    """
    Parses an ISO-8601 timestamp with a UTC offset, e.g.
    '2026-07-11T20:45:15.635+05:30', and returns a naive UTC datetime
    (matching the naive-UTC convention used everywhere else in this file,
    e.g. datetime.utcnow() / 'Created At'). Returns None if unparseable.
    """
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(str(s).strip())
        offset = dt.utcoffset()
        if offset is not None:
            dt = (dt - offset).replace(tzinfo=None)
        return dt
    except Exception:
        return None
 
 
def ist_date_str(dt_utc):
    """Same IST-shift convention as format_ist(), but date-only — used to
    enforce 'max one followup per lead per day'."""
    if not isinstance(dt_utc, datetime):
        return None
    ist = dt_utc + timedelta(hours=5, minutes=30)
    return ist.strftime("%Y-%m-%d")
 
 
def get_ist_now():
    """Current time shifted to IST, same convention as format_ist()."""
    return datetime.utcnow() + timedelta(hours=5, minutes=30)
 
 
def is_within_business_hours():
    """True between 10:00 AM and 7:00 PM IST (7pm itself is already closed)."""
    hour = get_ist_now().hour
    return FOLLOWUP_BUSINESS_START_HOUR <= hour < FOLLOWUP_BUSINESS_END_HOUR
 
 
def seconds_until_next_business_window():
    """How long to sleep before the next 10am-IST window opens."""
    ist_now = get_ist_now()
    today_start = ist_now.replace(hour=FOLLOWUP_BUSINESS_START_HOUR, minute=0, second=0, microsecond=0)
    target = today_start if ist_now.hour < FOLLOWUP_BUSINESS_START_HOUR else today_start + timedelta(days=1)
    return max(60, int((target - ist_now).total_seconds()))
 
 
# ---------------------------------------------------------------------
# MISTRAL — SHORT NATURAL FOLLOW-UP MESSAGE GENERATION
# ---------------------------------------------------------------------
def generate_followup_message(lead_ctx, end_data, call_logs, attempt_number, previous_messages=None):
    """
    Uses Mistral to write a short, natural WhatsApp follow-up for this
    lead, following the Nisha Homes message format: warm/conversational
    tone, one clear reply-word CTA per message (rotated so it's never the
    same word/phrasing twice to the same lead), a short brand sign-off,
    and a different "purpose" per attempt (see FOLLOWUP_STAGE_GUIDANCE).
    Returns plain text, or None on failure.
    """
    if not FOLLOWUP_MISTRAL_KEY:
        print("[followup] Mistral key not configured — skipping message generation")
        return None
 
    call_summary = []
    for c in (call_logs or [])[:5]:
        call_summary.append({
            "date": c.get("CallDateTimeFormatted"),
            "status": c.get("CallStatus"),
            "response": c.get("CustomerResponse"),
            "interest": c.get("InterestLevel"),
            "objection": c.get("Objection"),
            "remarks": c.get("CallerRemarks"),
        })
 
    previous_texts = [
        h.get("message", "") for h in (previous_messages or [])
        if h.get("message") and h.get("status") == "sent"
    ]
 
    context = {
        "customerName": lead_ctx.get("name"),
        "location": lead_ctx.get("location"),
        "propertyType": lead_ctx.get("propertyType"),
        "budget": lead_ctx.get("budget"),
        "configuration": lead_ctx.get("configuration"),
        "followupAttemptNumber": attempt_number,
        "lastCallStatus": (end_data or {}).get("Call Status", ""),
        "lastInterestLevel": (end_data or {}).get("Interest Level", ""),
        "lastCallerRemarks": (end_data or {}).get("Caller Remarks", ""),
        "callHistory": call_summary,
        "previouslySentToThisLead": previous_texts,
    }
 
    stage_guidance = FOLLOWUP_STAGE_GUIDANCE.get(attempt_number, FOLLOWUP_STAGE_GUIDANCE[3])
 
    system_prompt = (
        "You are 'Nisha' from Nisha Homes, a real-estate advisory business in Delhi NCR, "
        "writing a short WhatsApp follow-up to a property/accommodation lead. Write in first "
        "person as Nisha, in a warm, natural, conversational WhatsApp tone — short sentences, "
        "not a formal email, light and occasional emoji use (not excessive). Keep the whole "
        "message concise: about 2-4 short lines plus the sign-off.\n\n"
        f"{stage_guidance}\n\n"
        "Use the lead's own details and call history below (if any) so the message feels "
        "personal and relevant — reference their location, property type or budget if known, "
        "and briefly acknowledge anything specific from a past call (an objection, stated "
        "interest, etc). Do not invent details that aren't given.\n\n"
        "End the message with exactly ONE simple, easy-to-answer call-to-action asking them "
        "to reply with a single word — choose ONE word from this list, and pick one that is "
        f"different from anything already sent to this lead: {', '.join(FOLLOWUP_CTA_LIBRARY)}.\n\n"
        "Close with a short brand sign-off on its own line(s), reusing exactly this contact "
        f"block (you may shorten 'Nisha Homes' branding around it, but keep the numbers and "
        f"site as-is):\n{FOLLOWUP_CONTACT_LINE}\n\n"
        "CRITICAL — never repeat a message. If 'previouslySentToThisLead' below is non-empty, "
        "your new message must read differently from every message listed there, and must not "
        "reuse the same CTA word or phrases like 'just checking in'.\n\n"
        "Reply with ONLY the WhatsApp message text — no quotes, no preamble, no explanation."
    )
 
    payload = {
        "model": "mistral-large-latest",
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": json.dumps(context)},
        ],
        "temperature": 0.6,
    }
 
    try:
        resp = requests.post(
            MISTRAL_API_URL,
            headers={
                "Authorization": f"Bearer {FOLLOWUP_MISTRAL_KEY}",
                "Content-Type": "application/json",
            },
            json=payload,
            timeout=30,
        )
        resp.raise_for_status()
        text = resp.json()["choices"][0]["message"]["content"].strip()
        return text.strip('"').strip() or None
    except Exception as e:
        print(f"[followup] message generation failed: {e}")
        return None
 
 
# ---------------------------------------------------------------------
# ELIGIBILITY
# ---------------------------------------------------------------------
def get_eligible_followup_leads(limit=None):
    """
    Scans Leads + RentalLeads (buyer/rental only — agent/selling are
    excluded per spec) for leads that are:
      - created within the last FOLLOWUP_WINDOW_DAYS days
      - not already at FOLLOWUP_MAX_ATTEMPTS followups
      - not already followed up today (IST)
      - "quiet enough" since the customer's last message (>= 5 min,
        <= 7 days — using lastUserMsgAt if present, else Created At)
    Returns a list of lightweight lead-context dicts, oldest/least-followed
    first.
    """
    now = datetime.utcnow()
    window_start = now - timedelta(days=FOLLOWUP_WINDOW_DAYS)
    today_ist = ist_date_str(now)
 
    # cheap DB-side filter: only docs that have a Created At AND aren't
    # already fully-followed-up — full date parsing still happens in
    # Python below since Created At is stored as a string, same
    # limitation as the rest of this app's dashboard endpoints.
    base_query = {
        "Created At": {"$exists": True},
        "$or": [
            {"followup_send": {"$exists": False}},
            {"followup_send": {"$lt": FOLLOWUP_MAX_ATTEMPTS}},
        ],
    }
 
    candidates = []
    for coll_name in FOLLOWUP_COLLECTIONS:
        for lead in db[coll_name].find(base_query):
            created_dt = parse_created_at_str(lead.get("Created At"))
            if not created_dt or created_dt < window_start or created_dt > now:
                continue
 
            lead_type = _normalize_lead_type_field(lead.get("LeadType"))
            if lead_type in FOLLOWUP_LEAD_TYPES_EXCLUDED:
                continue
 
            followup_count = lead.get("followup_send", 0)
            try:
                followup_count = int(followup_count)
            except (TypeError, ValueError):
                followup_count = 0
            if followup_count >= FOLLOWUP_MAX_ATTEMPTS:
                continue
 
            last_followup_at = lead.get("lastFollowupAt")
            if isinstance(last_followup_at, datetime) and ist_date_str(last_followup_at) == today_ist:
                continue  # already followed up today
 
            last_msg_dt = parse_last_user_msg_at(lead.get("lastUserMsgAt")) or created_dt
            gap = now - last_msg_dt
            if gap < timedelta(minutes=FOLLOWUP_MIN_GAP_MINUTES) or gap > timedelta(days=FOLLOWUP_MAX_GAP_DAYS):
                continue
 
            phone = normalize_number(lead.get("Phone Number", ""))
            if not phone:
                continue
            if not phone.startswith("91"):
                phone = "91" + phone
 
            candidates.append({
                "leadId": lead["_id"],
                "collection": coll_name,
                "leadType": lead_type,
                "phone": phone,
                "name": lead.get("Lead Name") or lead.get("Name") or "Unknown",
                "location": lead.get("Location Interested In") or lead.get("Property Location") or "",
                "propertyType": lead.get("Property Type", ""),
                "budget": lead.get("Budget Range") or lead.get("Expected Price") or "",
                "configuration": lead.get("Configuration", ""),
                "createdAt": created_dt,
                "followupCount": followup_count,
            })
 
    candidates.sort(key=lambda c: (c["followupCount"], c["createdAt"]))
    if limit:
        return candidates[:limit]
    return candidates
 
 
# ---------------------------------------------------------------------
# SEND ONE FOLLOW-UP
# ---------------------------------------------------------------------
def process_single_followup(ctx):
    lead_id = ctx["leadId"]
    coll_name = ctx["collection"]
    phone = ctx["phone"]
    attempt_number = ctx["followupCount"] + 1
    now = datetime.utcnow()
 
    end_doc = db["endData"].find_one({"Number": phone}) or {}
    call_logs = list(call_logs_collection.find({"Number": phone}).sort("CreatedAt", -1).limit(5))
 
    # Fetch this lead's own follow-up history so Mistral never repeats a
    # message or CTA word it already sent.
    lead_doc = db[coll_name].find_one({"_id": lead_id}, {"followupHistory": 1}) or {}
    previous_messages = lead_doc.get("followupHistory", [])
 
    message = generate_followup_message(ctx, end_doc, call_logs, attempt_number, previous_messages)
    if not message:
        followup_logs_collection.insert_one({
            "leadId": str(lead_id), "collection": coll_name, "phone": phone,
            "name": ctx["name"], "attempt": attempt_number, "message": None,
            "status": "generation_failed", "createdAt": now,
        })
        print(f"[followup] message generation failed for {phone} ({ctx['name']}) — will retry next scan")
        return
 
    history_entry = {"attempt": attempt_number, "message": message, "sentAt": now}
 
    try:
        result = wp.send_whatsapp_message(phone, message)
        history_entry["status"] = "sent"
        history_entry["messageId"] = result.get("messageId")
 
        new_status = "manual_required" if attempt_number >= FOLLOWUP_MAX_ATTEMPTS else "pending"
        db[coll_name].update_one(
            {"_id": lead_id},
            {
                "$set": {
                    "followup_send": attempt_number,
                    "lastFollowupAt": now,
                    "followupStatus": new_status,
                },
                "$push": {"followupHistory": history_entry},
            },
        )
        followup_logs_collection.insert_one({
            "leadId": str(lead_id), "collection": coll_name, "phone": phone,
            "name": ctx["name"], "attempt": attempt_number, "message": message,
            "status": "sent", "messageId": result.get("messageId"), "createdAt": now,
        })
        print(f"[followup] sent attempt {attempt_number}/{FOLLOWUP_MAX_ATTEMPTS} to {phone} ({ctx['name']})")
 
    except wp.WirebaseError as e:
        history_entry["status"] = "failed"
        history_entry["error"] = str(e)
        # Don't increment followup_send on failure (so it's retried), but
        # DO stamp lastFollowupAt so a broken instance isn't hammered every
        # scan — it gets one attempt per day, same cadence as a real send.
        db[coll_name].update_one(
            {"_id": lead_id},
            {
                "$set": {"lastFollowupAt": now},
                "$push": {"followupHistory": history_entry},
            },
        )
        followup_logs_collection.insert_one({
            "leadId": str(lead_id), "collection": coll_name, "phone": phone,
            "name": ctx["name"], "attempt": attempt_number, "message": message,
            "status": "failed", "error": str(e), "createdAt": now,
        })
        print(f"[followup] FAILED attempt {attempt_number} to {phone}: {e}")
 
    invalidate_cache("followup_stats")
 
 
# ---------------------------------------------------------------------
# BATCH RUNNER — gathers eligible leads, sends in a paced/rate-limited way
# ---------------------------------------------------------------------
def run_followup_scan_and_send():
    while True:
        if not is_within_business_hours():
            print("[followup] outside business hours (10am-7pm IST) — pausing until next window")
            return
 
        eligible = get_eligible_followup_leads()
        if not eligible:
            print("[followup] no eligible leads this scan")
            return
 
        batch = eligible[:FOLLOWUP_HOURLY_LIMIT]
        print(f"[followup] processing batch of {len(batch)} lead(s) (of {len(eligible)} eligible)")
 
        for ctx in batch:
            if not is_within_business_hours():
                print("[followup] business hours ended mid-batch — pausing until next window")
                return
            try:
                process_single_followup(ctx)
            except Exception:
                import traceback
                traceback.print_exc()
 
            wait_minutes = round(random.uniform(FOLLOWUP_SEND_WAIT_MIN_MINUTES, FOLLOWUP_SEND_WAIT_MAX_MINUTES), 1)
            print(f"[followup] waiting {wait_minutes} min before the next send")
            time.sleep(wait_minutes * 60)
 
        if len(eligible) <= FOLLOWUP_HOURLY_LIMIT:
            return  # everyone eligible this scan has now been handled
 
        if not is_within_business_hours():
            print("[followup] business hours ended before next batch — pausing until next window")
            return
 
        cooldown_minutes = round(random.uniform(FOLLOWUP_BATCH_COOLDOWN_MIN_MINUTES, FOLLOWUP_BATCH_COOLDOWN_MAX_MINUTES), 1)
        print(f"[followup] batch cap ({FOLLOWUP_HOURLY_LIMIT}) reached — cooling down {cooldown_minutes} min")
        time.sleep(cooldown_minutes * 60)
 
 
# ---------------------------------------------------------------------
# MONGO-BACKED LOCK (guards against duplicate sends across multiple
# gunicorn workers/processes — still pure Flask/Mongo, no new tooling)
# ---------------------------------------------------------------------
def try_acquire_followup_lock(ttl_seconds):
    now = datetime.utcnow()
    expires = now + timedelta(seconds=ttl_seconds)
    owner = f"{socket.gethostname()}-{os.getpid()}"
 
    # Take over a lock that's missing or has expired.
    result = system_locks_collection.find_one_and_update(
        {"_id": "followup_scheduler", "lockedUntil": {"$lt": now}},
        {"$set": {"lockedUntil": expires, "lockedBy": owner, "lockedAt": now}},
    )
    if result is not None:
        return True
 
    # No stale lock found — try to create it fresh (only succeeds if no
    # doc exists yet at all).
    try:
        system_locks_collection.insert_one({
            "_id": "followup_scheduler", "lockedUntil": expires,
            "lockedBy": owner, "lockedAt": now,
        })
        return True
    except Exception:
        return False  # someone else holds a valid, non-expired lock
 
 
def followup_scheduler_loop():
    print("[followup] scheduler loop started")
    while True:
        try:
            if not is_within_business_hours():
                sleep_for = seconds_until_next_business_window()
                print(f"[followup] outside business hours (10am-7pm IST) — sleeping {sleep_for}s until next window")
                time.sleep(sleep_for)
                continue
 
            if try_acquire_followup_lock(ttl_seconds=FOLLOWUP_RESCAN_INTERVAL_SEC - 60):
                run_followup_scan_and_send()
            else:
                print("[followup] scheduler lock held by another process — skipping this cycle")
        except Exception:
            import traceback
            traceback.print_exc()
        time.sleep(FOLLOWUP_RESCAN_INTERVAL_SEC)
 
 
_followup_thread_started = False
 
 
def start_followup_scheduler():
    """Starts the background follow-up thread exactly once per process,
    and avoids a double-start under Flask's debug-mode reloader."""
    global _followup_thread_started
    if _followup_thread_started:
        return
    if os.environ.get("WERKZEUG_RUN_MAIN") == "true" or not app.debug:
        threading.Thread(target=followup_scheduler_loop, daemon=True, name="followup-scheduler").start()
        _followup_thread_started = True
        print("[followup] scheduler thread launched")
 
 
start_followup_scheduler()
 
 
# =====================================================================
# ROUTES
# =====================================================================
@app.route("/followups")
def followups_page():
    if not session.get("user_id") or session.get("role") not in ("admin", "emp"):
        return redirect("/")
    return render_template(
        "followups.html",
        employee_name=session.get("employee_name"),
        employee_number=session.get("employee_number"),
        role=session.get("role"),
    )
 
 
def _compute_followup_stats():
    now = datetime.utcnow()
    window_start = now - timedelta(days=FOLLOWUP_WINDOW_DAYS)
 
    counts = {0: 0, 1: 0, 2: 0}
    manual = 0
    query = {"Created At": {"$exists": True}}
 
    for coll_name in FOLLOWUP_COLLECTIONS:
        for lead in db[coll_name].find(query, {"followup_send": 1, "LeadType": 1, "Created At": 1}):
            lead_type = _normalize_lead_type_field(lead.get("LeadType"))
            if lead_type in FOLLOWUP_LEAD_TYPES_EXCLUDED:
                continue
            created_dt = parse_created_at_str(lead.get("Created At"))
            if not created_dt or created_dt < window_start or created_dt > now:
                continue
 
            count = lead.get("followup_send", 0)
            try:
                count = int(count)
            except (TypeError, ValueError):
                count = 0
 
            if count >= FOLLOWUP_MAX_ATTEMPTS:
                manual += 1
            else:
                counts[count] = counts.get(count, 0) + 1
 
    today_start = datetime(now.year, now.month, now.day)
    return {
        "pending": counts.get(0, 0),
        "followup1": counts.get(1, 0),
        "followup2": counts.get(2, 0),
        "manualRequired": manual,
        "sentToday": followup_logs_collection.count_documents({"status": "sent", "createdAt": {"$gte": today_start}}),
        "failedToday": followup_logs_collection.count_documents({"status": "failed", "createdAt": {"$gte": today_start}}),
        "sentTotal": followup_logs_collection.count_documents({"status": "sent"}),
        "failedTotal": followup_logs_collection.count_documents({"status": "failed"}),
    }
 
 
@app.route("/api/followup/stats", methods=["GET"])
def followup_stats():
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False, "message": "Staff only"}), 403
    data = cached("followup_stats", 20, _compute_followup_stats)
    return jsonify({"success": True, **data}), 200
 
 
@app.route("/api/followup/leads", methods=["GET"])
def followup_leads_list():
    if session.get("role") not in ("admin", "emp"):
        return jsonify({"success": False, "message": "Staff only"}), 403
 
    bucket = request.args.get("bucket", "pending")
    now = datetime.utcnow()
    window_start = now - timedelta(days=FOLLOWUP_WINDOW_DAYS)
 
    def _clean(v, default="-"):
        if v is None or v == "":
            return default
        if isinstance(v, float) and math.isnan(v):
            return default
        return v
 
    # --- activity-log buckets: sent / failed ---
    if bucket in ("sent", "failed"):
        logs = list(followup_logs_collection.find({"status": bucket}).sort("createdAt", -1).limit(300))
        data = [{
            "leadId": l.get("leadId"),
            "collection": l.get("collection"),
            "name": _clean(l.get("name")),
            "phone": l.get("phone"),
            "attempt": l.get("attempt"),
            "message": l.get("message"),
            "status": l.get("status"),
            "error": l.get("error"),
            "at": format_ist(l.get("createdAt")) if isinstance(l.get("createdAt"), datetime) else "-",
        } for l in logs]
        return jsonify({"success": True, "bucket": bucket, "count": len(data), "data": data}), 200
 
    # --- lead-stage buckets: pending / followup1 / followup2 / manual ---
    target_count = {"pending": 0, "followup1": 1, "followup2": 2}.get(bucket)
    is_manual = bucket in ("followup3", "manual")
    if target_count is None and not is_manual:
        return jsonify({"success": False, "message": "Invalid bucket"}), 400
 
    out = []
    for coll_name in FOLLOWUP_COLLECTIONS:
        for lead in db[coll_name].find():
            lead_type = _normalize_lead_type_field(lead.get("LeadType"))
            if lead_type in FOLLOWUP_LEAD_TYPES_EXCLUDED:
                continue
            created_dt = parse_created_at_str(lead.get("Created At"))
            if not created_dt or created_dt < window_start or created_dt > now:
                continue
 
            count = lead.get("followup_send", 0)
            try:
                count = int(count)
            except (TypeError, ValueError):
                count = 0
 
            if is_manual:
                if count < FOLLOWUP_MAX_ATTEMPTS:
                    continue
            else:
                if count != target_count:
                    continue
 
            phone = normalize_number(lead.get("Phone Number", ""))
            if phone and not phone.startswith("91"):
                phone = "91" + phone
 
            end_doc = (db["endData"].find_one({"Number": phone}) or {}) if phone else {}
 
            out.append({
                "id": str(lead["_id"]),
                "collection": coll_name,
                "leadType": lead_type,
                "name": _clean(lead.get("Lead Name") or lead.get("Name"), "Unknown"),
                "phone": phone or "-",
                "location": _clean(lead.get("Location Interested In") or lead.get("Property Location")),
                "propertyType": _clean(lead.get("Property Type")),
                "budget": _clean(lead.get("Budget Range") or lead.get("Expected Price")),
                "createdAt": format_ist(created_dt),
                "followupCount": count,
                "lastFollowupAt": format_ist(lead.get("lastFollowupAt")) if isinstance(lead.get("lastFollowupAt"), datetime) else "-",
                "callStatus": _clean(end_doc.get("Call Status")),
                "interestLevel": _clean(end_doc.get("Interest Level")),
                "followupHistory": [
                    {
                        "attempt": h.get("attempt"),
                        "message": h.get("message"),
                        "status": h.get("status"),
                        "sentAt": format_ist(h.get("sentAt")) if isinstance(h.get("sentAt"), datetime) else "-",
                    }
                    for h in (lead.get("followupHistory") or [])
                ],
            })
 
    out.sort(key=lambda x: x["createdAt"], reverse=True)
    return jsonify({"success": True, "bucket": bucket, "count": len(out), "data": out}), 200
 
 
@app.route("/api/followup/trigger", methods=["POST"])
def followup_trigger():
    """Admin-only manual kick, for testing/ops — runs one scan in the
    background so the request returns immediately."""
    if session.get("role") != "admin":
        return jsonify({"success": False, "message": "Admin only"}), 403
    threading.Thread(target=run_followup_scan_and_send, daemon=True, name="followup-manual-trigger").start()
    return jsonify({"success": True, "message": "Follow-up scan started in the background"}), 200
 

        
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000)